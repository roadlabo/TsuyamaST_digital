from __future__ import annotations

# =========================================
# Imports
# =========================================
import argparse
import csv
import json
import logging
import os
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

os.environ.setdefault("OPENCV_FFMPEG_CAPTURE_OPTIONS", "rtsp_transport;tcp|fflags;nobuffer|flags;low_delay")
os.environ.setdefault("OPENCV_VIDEOIO_DEBUG", "0")
os.environ.setdefault("OPENCV_LOG_LEVEL", "ERROR")

import cv2
import numpy as np
import pandas as pd
import torch
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from PyQt6 import QtCore, QtGui, QtWidgets
from ultralytics import YOLO

APP_DIR = Path(__file__).resolve().parents[1]
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from importlib import import_module
_congestion_common = import_module("10_common.congestion_common")
CongestionSmoother = _congestion_common.CongestionSmoother
level_style = _congestion_common.level_style


def setup_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / "ai_monitor.log"
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    root = logging.getLogger()
    if root.handlers:
        return
    root.setLevel(logging.INFO)
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    root.addHandler(stream_handler)
    root.addHandler(file_handler)


def _sleep_backoff(i: int, cap: float = 0.5) -> None:
    time.sleep(min(cap, 0.05 * (2 ** i)))


def safe_replace(tmp_path: Path, dst_path: Path, retries: int = 10) -> None:
    last_exc: Exception | None = None
    for i in range(max(1, int(retries))):
        try:
            os.replace(tmp_path, dst_path)
            return
        except (PermissionError, OSError) as exc:
            last_exc = exc
            logging.warning("[WARN] ai_status write retry %d/%d failed: %s", i + 1, retries, exc)
            _sleep_backoff(i)
    try:
        tmp_path.replace(dst_path)
        return
    except Exception as exc:
        raise RuntimeError(f"safe_replace failed: {tmp_path} -> {dst_path} ({last_exc or exc})") from exc


def write_json_atomic(path: Path, payload: dict[str, Any], retries: int = 10) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    with tmp_path.open("w", encoding="utf-8") as fh:
        fh.write(text)
        fh.flush()
        os.fsync(fh.fileno())
    try:
        safe_replace(tmp_path, path, retries=retries)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except Exception:
                pass


def safe_read_json(path: Path, default: Any, retries: int = 3) -> Any:
    if not path.exists():
        return default
    last_exc: Exception | None = None
    for i in range(max(1, int(retries))):
        try:
            with path.open("r", encoding="utf-8") as fh:
                return json.load(fh)
        except FileNotFoundError:
            return default
        except (json.JSONDecodeError, PermissionError, OSError) as exc:
            last_exc = exc
            _sleep_backoff(i, cap=0.2)
        except Exception as exc:
            last_exc = exc
            break
    logging.warning("safe_read_json failed: %s (%s)", path, last_exc)
    return default

# =========================================
# Constants / CLASS_MAP
# =========================================
CLASS_MAP = {
    0: "人",
    1: "自転車",
    2: "車",
    3: "オートバイ",
    4: "飛行機",
    5: "バス",
    6: "電車",
    7: "トラック",
}

# =========================================
# Default Settings
# =========================================
DEFAULT_SYSTEM_CONFIG: dict[str, Any] = {
    "model_path": "yolo11m.pt",
    "device_preference": "auto",
    "metrics_save_interval_sec": 5,
    "ui_refresh_interval_ms": 500,
    "ai_status_json_path": "app/11_config/ai_status.json",
    "status_update_interval_sec": 10,
    "output_root": "app/04_ai_monitor/data",
    "display_update_interval_ms": 200,
    "graph_update_interval_sec": 10,
    "level2_threshold": 8.0,
    "level3_threshold": 12.0,
    "level4_threshold": 16.0,
    "congestion_smoothing_window": 6,
}

DEFAULT_CAMERA_SETTINGS: dict[str, Any] = {
    "cameras": [
        {
            "camera_id": 1,
            "camera_name": "Camera1",
            "stream_url": "0",
            "enabled": True,
            "line_start": [100, 300],
            "line_end": [1000, 300],
            "exclude_polygon": [],
            "congestion_threshold": 5,
            "long_stay_minutes": 15,
            "long_stay_trigger_count": 1,
            "stay_zone_polygon": [],
            "reconnect_sec": 3,
            "tracking_method": "bytetrack",
            "yolo_model": "yolo11m.pt",
            "confidence_threshold": 0.25,
            "iou_threshold": 0.5,
            "frame_skip": 1,
            "imgsz": 512,
            "target_classes": [2, 3, 5, 7],
            "bt_track_high_thresh": 0.3,
            "bt_track_low_thresh": 0.1,
            "bt_match_thresh": 0.8,
            "bt_track_buffer": 30,
            "crossing_judgment_pattern": "line_cross",
            "distance_threshold": 25.0,
            "congestion_calculation_interval": 3,
            "enable_congestion": True,
            "line_direction_mode": "line_vector",
            "display_scale": 0.8,
        },
        {
            "camera_id": 2,
            "camera_name": "Camera2",
            "stream_url": "1",
            "enabled": True,
            "line_start": [100, 300],
            "line_end": [1000, 300],
            "exclude_polygon": [],
            "congestion_threshold": 5,
            "long_stay_minutes": 15,
            "long_stay_trigger_count": 1,
            "stay_zone_polygon": [],
            "reconnect_sec": 3,
            "tracking_method": "bytetrack",
            "yolo_model": "yolo11m.pt",
            "confidence_threshold": 0.25,
            "iou_threshold": 0.5,
            "frame_skip": 1,
            "imgsz": 640,
            "target_classes": [2, 3, 5, 7],
            "bt_track_high_thresh": 0.3,
            "bt_track_low_thresh": 0.1,
            "bt_match_thresh": 0.8,
            "bt_track_buffer": 30,
            "crossing_judgment_pattern": "line_cross",
            "distance_threshold": 25.0,
            "congestion_calculation_interval": 3,
            "enable_congestion": True,
            "line_direction_mode": "line_vector",
            "display_scale": 0.8,
        },
        {
            "camera_id": 3,
            "camera_name": "Camera3",
            "stream_url": "2",
            "enabled": True,
            "line_start": [100, 300],
            "line_end": [1000, 300],
            "exclude_polygon": [],
            "congestion_threshold": 5,
            "long_stay_minutes": 15,
            "long_stay_trigger_count": 1,
            "stay_zone_polygon": [],
            "reconnect_sec": 3,
            "tracking_method": "bytetrack",
            "yolo_model": "yolo11m.pt",
            "confidence_threshold": 0.25,
            "iou_threshold": 0.5,
            "frame_skip": 1,
            "imgsz": 512,
            "target_classes": [2, 3, 5, 7],
            "bt_track_high_thresh": 0.3,
            "bt_track_low_thresh": 0.1,
            "bt_match_thresh": 0.8,
            "bt_track_buffer": 30,
            "crossing_judgment_pattern": "line_cross",
            "distance_threshold": 25.0,
            "congestion_calculation_interval": 3,
            "enable_congestion": True,
            "line_direction_mode": "line_vector",
            "display_scale": 0.8,
        },
    ]
}

KEEP_CAMERA_KEYS = [
    "camera_id",
    "camera_name",
    "stream_url",
    "enabled",
    "line_start",
    "line_end",
    "exclude_polygon",
    "stay_zone_polygon",
    "congestion_threshold",
    "long_stay_minutes",
]

# =========================================
# Dataclasses
# =========================================
@dataclass
class AppConfig:
    root_dir: Path
    system_config_path: Path
    camera_settings_path: Path
    system: dict[str, Any]
    cameras: list[dict[str, Any]]


@dataclass
class CongestionState:
    frame_motion_scores: list[float] = field(default_factory=list)
    smoothed_motion_scores: list[float] = field(default_factory=list)
    frame_time_stamps: list[datetime] = field(default_factory=list)
    frame_cumulative_motion_score: float = 0.0
    window_frame_count: int = 0
    current_congestion_index: float = 0.0
    current_smoothed_index: float = 0.0
    window_start: datetime | None = None
    previous_positions: dict[int, tuple[float, float]] = field(default_factory=dict)
    last_seen_at: dict[int, datetime] = field(default_factory=dict)


@dataclass
class CounterState:
    counted_track_ids: set[int] = field(default_factory=set)
    previous_side: dict[int, float] = field(default_factory=dict)
    pass_bins_ltor: list[int] = field(default_factory=lambda: [0] * 144)
    pass_bins_rtol: list[int] = field(default_factory=lambda: [0] * 144)


@dataclass
class TrackState:
    first_seen: dict[int, datetime] = field(default_factory=dict)
    long_stay_emitted: set[int] = field(default_factory=set)


@dataclass
class WakimuraAlphaState:
    exit_timestamps: list[datetime] = field(default_factory=list)
    high_load_active: bool = False
    high_load_session_stays: list[float] = field(default_factory=list)
    high_load_alpha: float | None = None
    current_alpha: float = 0.0
    current_alpha_window: float = 0.0
    current_n_out: int = 0
    current_avg_stay_sec: float = 0.0


# =========================================
# Config Manager
# =========================================
class ConfigManager:
    def __init__(self, root_dir: Path):
        self.root_dir = root_dir
        self.config_dir = root_dir / "config"
        self.config_dir.mkdir(parents=True, exist_ok=True)
        self.system_config_path = self.config_dir / "system_config.json"
        self.camera_settings_path = self.config_dir / "camera_settings.json"

    def ensure_defaults(self) -> None:
        if not self.system_config_path.exists():
            self._atomic_write_json(self.system_config_path, DEFAULT_SYSTEM_CONFIG)
        if not self.camera_settings_path.exists():
            slim_defaults = [self._to_slim_camera_settings(cam) for cam in DEFAULT_CAMERA_SETTINGS.get("cameras", [])]
            self._atomic_write_json(self.camera_settings_path, {"cameras": slim_defaults})

    def load(self) -> AppConfig:
        self.ensure_defaults()
        system = safe_read_json(self.system_config_path, DEFAULT_SYSTEM_CONFIG.copy(), retries=3)
        camera_dict = safe_read_json(self.camera_settings_path, {"cameras": []}, retries=3)
        loaded_cameras = camera_dict.get("cameras", [])
        default_map = {int(c["camera_id"]): dict(c) for c in DEFAULT_CAMERA_SETTINGS.get("cameras", [])}
        cameras: list[dict[str, Any]] = []
        for cam in loaded_cameras:
            raw = dict(cam)
            if "line_points" in raw and "line_start" not in raw and "line_end" not in raw:
                points = raw.get("line_points") or []
                if isinstance(points, list) and len(points) >= 2:
                    raw["line_start"] = points[0]
                    raw["line_end"] = points[1]
            cid = int(raw.get("camera_id", -1))
            base = dict(default_map.get(cid, {}))
            base.update(raw)
            base.pop("line_points", None)
            base.pop("direction", None)
            cameras.append(base)
        return AppConfig(self.root_dir, self.system_config_path, self.camera_settings_path, system, cameras)

    def save_camera_settings(self, cameras: list[dict[str, Any]]) -> None:
        slim = [self._to_slim_camera_settings(cam) for cam in cameras]
        self._atomic_write_json(self.camera_settings_path, {"cameras": slim})

    def save_system_settings(self, system: dict[str, Any]) -> None:
        self._atomic_write_json(self.system_config_path, system)

    @staticmethod
    def _atomic_write_json(path: Path, data: dict[str, Any]) -> None:
        write_json_atomic(path, data, retries=10)

    @staticmethod
    def _to_slim_camera_settings(camera: dict[str, Any]) -> dict[str, Any]:
        return {k: camera[k] for k in KEEP_CAMERA_KEYS if k in camera}


# =========================================
# Congestion Logic
# =========================================
class CongestionScorer:
    """AICount11.py の congestion 算出式を監視向けに時間窓化して適用。3秒窓のフレーム平均で更新する。"""

    def __init__(self, interval_sec: int = 3, day_keep: int = 1, smoothing_window: int = 6):
        self.interval_sec = max(1, int(interval_sec))
        self.day_keep = max(1, day_keep)
        self.state = CongestionState()
        self.smoother = CongestionSmoother(smoothing_window)

    def update_interval(self, interval_sec: int) -> None:
        self.interval_sec = max(1, int(interval_sec))

    def update_smoothing_window(self, window_size: int) -> None:
        self.smoother.update_window_size(window_size)

    def _compute_frame_motion_score(self, tracks: list[dict], frame_width: int, now: datetime) -> float:
        if frame_width <= 0:
            return 0.0

        total = 0.0
        for tr in tracks:
            track_id = int(tr["track_id"])
            cx, cy = tr["center"]
            prev_x, prev_y = self.state.previous_positions.get(track_id, (cx, cy))
            distance = ((cx - prev_x) ** 2 + (cy - prev_y) ** 2) ** 0.5
            total += 1.0 / (1.0 + (distance / frame_width) * 500.0)

            self.state.previous_positions[track_id] = (cx, cy)
            self.state.last_seen_at[track_id] = now

        stale_before = now - timedelta(seconds=max(10, self.interval_sec * 3))
        stale_ids = [tid for tid, ts in self.state.last_seen_at.items() if ts < stale_before]
        for tid in stale_ids:
            self.state.last_seen_at.pop(tid, None)
            self.state.previous_positions.pop(tid, None)

        return total

    def update(self, tracks: list[dict], now: datetime, frame_width: int) -> float:
        if self.state.window_start is None:
            self.state.window_start = now
        frame_score = self._compute_frame_motion_score(tracks, frame_width, now)
        self.state.frame_cumulative_motion_score += frame_score
        self.state.window_frame_count += 1
        elapsed = (now - self.state.window_start).total_seconds()
        if elapsed < self.interval_sec:
            return self.state.current_congestion_index

        frame_count = max(1, self.state.window_frame_count)
        value = round(self.state.frame_cumulative_motion_score / frame_count, 3)
        self.state.frame_motion_scores.append(value)
        self.state.frame_time_stamps.append(now)
        self.state.current_congestion_index = value
        self.state.current_smoothed_index = round(self.smoother.add(value), 3)
        self.state.smoothed_motion_scores.append(self.state.current_smoothed_index)
        self.state.frame_cumulative_motion_score = 0.0
        self.state.window_frame_count = 0
        self.state.window_start = now

        day_ago = now - timedelta(days=self.day_keep)
        while self.state.frame_time_stamps and self.state.frame_time_stamps[0] < day_ago:
            self.state.frame_time_stamps.pop(0)
            self.state.frame_motion_scores.pop(0)
            if self.state.smoothed_motion_scores:
                self.state.smoothed_motion_scores.pop(0)
        return value


# =========================================
# Line Counter Logic
# =========================================
class LineCounter:
    def __init__(self, line_points: list[list[int]]):
        self.line_points = line_points
        self.state = CounterState()

    def update_line(self, line_points: list[list[int]]) -> None:
        self.line_points = line_points
        self.state.previous_side.clear()
        self.state.counted_track_ids.clear()

    def _signed_side(self, point: tuple[float, float]) -> float:
        p1, p2 = self.line_points
        vx, vy = p2[0] - p1[0], p2[1] - p1[1]
        wx, wy = point[0] - p1[0], point[1] - p1[1]
        return vx * wy - vy * wx

    def update(self, track_id: int, point: tuple[float, float], class_name: str, now: datetime):
        if len(self.line_points) != 2:
            return None
        current_side = self._signed_side(point)
        prev_side = self.state.previous_side.get(track_id)
        self.state.previous_side[track_id] = current_side
        if prev_side is None or track_id in self.state.counted_track_ids:
            return None

        crossed = (prev_side < 0 <= current_side) or (prev_side > 0 >= current_side)
        if not crossed:
            return None

        direction = "LtoR" if prev_side < current_side else "RtoL"
        self.state.counted_track_ids.add(track_id)
        bin_index = (now.hour * 60 + now.minute) // 10
        if 0 <= bin_index < 144:
            if direction == "LtoR":
                self.state.pass_bins_ltor[bin_index] += 1
            else:
                self.state.pass_bins_rtol[bin_index] += 1

        return {
            "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
            "track_id": track_id,
            "class_name": class_name,
            "direction": direction,
        }


class WakimuraAlphaCalculator:
    """20_rotary_efficiency_analysis.py を基にした参考指標（LEVEL判定には不使用）。"""

    def __init__(
        self,
        rotary_capacity: int = 10,
        base_stay_time_sec: float = 60.0,
        window_seconds: int = 300,
        high_load_vehicle_threshold: int = 7,
    ):
        self.rotary_capacity = max(1, int(rotary_capacity))
        self.base_stay_time_sec = max(1e-6, float(base_stay_time_sec))
        self.window_seconds = max(1, int(window_seconds))
        self.high_load_vehicle_threshold = max(1, int(high_load_vehicle_threshold))
        self.state = WakimuraAlphaState()

    def record_exit(self, now: datetime) -> None:
        self.state.exit_timestamps.append(now)

    def update(self, now: datetime, vehicle_count: int, avg_stay_sec: float) -> dict[str, Any]:
        window_start = now - timedelta(seconds=self.window_seconds)
        self.state.exit_timestamps = [ts for ts in self.state.exit_timestamps if ts >= window_start]
        n_out = len(self.state.exit_timestamps)
        c_nominal = 3600.0 * self.rotary_capacity / self.base_stay_time_sec
        c_effective = 3600.0 * n_out / float(self.window_seconds)
        alpha_window = (c_effective / c_nominal) if c_nominal > 0 else 0.0

        if vehicle_count >= self.high_load_vehicle_threshold:
            if not self.state.high_load_active:
                self.state.high_load_active = True
                self.state.high_load_session_stays.clear()
                self.state.high_load_alpha = None
            if avg_stay_sec > 0.0:
                self.state.high_load_session_stays.append(float(avg_stay_sec))
            if self.state.high_load_session_stays:
                session_avg = float(np.mean(self.state.high_load_session_stays))
                if session_avg > 0.0:
                    self.state.high_load_alpha = self.base_stay_time_sec / session_avg
        else:
            self.state.high_load_active = False
            self.state.high_load_session_stays.clear()
            self.state.high_load_alpha = None

        alpha = alpha_window
        if self.state.high_load_active and self.state.high_load_alpha is not None:
            alpha = self.state.high_load_alpha

        self.state.current_alpha = float(alpha)
        self.state.current_alpha_window = float(alpha_window)
        self.state.current_n_out = int(n_out)
        self.state.current_avg_stay_sec = max(0.0, float(avg_stay_sec))
        return {
            "wakimura_alpha": float(self.state.current_alpha),
            "wakimura_alpha_window": float(self.state.current_alpha_window),
            "wakimura_n_out": int(self.state.current_n_out),
            "wakimura_avg_stay_sec": float(self.state.current_avg_stay_sec),
            "wakimura_high_load_mode": bool(self.state.high_load_active),
        }


# =========================================
# Status Manager
# =========================================
class StatusManager:
    def __init__(self, ai_status_path: Path, system_cfg: dict[str, Any]):
        self.ai_status_path = ai_status_path
        self.system_cfg = system_cfg
        self._last_level: int | None = None
        self._last_write_time: float = 0.0
        self.last_failure_at: float = 0.0

    def update_if_needed(self, level: int) -> None:
        now = time.time()
        update_interval_sec = max(1, int(self.system_cfg.get("status_update_interval_sec", 10)))
        if level == self._last_level and (now - self._last_write_time) < update_interval_sec:
            return
        payload = {
            "congestion_level": int(level),
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        try:
            self._atomic_write_json(self.ai_status_path, payload)
        except Exception as exc:
            now_ts = time.time()
            if now_ts - self.last_failure_at > 5.0:
                self.last_failure_at = now_ts
                logging.warning("ai_status write skipped: %s", exc)
            return
        self._last_level = int(level)
        self._last_write_time = now

    @staticmethod
    def _atomic_write_json(path: Path, data: dict) -> None:
        write_json_atomic(path, data, retries=10)


# =========================================
# Report Writer
# =========================================
class ReportWriter:
    def __init__(self, output_root: Path):
        self.output_root = output_root

    def write_daily_report(self, target_date: date, cameras: list[dict], metrics_root: Path) -> Path:
        del cameras, metrics_root
        out_path = self.output_root / "reports" / "daily" / f"{target_date.isoformat()}.xlsx"
        if not out_path.exists():
            raise FileNotFoundError(f"日次Excelが見つかりません: {out_path}")
        wb = load_workbook(out_path)
        if "Graph" not in wb.sheetnames:
            wb.create_sheet("Graph")
        wb.save(out_path)
        return out_path

    def write_monthly_report(self, target_month: str, metrics_root: Path, cameras: list[dict]) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "monthly"
        self._style_sheet(ws)
        ws.append(
            [
                "date",
                "max_level",
                "cam1_avg_congestion",
                "cam1_total_ltor",
                "cam1_total_rtol",
                "cam2_avg_congestion",
                "cam2_total_ltor",
                "cam2_total_rtol",
                "cam3_avg_congestion",
                "cam3_total_ltor",
                "cam3_total_rtol",
            ]
        )
        del metrics_root, cameras

        daily_dir = self.output_root / "reports" / "daily"
        for file in sorted(daily_dir.glob(f"{target_month}-*.xlsx")):
            if file.name.startswith("daily_report_"):
                continue
            try:
                data_df = pd.read_excel(file, sheet_name="Data")
            except Exception:
                continue
            if data_df.empty:
                continue

            ws.append(
                [
                    file.stem,
                    int(data_df["渋滞LEVEL"].max()),
                    round(float(data_df["Camera1 渋滞指数"].mean()), 3),
                    int(data_df["Camera1 LtoR"].sum()),
                    int(data_df["Camera1 RtoL"].sum()),
                    round(float(data_df["Camera2 渋滞指数"].mean()), 3),
                    int(data_df["Camera2 LtoR"].sum()),
                    int(data_df["Camera2 RtoL"].sum()),
                    round(float(data_df["Camera3 渋滞指数"].mean()), 3),
                    int(data_df["Camera3 LtoR"].sum()),
                    int(data_df["Camera3 RtoL"].sum()),
                ]
            )

        out_path = self.output_root / "reports" / "monthly" / f"monthly_report_{target_month}.xlsx"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def _add_camera_sheet(self, wb: Workbook, cam: dict, pass_df: pd.DataFrame, metric_df: pd.DataFrame, long_df: pd.DataFrame) -> None:
        ws = wb.create_sheet(f"cam{cam['camera_id']}")
        self._style_sheet(ws)
        ws["A1"] = f"{cam['camera_name']} Detail"
        ws["A1"].font = Font(bold=True, size=14, color="00D7FF")

        hist = [0] * 144
        if not pass_df.empty:
            pass_df["timestamp"] = pd.to_datetime(pass_df["timestamp"])
            for ts in pass_df["timestamp"]:
                hist[(ts.hour * 60 + ts.minute) // 10] += 1

        ws.append(["bin", "pass_count"])
        for i, v in enumerate(hist):
            ws.append([i, v])

        bar = BarChart()
        bar.title = "Pass Histogram"
        data = Reference(ws, min_col=2, min_row=3, max_row=146)
        cats = Reference(ws, min_col=1, min_row=3, max_row=146)
        bar.add_data(data, titles_from_data=False)
        bar.set_categories(cats)
        bar.height = 6
        bar.width = 13
        ws.add_chart(bar, "D3")

        start_row = 150
        ws[f"A{start_row}"] = "congestion_time_series"
        ws.append(["timestamp", "congestion_score"])
        for _, r in metric_df[["timestamp", "congestion_score"]].iterrows() if not metric_df.empty else []:
            ws.append([r["timestamp"], float(r["congestion_score"])])

        if len(metric_df) > 2:
            line = LineChart()
            line.title = "Congestion Score"
            dref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + len(metric_df))
            cref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(metric_df))
            line.add_data(dref, titles_from_data=False)
            line.set_categories(cref)
            line.height = 5
            line.width = 13
            ws.add_chart(line, "D18")

        ls_row = 18
        ws[f"A{ls_row}"] = "long_stay_list"
        ws[f"A{ls_row+1}"] = "track_id"
        ws[f"B{ls_row+1}"] = "stay_minutes"
        for i, (_, r) in enumerate(long_df.iterrows() if not long_df.empty else []):
            ws[f"A{ls_row+2+i}"] = int(r["track_id"])
            ws[f"B{ls_row+2+i}"] = float(r["stay_minutes"])

    def _style_sheet(self, ws):
        fill = PatternFill("solid", fgColor="101922")
        thin = Side(style="thin", color="1A9FB6")
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 22
        for row in range(1, 220):
            for col in range(1, 8):
                c = ws.cell(row=row, column=col)
                c.fill = fill
                c.font = Font(color="FFFFFF", size=10)
                c.alignment = Alignment(vertical="center")
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

class TenMinuteRecordWriter:
    HEADERS = [
        "時刻（10分単位）",
        "渋滞LEVEL",
        "Camera1 渋滞指数",
        "Camera1 脇村指標α",
        "Camera1 LtoR",
        "Camera1 RtoL",
        "Camera2 渋滞指数",
        "Camera2 脇村指標α",
        "Camera2 LtoR",
        "Camera2 RtoL",
        "Camera3 渋滞指数",
        "Camera3 脇村指標α",
        "Camera3 LtoR",
        "Camera3 RtoL",
    ]

    def __init__(self, output_root: Path):
        self.output_root = output_root / "reports" / "daily"

    @staticmethod
    def _floor_to_10min(dt: datetime) -> datetime:
        return dt.replace(minute=(dt.minute // 10) * 10, second=0, microsecond=0)

    def collect_sample(self, now: datetime, level: int, latest_payloads: dict[int, dict[str, Any]]) -> dict[str, Any]:
        sample: dict[str, Any] = {"time": now, "level": int(level), "cameras": {}}
        for cid in (1, 2, 3):
            payload = latest_payloads.get(cid, {})
            sample["cameras"][cid] = {
                "index": float(payload.get("smoothed_congestion_score", payload.get("congestion_score", 0.0))),
                "wakimura_alpha": float(payload.get("wakimura_alpha", 0.0)),
                "ltor_total": int(payload.get("count_ltor", 0)),
                "rtol_total": int(payload.get("count_rtol", 0)),
            }
        return sample

    def aggregate_10min(self, bin_start: datetime, samples: list[dict[str, Any]]) -> list[Any]:
        if not samples:
            return [bin_start.strftime("%H:%M"), 1, 0.0, 0.0, 0, 0, 0.0, 0.0, 0, 0, 0.0, 0.0, 0, 0]
        level_max = max(int(s["level"]) for s in samples)
        row: list[Any] = [bin_start.strftime("%H:%M"), level_max]
        for cid in (1, 2, 3):
            indices = [float(s["cameras"][cid]["index"]) for s in samples]
            wakimuras = [float(s["cameras"][cid]["wakimura_alpha"]) for s in samples]
            ltor_values = [int(s["cameras"][cid]["ltor_total"]) for s in samples]
            rtol_values = [int(s["cameras"][cid]["rtol_total"]) for s in samples]
            ltor = max(0, max(ltor_values) - min(ltor_values)) if ltor_values else 0
            rtol = max(0, max(rtol_values) - min(rtol_values)) if rtol_values else 0
            avg_index = round(float(np.mean(indices)), 3) if indices else 0.0
            avg_wakimura = round(float(np.mean(wakimuras)), 3) if wakimuras else 0.0
            row.extend([avg_index, avg_wakimura, ltor, rtol])
        return row

    def append_row(self, target_date: date, row: list[Any]) -> Path:
        wb, path = self._load_or_create(target_date)
        ws = wb["Data"]
        ws.append(row)
        wb.save(path)
        return path

    def finalize_day_graph(self, target_date: date) -> Path | None:
        path = self.output_root / f"{target_date.isoformat()}.xlsx"
        if not path.exists():
            return None
        wb = load_workbook(path)
        ws_data = wb["Data"] if "Data" in wb.sheetnames else wb.active
        if "Graph" in wb.sheetnames:
            del wb["Graph"]
        ws_graph = wb.create_sheet("Graph")
        last_row = ws_data.max_row
        if last_row < 2:
            wb.save(path)
            return path

        cats = Reference(ws_data, min_col=1, min_row=2, max_row=last_row)

        chart_level = LineChart()
        chart_level.title = "渋滞レベル推移"
        chart_level.y_axis.scaling.min = 1
        chart_level.y_axis.scaling.max = 4
        chart_level.height = 6
        chart_level.width = 10
        chart_level.add_data(Reference(ws_data, min_col=2, min_row=1, max_row=last_row), titles_from_data=True)
        chart_level.set_categories(cats)
        ws_graph.add_chart(chart_level, "A1")

        chart_index = LineChart()
        chart_index.title = "カメラ別渋滞指数"
        chart_index.height = 6
        chart_index.width = 10
        for col in (3, 7, 11):
            chart_index.add_data(Reference(ws_data, min_col=col, max_col=col, min_row=1, max_row=last_row), titles_from_data=True)
        chart_index.set_categories(cats)
        ws_graph.add_chart(chart_index, "A18")

        chart_traffic = BarChart()
        chart_traffic.title = "交通量（LtoR / RtoL）"
        chart_traffic.height = 6
        chart_traffic.width = 10
        for col in (5, 6, 9, 10, 13, 14):
            chart_traffic.add_data(Reference(ws_data, min_col=col, max_col=col, min_row=1, max_row=last_row), titles_from_data=True)
        chart_traffic.set_categories(cats)
        ws_graph.add_chart(chart_traffic, "L1")

        ws_graph.page_setup.paperSize = ws_graph.PAPERSIZE_A4
        ws_graph.page_setup.orientation = ws_graph.ORIENTATION_LANDSCAPE
        ws_graph.page_setup.fitToWidth = 1
        ws_graph.page_setup.fitToHeight = 1
        ws_graph.page_margins = PageMargins(left=0.2, right=0.2, top=0.3, bottom=0.3)
        wb.save(path)
        return path

    def _load_or_create(self, target_date: date) -> tuple[Workbook, Path]:
        self.output_root.mkdir(parents=True, exist_ok=True)
        path = self.output_root / f"{target_date.isoformat()}.xlsx"
        if path.exists():
            wb = load_workbook(path)
            ws = wb["Data"] if "Data" in wb.sheetnames else wb.active
            existing_headers = [ws.cell(row=1, column=i + 1).value for i in range(len(self.HEADERS))]
            if existing_headers != self.HEADERS:
                wb = Workbook()
                ws = wb.active
                ws.title = "Data"
                ws.append(self.HEADERS)
                wb.create_sheet("Graph")
                wb.save(path)
            return wb, path
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(self.HEADERS)
        wb.create_sheet("Graph")
        wb.save(path)
        return wb, path


# =========================================
# UI Panels
# =========================================
class ClickableImageLabel(QtWidgets.QLabel):
    point_clicked = QtCore.pyqtSignal(int, int)

    def __init__(self, text: str = "", parent=None):
        super().__init__(text, parent)
        self.source_size = (1, 1)

    def set_source_size(self, width: int, height: int) -> None:
        self.source_size = (max(1, width), max(1, height))

    def mousePressEvent(self, event: QtGui.QMouseEvent) -> None:
        pix = self.pixmap()
        if pix is None:
            return
        scaled = pix.size()
        off_x = (self.width() - scaled.width()) // 2
        off_y = (self.height() - scaled.height()) // 2
        local_x = event.pos().x() - off_x
        local_y = event.pos().y() - off_y
        if local_x < 0 or local_y < 0 or local_x >= scaled.width() or local_y >= scaled.height():
            return
        src_w, src_h = self.source_size
        x = int(local_x * src_w / max(1, scaled.width()))
        y = int(local_y * src_h / max(1, scaled.height()))
        self.point_clicked.emit(x, y)
        super().mousePressEvent(event)


def _segments_intersect(p1, p2, p3, p4) -> bool:
    def ccw(a, b, c):
        return (c[1] - a[1]) * (b[0] - a[0]) > (b[1] - a[1]) * (c[0] - a[0])
    return ccw(p1, p3, p4) != ccw(p2, p3, p4) and ccw(p1, p2, p3) != ccw(p1, p2, p4)


def has_self_intersection(points: list[list[int]]) -> bool:
    if len(points) < 4:
        return False
    n = len(points)
    for i in range(n):
        a1 = points[i]
        a2 = points[(i + 1) % n]
        for j in range(i + 1, n):
            if abs(i - j) <= 1 or (i == 0 and j == n - 1):
                continue
            b1 = points[j]
            b2 = points[(j + 1) % n]
            if _segments_intersect(a1, a2, b1, b2):
                return True
    return False


class CameraSettingsDialog(QtWidgets.QDialog):
    def __init__(self, camera_cfg: dict[str, Any], latest_frame=None, parent=None, initial_mode: str = "line"):
        super().__init__(parent)
        self.setWindowTitle(f"設定: {camera_cfg['camera_name']}")
        self.resize(1000, 780)
        self.camera_cfg = dict(camera_cfg)
        self.line_points = [camera_cfg.get("line_start", [100, 100])[:], camera_cfg.get("line_end", [400, 100])[:]]
        self.exclude_polygon: list[list[int]] = [p[:] for p in camera_cfg.get("exclude_polygon", [])]
        self.mode = initial_mode if initial_mode in {"line", "poly"} else "line"

        root = QtWidgets.QVBoxLayout(self)
        tabs = QtWidgets.QTabWidget()
        root.addWidget(tabs)

        basic = QtWidgets.QWidget()
        tabs.addTab(basic, "基本")
        form = QtWidgets.QFormLayout(basic)
        self.name_edit = QtWidgets.QLineEdit(camera_cfg.get("camera_name", ""))
        self.url_edit = QtWidgets.QLineEdit(str(camera_cfg.get("stream_url", "")))
        form.addRow("カメラ名", self.name_edit)
        form.addRow("stream_url", self.url_edit)

        ai_tab = QtWidgets.QWidget()
        tabs.addTab(ai_tab, "解析条件")
        ai_form = QtWidgets.QFormLayout(ai_tab)

        self.yolo_model = QtWidgets.QLineEdit(str(camera_cfg.get("yolo_model", "yolo11m.pt")))
        self.conf = QtWidgets.QDoubleSpinBox(); self.conf.setRange(0, 1); self.conf.setSingleStep(0.01); self.conf.setValue(float(camera_cfg.get("confidence_threshold", 0.25)))
        self.iou = QtWidgets.QDoubleSpinBox(); self.iou.setRange(0, 1); self.iou.setSingleStep(0.01); self.iou.setValue(float(camera_cfg.get("iou_threshold", 0.5)))
        self.frame_skip = QtWidgets.QSpinBox(); self.frame_skip.setRange(1, 30); self.frame_skip.setValue(int(camera_cfg.get("frame_skip", 1)))
        self.imgsz = QtWidgets.QSpinBox(); self.imgsz.setRange(320, 2048); self.imgsz.setSingleStep(32); self.imgsz.setValue(int(camera_cfg.get("imgsz", 640)))
        self.bt_hi = QtWidgets.QDoubleSpinBox(); self.bt_hi.setRange(0, 1); self.bt_hi.setValue(float(camera_cfg.get("bt_track_high_thresh", 0.3)))
        self.bt_lo = QtWidgets.QDoubleSpinBox(); self.bt_lo.setRange(0, 1); self.bt_lo.setValue(float(camera_cfg.get("bt_track_low_thresh", 0.1)))
        self.bt_match = QtWidgets.QDoubleSpinBox(); self.bt_match.setRange(0, 1); self.bt_match.setValue(float(camera_cfg.get("bt_match_thresh", 0.8)))
        self.bt_buffer = QtWidgets.QSpinBox(); self.bt_buffer.setRange(1, 1000); self.bt_buffer.setValue(int(camera_cfg.get("bt_track_buffer", 30)))
        self.crossing = QtWidgets.QLineEdit(str(camera_cfg.get("crossing_judgment_pattern", "line_cross")))
        self.distance_th = QtWidgets.QDoubleSpinBox(); self.distance_th.setRange(1, 1000); self.distance_th.setValue(float(camera_cfg.get("distance_threshold", 25.0)))
        self.cong_interval = QtWidgets.QSpinBox(); self.cong_interval.setRange(1, 60); self.cong_interval.setValue(int(camera_cfg.get("congestion_calculation_interval", 3)))
        self.enable_cong = QtWidgets.QCheckBox("enable_congestion"); self.enable_cong.setChecked(bool(camera_cfg.get("enable_congestion", True)))
        self.spin_threshold = QtWidgets.QDoubleSpinBox(); self.spin_threshold.setRange(0.0, 50.0); self.spin_threshold.setDecimals(2); self.spin_threshold.setSingleStep(0.1); self.spin_threshold.setValue(float(camera_cfg.get("congestion_threshold", 5)))
        self.spin_stay = QtWidgets.QSpinBox(); self.spin_stay.setRange(1, 120); self.spin_stay.setValue(int(camera_cfg.get("long_stay_minutes", 15)))

        ai_form.addRow("yolo_model（未設定時のみ使用）", self.yolo_model)
        ai_form.addRow("confidence_threshold", self.conf)
        ai_form.addRow("iou_threshold", self.iou)
        ai_form.addRow("frame_skip", self.frame_skip)
        ai_form.addRow("imgsz", self.imgsz)
        ai_form.addRow("bt_track_high_thresh", self.bt_hi)
        ai_form.addRow("bt_track_low_thresh", self.bt_lo)
        ai_form.addRow("bt_match_thresh", self.bt_match)
        ai_form.addRow("bt_track_buffer", self.bt_buffer)
        ai_form.addRow("crossing_judgment_pattern", self.crossing)
        ai_form.addRow("distance_threshold", self.distance_th)
        ai_form.addRow("congestion_calculation_interval", self.cong_interval)
        ai_form.addRow(self.enable_cong)
        ai_form.addRow("渋滞指数閾値", self.spin_threshold)
        ai_form.addRow("長時間滞在(分)", self.spin_stay)

        class_group = QtWidgets.QGroupBox("対象クラス (0-7)")
        class_layout = QtWidgets.QGridLayout(class_group)
        self.class_checks: dict[int, QtWidgets.QCheckBox] = {}
        selected = set(int(x) for x in camera_cfg.get("target_classes", [2, 3, 5, 7]))
        for i in range(8):
            cb = QtWidgets.QCheckBox(f"{CLASS_MAP[i]}({i})")
            cb.setChecked(i in selected)
            self.class_checks[i] = cb
            class_layout.addWidget(cb, i // 4, i % 4)
        ai_form.addRow(class_group)

        self.image = ClickableImageLabel("snapshot")
        self.image.setMinimumHeight(340)
        self.image.setStyleSheet("background:#0c0f16;border:1px solid #00D7FF;")
        self.image.point_clicked.connect(self._on_click)
        root.addWidget(self.image)

        self.mode_label = QtWidgets.QLabel("")
        self.mode_label.setStyleSheet("color:#d4e6ff;font-weight:bold;")
        root.addWidget(self.mode_label)

        row = QtWidgets.QHBoxLayout()
        btn_line = QtWidgets.QPushButton("ライン設定")
        btn_line.clicked.connect(lambda: self._set_mode("line"))
        btn_poly = QtWidgets.QPushButton("除外エリア設定")
        btn_poly.clicked.connect(lambda: self._set_mode("poly"))
        btn_finish_poly = QtWidgets.QPushButton("指定終了")
        btn_finish_poly.clicked.connect(self._finish_polygon)
        btn_reset = QtWidgets.QPushButton("やり直し")
        btn_reset.clicked.connect(self._reset_mode)
        row.addWidget(btn_line); row.addWidget(btn_poly); row.addWidget(btn_finish_poly); row.addWidget(btn_reset)
        root.addLayout(row)

        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Ok | QtWidgets.QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._validate_and_accept)
        bb.rejected.connect(self.reject)
        root.addWidget(bb)

        self.snapshot = latest_frame
        if self.snapshot is None:
            self.snapshot = np.zeros((360, 640, 3), dtype=np.uint8)
            cv2.putText(self.snapshot, "No live frame", (30, 180), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)
        self._update_mode_label()
        self._render_snapshot()

    def _set_mode(self, mode: str) -> None:
        self.mode = mode
        if mode == "line":
            QtWidgets.QMessageBox.information(self, "ライン設定", "ライン設定は2点を取り直します。")
        self._update_mode_label()

    def _update_mode_label(self) -> None:
        txt = "現在モード: ライン設定" if self.mode == "line" else "現在モード: 除外エリア設定"
        self.mode_label.setText(txt)

    def _on_click(self, x: int, y: int) -> None:
        if self.mode == "line":
            if len(self.line_points) >= 2:
                self.line_points = []
            self.line_points.append([x, y])
        else:
            self.exclude_polygon.append([x, y])
        self._render_snapshot()

    def _finish_polygon(self) -> None:
        if len(self.exclude_polygon) < 3:
            QtWidgets.QMessageBox.warning(self, "警告", "除外エリアは3点以上必要です。")
            return
        if has_self_intersection(self.exclude_polygon):
            QtWidgets.QMessageBox.warning(self, "警告", "自己交差ポリゴンは設定できません。やり直してください。")
            self.exclude_polygon = []
        self._render_snapshot()

    def _reset_mode(self) -> None:
        if self.mode == "line":
            self.line_points = []
        else:
            self.exclude_polygon = []
        self._render_snapshot()

    def _render_snapshot(self) -> None:
        frame = self.snapshot.copy()
        if len(self.line_points) == 2:
            cv2.line(frame, tuple(self.line_points[0]), tuple(self.line_points[1]), (0, 255, 255), 2)
            for p in self.line_points:
                cv2.circle(frame, tuple(p), 4, (0, 255, 255), -1)
                cv2.putText(frame, f"{p[0]},{p[1]}", (p[0] + 5, p[1] - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)
        for i, p in enumerate(self.exclude_polygon):
            cv2.circle(frame, tuple(p), 4, (150, 150, 150), -1)
            cv2.putText(frame, str(i + 1), (p[0] + 5, p[1] - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)
        if len(self.exclude_polygon) >= 2:
            cv2.polylines(frame, [np.array(self.exclude_polygon, np.int32)], len(self.exclude_polygon) >= 3, (120, 120, 120), 2)

        h, w, _ = frame.shape
        self.image.set_source_size(w, h)
        qimg = QtGui.QImage(frame.data, w, h, frame.strides[0], QtGui.QImage.Format.Format_BGR888)
        self.image.setPixmap(QtGui.QPixmap.fromImage(qimg).scaled(self.image.size(), QtCore.Qt.AspectRatioMode.KeepAspectRatio))

    def resizeEvent(self, event: QtGui.QResizeEvent) -> None:
        super().resizeEvent(event)
        self._render_snapshot()

    def _validate_and_accept(self) -> None:
        selected_classes = [i for i, cb in self.class_checks.items() if cb.isChecked()]
        if not selected_classes:
            QtWidgets.QMessageBox.warning(self, "警告", "対象クラスを1つ以上選択してください。")
            return
        if len(self.line_points) != 2:
            QtWidgets.QMessageBox.warning(self, "警告", "ラインは2点指定してください。")
            return
        self.accept()

    def get_updated_config(self) -> dict[str, Any]:
        cfg = dict(self.camera_cfg)
        cfg["camera_name"] = self.name_edit.text().strip() or cfg["camera_name"]
        cfg["stream_url"] = self.url_edit.text().strip()
        cfg["line_start"] = self.line_points[0]
        cfg["line_end"] = self.line_points[1]
        cfg["exclude_polygon"] = self.exclude_polygon
        cfg["yolo_model"] = self.yolo_model.text().strip() or "yolo11m.pt"
        cfg["confidence_threshold"] = float(self.conf.value())
        cfg["iou_threshold"] = float(self.iou.value())
        cfg["frame_skip"] = int(self.frame_skip.value())
        cfg["imgsz"] = int(self.imgsz.value())
        cfg["target_classes"] = [i for i, cb in self.class_checks.items() if cb.isChecked()]
        cfg["bt_track_high_thresh"] = float(self.bt_hi.value())
        cfg["bt_track_low_thresh"] = float(self.bt_lo.value())
        cfg["bt_match_thresh"] = float(self.bt_match.value())
        cfg["bt_track_buffer"] = int(self.bt_buffer.value())
        cfg["crossing_judgment_pattern"] = self.crossing.text().strip() or "line_cross"
        cfg["distance_threshold"] = float(self.distance_th.value())
        cfg["congestion_calculation_interval"] = int(self.cong_interval.value())
        cfg["enable_congestion"] = bool(self.enable_cong.isChecked())
        cfg["congestion_threshold"] = float(self.spin_threshold.value())
        cfg["long_stay_minutes"] = int(self.spin_stay.value())
        return cfg


class CombinedTimelineGraph(QtWidgets.QWidget):
    def __init__(self, mode: str = "line", parent=None):
        super().__init__(parent)
        self.mode = mode
        self.title = ""
        self.today_points: list[tuple[datetime, float]] = []
        self.prev_points: list[tuple[datetime, float]] = []
        self.today_values: list[float] = []
        self.prev_values: list[float] = []
        self.threshold: float | None = None
        self.show_threshold = False
        self.y_min_override: float | None = None
        self.y_max_override: float | None = None
        self.y_axis_labels: dict[float, str] = {}
        self.setFixedHeight(75)

    def set_line_data(self, prev_points: list[tuple[datetime, float]], today_points: list[tuple[datetime, float]], title: str, threshold: float | None = None, show_threshold: bool = True) -> None:
        self.mode = "line"
        self.prev_points = prev_points
        self.today_points = today_points
        self.prev_values = []
        self.today_values = []
        self.title = title
        self.threshold = threshold
        self.show_threshold = show_threshold
        self.update()

    def set_bar_data(self, prev_values: list[int], today_values: list[int], title: str) -> None:
        self.mode = "bar"
        self.prev_values = [float(v) for v in prev_values]
        self.today_values = [float(v) for v in today_values]
        self.prev_points = []
        self.today_points = []
        self.title = title
        self.threshold = None
        self.show_threshold = False
        self.update()

    def set_y_axis_config(self, y_min: float | None = None, y_max: float | None = None, labels: dict[float, str] | None = None) -> None:
        self.y_min_override = y_min
        self.y_max_override = y_max
        self.y_axis_labels = labels or {}
        self.update()

    def paintEvent(self, event: QtGui.QPaintEvent) -> None:
        super().paintEvent(event)
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        painter.fillRect(self.rect(), QtGui.QColor("#0f1620"))

        y_axis_w = 28
        right_margin = 5
        header_h = 22
        top_margin = header_h + 2
        bottom_margin = 15
        plot = QtCore.QRectF(y_axis_w, top_margin, max(10, self.width() - y_axis_w - right_margin), max(10, self.height() - top_margin - bottom_margin))

        painter.setPen(QtGui.QPen(QtGui.QColor("#1d6f8b"), 1))
        painter.drawRoundedRect(self.rect().adjusted(0, 0, -1, -1), 3, 3)
        painter.drawRect(plot)
        painter.setPen(QtGui.QColor("#cfefff"))
        title_rect = QtCore.QRectF(0, 0, self.width(), header_h)
        painter.drawText(
            title_rect.adjusted(8, 0, -8, 0),
            QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter,
            self.title,
        )

        if self.mode == "line":
            ys = [v for _, v in self.prev_points] + [v for _, v in self.today_points]
        else:
            ys = self.prev_values[:] + self.today_values[:]
        if self.show_threshold and self.threshold is not None:
            ys.append(float(self.threshold))
        y_max = float(self.y_max_override) if self.y_max_override is not None else max(1.0, max(ys) if ys else 1.0)
        y_min = float(self.y_min_override) if self.y_min_override is not None else 0.0
        if y_max <= y_min:
            y_max = y_min + 1.0

        if self.y_axis_labels:
            ticks = [(float(v), str(lbl)) for v, lbl in sorted(self.y_axis_labels.items(), key=lambda x: x[0]) if y_min <= float(v) <= y_max]
        else:
            ticks = []
            for i in range(5):
                ratio = i / 4.0
                value = y_min + (y_max - y_min) * ratio
                ticks.append((value, f"{value:.1f}"))

        for value, tick_label in ticks:
            ratio = (value - y_min) / (y_max - y_min)
            y = plot.bottom() - ratio * plot.height()
            painter.setPen(QtGui.QPen(QtGui.QColor("#274457"), 1))
            painter.drawLine(QtCore.QPointF(plot.left(), y), QtCore.QPointF(plot.right(), y))
            painter.setPen(QtGui.QColor("#8db6c7"))
            painter.drawText(2, int(y) + 4, tick_label)

        x_ticks = list(range(25))
        for h in x_ticks:
            x = plot.left() + (h / 24.0) * plot.width()
            painter.setPen(QtGui.QPen(QtGui.QColor("#274457"), 1))
            painter.drawLine(QtCore.QPointF(x, plot.top()), QtCore.QPointF(x, plot.bottom()))
            painter.setPen(QtGui.QColor("#8db6c7"))
            text = f"{h}"
            text_w = painter.fontMetrics().horizontalAdvance(text)
            painter.drawText(int(x - text_w / 2), self.height() - 6, text)

        if self.mode == "line" and (self.prev_points or self.today_points):
            for points, color in [(self.prev_points, QtGui.QColor("#2f7dff")), (self.today_points, QtGui.QColor("#ff3b3b"))]:
                if not points:
                    continue
                path = QtGui.QPainterPath()
                for i, (ts, value) in enumerate(points):
                    sec = ts.hour * 3600 + ts.minute * 60 + ts.second
                    x = plot.left() + (sec / 86400.0) * plot.width()
                    y = plot.bottom() - ((value - y_min) / (y_max - y_min)) * plot.height()
                    if i == 0:
                        path.moveTo(x, y)
                    else:
                        path.lineTo(x, y)
                painter.setPen(QtGui.QPen(color, 2.0))
                painter.drawPath(path)
            if self.show_threshold and self.threshold is not None:
                th_y = plot.bottom() - ((self.threshold - y_min) / (y_max - y_min)) * plot.height()
                painter.setPen(QtGui.QPen(QtGui.QColor("#ffd400"), 2.0))
                painter.drawLine(QtCore.QPointF(plot.left(), th_y), QtCore.QPointF(plot.right(), th_y))
        elif self.mode == "bar" and (self.prev_values or self.today_values):
            n = max(1, len(self.prev_values), len(self.today_values))
            slot_w = plot.width() / n
            bar_w = max(1.8, slot_w * 0.38)
            for i in range(n):
                prev_val = self.prev_values[i] if i < len(self.prev_values) else 0.0
                today_val = self.today_values[i] if i < len(self.today_values) else 0.0
                center_x = plot.left() + (i + 0.5) * slot_w
                prev_h = ((prev_val - y_min) / (y_max - y_min)) * plot.height()
                today_h = ((today_val - y_min) / (y_max - y_min)) * plot.height()
                offset = min(slot_w * 0.22, max(2.0, bar_w * 0.7))
                prev_x = center_x - offset - bar_w / 2
                today_x = center_x + offset - bar_w / 2
                painter.fillRect(QtCore.QRectF(prev_x, plot.bottom() - prev_h, bar_w, prev_h), QtGui.QColor("#2f7dff"))
                painter.fillRect(QtCore.QRectF(today_x, plot.bottom() - today_h, bar_w, today_h), QtGui.QColor("#ff3b3b"))

        threshold_text = f"TH={self.threshold:.2f}" if (self.mode == "line" and self.show_threshold and self.threshold is not None) else None
        self._draw_legend(painter, title_rect, threshold_text)

    def _draw_legend(self, painter: QtGui.QPainter, header_rect: QtCore.QRectF, threshold_text: str | None = None) -> None:
        legend = [("前日", QtGui.QColor("#2f7dff")), ("当日", QtGui.QColor("#ff3b3b"))]
        if self.mode == "line" and self.show_threshold:
            legend.append(("閾値", QtGui.QColor("#ffd400")))
        y = int(header_rect.center().y())
        x = int(header_rect.right()) - 8
        if threshold_text:
            painter.setPen(QtGui.QColor("#ffd400"))
            th_w = painter.fontMetrics().horizontalAdvance(threshold_text)
            x -= th_w
            painter.drawText(x, y + 4, threshold_text)
            x -= 12
        for label, color in reversed(legend):
            text_w = painter.fontMetrics().horizontalAdvance(label)
            x -= text_w
            painter.setPen(QtGui.QColor("#d9ecff"))
            painter.drawText(x, y + 4, label)
            x -= 16
            painter.setPen(QtGui.QPen(color, 2))
            painter.drawLine(x, y, x + 12, y)
            x -= 10


class CongestionIndexBar(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.score = 0.0
        self.threshold = 5.0
        self.setMinimumHeight(30)
        self.setMaximumHeight(30)

    def set_values(self, score: float, threshold: float) -> None:
        self.score = float(score)
        self.threshold = float(threshold)
        self.update()

    def paintEvent(self, event: QtGui.QPaintEvent) -> None:
        super().paintEvent(event)
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        rect = self.rect().adjusted(1, 1, -1, -1)
        painter.setPen(QtGui.QPen(QtGui.QColor("#1f4f7a"), 1))
        painter.setBrush(QtGui.QColor("#0c1a24"))
        painter.drawRoundedRect(rect, 4, 4)

        score_limited = max(0.0, min(20.0, self.score))
        threshold_limited = max(0.0, min(20.0, self.threshold))
        ratio = score_limited / 20.0
        fill = QtCore.QRectF(rect.left(), rect.top(), rect.width() * ratio, rect.height())
        bar_color = QtGui.QColor("#ff3b3b" if self.score >= self.threshold else "#2f7dff")
        painter.fillRect(fill, bar_color)

        th_x = rect.left() + rect.width() * (threshold_limited / 20.0)
        painter.setPen(QtGui.QPen(QtGui.QColor("#ffd400"), 2))
        painter.drawLine(QtCore.QPointF(th_x, rect.top()), QtCore.QPointF(th_x, rect.bottom()))

        painter.setPen(QtGui.QColor("#ffffff"))
        font = painter.font()
        font.setPointSizeF(max(12.0, font.pointSizeF() * 1.2))
        font.setBold(True)
        painter.setFont(font)
        painter.drawText(rect, QtCore.Qt.AlignmentFlag.AlignCenter, f"渋滞指数 {self.score:.1f}")


class CameraPanel(QtWidgets.QFrame):
    line_setting_requested = QtCore.pyqtSignal(int)
    exclude_setting_requested = QtCore.pyqtSignal(int)
    camera_setting_requested = QtCore.pyqtSignal(int)
    threshold_changed = QtCore.pyqtSignal(int, float)

    def __init__(self, camera_cfg: dict[str, Any], parent=None):
        super().__init__(parent)
        self.camera_cfg = camera_cfg
        self.camera_id = camera_cfg["camera_id"]
        self._latest_pixmap: QtGui.QPixmap | None = None
        self._last_frame_size: tuple[int | None, int | None] = (None, None)
        self._stay_entries: list[list[float] | tuple[int, float]] = []
        self._last_stay_signature: tuple[tuple[int, int], ...] = ()
        self._last_stay_visible_count = -1
        self.stay_card_widgets: dict[int, QtWidgets.QWidget] = {}
        self.stay_empty_label: QtWidgets.QLabel | None = None
        self._status_connected = False
        self.is_king = self.camera_id == 2
        self.max_video_height = 270
        self.last_graph_update_ts = 0.0
        self.graph_update_interval_sec = 1.0
        self._last_graph_revision_ts = -1.0
        self._last_hist_revision = (-1, -1)
        display_name_map = {2: "KING", 1: "QUEEN", 3: "JACK"}
        role_name = display_name_map.get(self.camera_id, f"CAM{self.camera_id}")
        panel_tint_map = {
            "KING": "rgba(201,162,39,0.10)",
            "QUEEN": "rgba(194,24,91,0.10)",
            "JACK": "rgba(21,101,192,0.10)",
        }
        panel_tint = panel_tint_map.get(role_name, "rgba(15,26,40,0.95)")
        self.setStyleSheet(
            f"QFrame{{background:{panel_tint};border:1px solid #169db8;border-radius:6px;}} QLabel{{color:#cfefff;}}"
        )
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(3, 3, 3, 3)
        root.setSpacing(1)
        self.setMinimumWidth(1272)
        self.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred)

        top_row = QtWidgets.QHBoxLayout()
        top_row.setContentsMargins(0, 0, 0, 0)
        top_row.setSpacing(3)

        self.video_box = QtWidgets.QWidget()
        video_layout = QtWidgets.QVBoxLayout(self.video_box)
        video_layout.setContentsMargins(0, 0, 0, 0)
        video_layout.setSpacing(2)
        video_layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignTop)
        self.video = QtWidgets.QLabel("video")
        self.video.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.video.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred)
        self.video.setStyleSheet("background:#010203;border:1px solid #00a6d6;")
        video_layout.addWidget(self.video, 0, QtCore.Qt.AlignmentFlag.AlignTop)
        self.video_box.setSizePolicy(QtWidgets.QSizePolicy.Policy.Fixed, QtWidgets.QSizePolicy.Policy.Fixed)
        top_row.addWidget(self.video_box, 3, QtCore.Qt.AlignmentFlag.AlignTop)

        self.right_box = QtWidgets.QWidget()
        self.right_box.setMinimumHeight(0)
        self.right_box.setFixedHeight(self.max_video_height)
        self.right_box.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)
        right = QtWidgets.QVBoxLayout(self.right_box)
        right.setContentsMargins(0, 0, 0, 0)
        right.setSpacing(1)
        right.setAlignment(QtCore.Qt.AlignmentFlag.AlignTop)

        role_style = {
            "KING": ("#C9A227", "#111111"),
            "QUEEN": ("#C2185B", "#FFFFFF"),
            "JACK": ("#1565C0", "#FFFFFF"),
        }
        role_bg, role_fg = role_style.get(role_name, ("#2A3B47", "#FFFFFF"))
        self.role_badge = QtWidgets.QLabel(role_name)
        self.role_badge.setFixedHeight(38)
        self.role_badge.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.role_badge.setStyleSheet(
            f"background:{role_bg};color:{role_fg};border:1px solid #c8d4e0;border-radius:6px;font-size:20px;font-weight:900;letter-spacing:1px;padding:0px 4px;"
        )
        right.addWidget(self.role_badge, 0, QtCore.Qt.AlignmentFlag.AlignTop)

        self.title = QtWidgets.QLabel("")
        self.title.setStyleSheet("font-size:11px;color:#00D7FF;font-weight:bold;line-height:1.0em;")
        self.title.setWordWrap(False)
        self.title.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)
        right.addWidget(self.title, 0, QtCore.Qt.AlignmentFlag.AlignTop)

        self.congestion_bar = CongestionIndexBar()

        level_row = QtWidgets.QHBoxLayout()
        level_row.setContentsMargins(0, 0, 0, 0)
        level_row.setSpacing(4)

        level_row.addWidget(self.congestion_bar, 2)

        th_box = QtWidgets.QWidget()
        th_box_layout = QtWidgets.QHBoxLayout(th_box)
        th_box_layout.setContentsMargins(0, 0, 0, 0)
        th_box_layout.setSpacing(3)
        th_label = QtWidgets.QLabel("しきい値")
        th_label.setStyleSheet("font-size:12px;color:#00d9ff;font-weight:bold;")
        self.threshold_edit = QtWidgets.QLineEdit()
        self.threshold_edit.setStyleSheet("QLineEdit{background:#08121c;color:#8ff6ff;border:1px solid #00a9d6;padding:2px 6px;border-radius:4px;font-weight:bold;}")
        self.threshold_edit.setValidator(QtGui.QDoubleValidator(0.0, 20.0, 1, self.threshold_edit))
        self.threshold_edit.setPlaceholderText("0.0-20.0")
        self.threshold_edit.returnPressed.connect(self._on_threshold_enter_pressed)
        th_box_layout.addWidget(th_label)
        th_box_layout.addWidget(self.threshold_edit, 1)
        level_row.addWidget(th_box, 1)
        right.addSpacing(5)
        right.addLayout(level_row)
        right.addSpacing(5)

        count_row = QtWidgets.QHBoxLayout()
        count_row.setContentsMargins(0, 0, 0, 0)
        count_row.setSpacing(1)
        self.ltor_card = QtWidgets.QLabel("")
        self.rtol_card = QtWidgets.QLabel("")
        for card in (self.ltor_card, self.rtol_card):
            card.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            card.setTextFormat(QtCore.Qt.TextFormat.RichText)
            card.setStyleSheet("font-size:10px;font-weight:700;background:#071925;border:1px solid #14b6dc;color:#98f5ff;border-radius:6px;padding:1px;")
            card.setFixedHeight(44)
            card.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)
            count_row.addWidget(card, 1)
        right.addLayout(count_row)

        self.wakimura_cards: dict[str, QtWidgets.QLabel] = {}
        self.wakimura_frames: dict[str, QtWidgets.QFrame] = {}
        self.wakimura_row: QtWidgets.QWidget | None = None
        if self.is_king:
            wakimura_card_height = 74
            self.wakimura_row = QtWidgets.QWidget()
            self.wakimura_row.setFixedHeight(wakimura_card_height)
            self.wakimura_row.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)
            wak_row_layout = QtWidgets.QHBoxLayout(self.wakimura_row)
            wak_row_layout.setContentsMargins(0, 0, 0, 0)
            wak_row_layout.setSpacing(1)
            for key, title in (
                ("alpha", "運用効率指標"),
                ("stay", "高付加判定"),
                ("exit", "流れ評価[WIN]"),
                ("time", "滞留評価[HL]"),
            ):
                card = QtWidgets.QFrame()
                card.setStyleSheet("background:#081225;border:1px solid #6a4cff;border-radius:6px;")
                card.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)
                card.setFixedHeight(wakimura_card_height)
                card_layout = QtWidgets.QVBoxLayout(card)
                card_layout.setContentsMargins(3, 2, 3, 2)
                card_layout.setSpacing(0)
                title_label = QtWidgets.QLabel(title)
                title_label.setStyleSheet("font-size:12px;color:#c6afff;font-weight:800;")
                title_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                value_label = QtWidgets.QLabel("--")
                value_label.setStyleSheet("font-size:11px;color:#d8c8ff;font-weight:500;")
                value_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignTop)
                value_label.setWordWrap(True)
                card_layout.addWidget(title_label)
                card_layout.addWidget(value_label, 1)
                wak_row_layout.addWidget(card, 1)
                self.wakimura_cards[key] = value_label
                self.wakimura_frames[key] = card

        right.addSpacing(5)
        self.stay_head = QtWidgets.QLabel("1分以上滞在ID")
        self.stay_head.setStyleSheet("font-size:11px;font-weight:bold;color:#9de7ff;padding:0px;margin:0px;")
        self.stay_head.setContentsMargins(0, 0, 0, 0)
        right.addWidget(self.stay_head)
        self.stay_grid = QtWidgets.QGridLayout()
        self.stay_grid.setContentsMargins(0, 0, 0, 0)
        self.stay_grid.setHorizontalSpacing(2)
        self.stay_grid.setVerticalSpacing(1)
        self.stay_grid.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignTop)
        self.stay_box = QtWidgets.QWidget()
        self.stay_box.setLayout(self.stay_grid)
        self.stay_box.setStyleSheet("background:#07131f;border:1px solid #145c7a;border-radius:6px;")
        self.stay_box.setSizePolicy(QtWidgets.QSizePolicy.Policy.Preferred, QtWidgets.QSizePolicy.Policy.Minimum)
        self.stay_box.setMinimumHeight(44)
        right.addWidget(self.stay_box)
        right.addSpacing(5)
        if self.is_king and self.wakimura_row is not None:
            right.addWidget(self.wakimura_row)
        self._render_stay_cards([])
        self._update_count_cards(0, 0)
        # UI余白圧縮はディスプレイ内に収めるための調整。

        top_row.addWidget(self.right_box, 0, QtCore.Qt.AlignmentFlag.AlignTop)
        top_row.setStretch(0, 0)
        top_row.setStretch(1, 1)
        root.addLayout(top_row)

        self.graphs: list[CombinedTimelineGraph] = []
        graphs_box = QtWidgets.QWidget()
        graphs_layout = QtWidgets.QVBoxLayout(graphs_box)
        graphs_layout.setContentsMargins(0, 1, 0, 0)
        graphs_layout.setSpacing(3)
        for _ in range(3):
            g = CombinedTimelineGraph("line")
            g.setFixedHeight(75)
            self.graphs.append(g)
            graphs_layout.addWidget(g)
        root.addWidget(graphs_box)
        self._update_title()
        self.threshold_edit.setText(f"{float(self.camera_cfg.get('congestion_threshold', 5.0)):.1f}")

    def update_view(self, payload: dict[str, Any]) -> None:
        frame = payload.get("frame")
        if frame is not None:
            h, w, _ = frame.shape
            src_w = payload.get("source_frame_width")
            src_h = payload.get("source_frame_height")
            if src_w and src_h:
                self._last_frame_size = (int(src_w), int(src_h))
            qimg = QtGui.QImage(frame.data, w, h, frame.strides[0], QtGui.QImage.Format.Format_BGR888)
            self._latest_pixmap = QtGui.QPixmap.fromImage(qimg)
            self._update_video_pixmap()

        score = float(payload.get("congestion_score", 0))
        threshold = float(payload.get("threshold", 5))
        status_raw = str(payload.get("status", ""))
        self._status_connected = status_raw == "RUNNING"

        ltor = payload.get("pass_bins_ltor", [0] * 144)
        rtol = payload.get("pass_bins_rtol", [0] * 144)
        self._update_title(payload.get("camera_name"), payload.get("stream_name"))
        if not self.threshold_edit.hasFocus():
            self.threshold_edit.setText(f"{threshold:.1f}")
        self._update_count_cards(int(sum(ltor)), int(sum(rtol)))
        long_stay_list = payload.get("long_stay_list", [])
        signature = self._build_stay_signature(long_stay_list)
        if signature != self._last_stay_signature:
            self._last_stay_signature = signature
            self._render_stay_cards(long_stay_list)
        self._update_congestion_bar(score, threshold)
        if self.is_king:
            wak_alpha = float(payload.get("wakimura_alpha", 0.0))
            wak_mode = bool(payload.get("wakimura_high_load_mode", False))
            self._update_wakimura_card(payload, wak_alpha, wak_mode)
        now_ts = time.time()
        graph_revision_ts = float(payload.get("graph_revision_ts", 0.0))
        should_update_line = (
            graph_revision_ts > self._last_graph_revision_ts
            and (now_ts - self.last_graph_update_ts >= self.graph_update_interval_sec)
        )
        hist_revision = (int(sum(ltor)), int(sum(rtol)))
        should_update_hist = hist_revision != self._last_hist_revision
        if should_update_line:
            self.graphs[0].set_line_data(payload.get("prev_congestion_points", []), payload.get("congestion_points", []), "渋滞指数", threshold=threshold, show_threshold=True)
            self._last_graph_revision_ts = graph_revision_ts
            self.last_graph_update_ts = now_ts
        if should_update_hist:
            self.graphs[1].set_bar_data(payload.get("hist_prev_ltor", [0] * 144), ltor, "LtoR")
            self.graphs[2].set_bar_data(payload.get("hist_prev_rtol", [0] * 144), rtol, "RtoL")
            self._last_hist_revision = hist_revision

    def resizeEvent(self, event: QtGui.QResizeEvent) -> None:
        super().resizeEvent(event)
        self._update_video_pixmap()
        self._relayout_stay_cards()

    def _update_congestion_bar(self, score: float, threshold: float) -> None:
        self.congestion_bar.set_values(score, threshold)

    def set_status(self, status_text: str) -> None:
        self._status_connected = status_text == "RUNNING"
        self._update_title()

    def _update_video_pixmap(self) -> None:
        if self._latest_pixmap is None:
            return
        frame_w, frame_h = self._last_frame_size
        if not frame_w or not frame_h:
            frame_w = max(1, self._latest_pixmap.width())
            frame_h = max(1, self._latest_pixmap.height())
        current_display_h = max(1, self.max_video_height)
        display_h = max(1, int(current_display_h * 1.1))
        aspect = frame_w / frame_h if frame_h else 1.0
        display_w = max(1, int(display_h * aspect))
        self.video_box.setFixedWidth(display_w)
        self.video_box.setFixedHeight(display_h)
        self.video.setFixedSize(display_w, display_h)
        self.right_box.setFixedHeight(display_h)
        qimg = self._latest_pixmap.toImage()
        pix = QtGui.QPixmap.fromImage(qimg).scaled(
            display_w,
            display_h,
            QtCore.Qt.AspectRatioMode.IgnoreAspectRatio,
            QtCore.Qt.TransformationMode.SmoothTransformation,
        )
        self.video.setPixmap(pix)

    def _update_title(self, camera_name: str | None = None, stream_name: str | None = None) -> None:
        display_name_map = {2: "KING", 1: "QUEEN", 3: "JACK"}
        fallback_name = display_name_map.get(self.camera_id, f"Camera{self.camera_id}")
        cam_name = display_name_map.get(self.camera_id) or camera_name or self.camera_cfg.get("camera_name", fallback_name)
        raw_stream = stream_name or self.camera_cfg.get("stream_name") or self.camera_cfg.get("stream_url")
        stream = str(raw_stream).strip() if raw_stream else f"stream{self.camera_id}"
        status_text = "受信中" if self._status_connected else "通信無し"
        title_text = f"{cam_name} ｜ {stream} ｜ {status_text}"
        if self.title.width() > 8:
            metrics = self.title.fontMetrics()
            title_text = metrics.elidedText(title_text, QtCore.Qt.TextElideMode.ElideRight, self.title.width())
        self.title.setText(title_text)

    def _on_threshold_enter_pressed(self) -> None:
        current = float(self.camera_cfg.get("congestion_threshold", 5.0))
        text = self.threshold_edit.text().strip()
        if not text:
            self.threshold_edit.setText(f"{current:.1f}")
            return
        try:
            value = float(text)
        except ValueError:
            self.threshold_edit.setText(f"{current:.1f}")
            return
        value = max(0.0, min(20.0, value))
        self.camera_cfg["congestion_threshold"] = float(value)
        self.threshold_edit.setText(f"{value:.1f}")
        self._update_congestion_bar(self.congestion_bar.score, float(value))
        self.threshold_changed.emit(self.camera_id, float(value))

    def _update_count_cards(self, ltor_total: int, rtol_total: int) -> None:
        self.ltor_card.setText(
            f"<div style='text-align:center;line-height:1.05em;'><span style='font-size:10px;font-weight:700;'>LtoR</span><br><span style='font-size:22px;font-weight:900;'>{ltor_total}</span></div>"
        )
        self.rtol_card.setText(
            f"<div style='text-align:center;line-height:1.05em;'><span style='font-size:10px;font-weight:700;'>RtoL</span><br><span style='font-size:22px;font-weight:900;'>{rtol_total}</span></div>"
        )

    def _render_stay_cards(self, entries: list[list[float] | tuple[int, float]]) -> None:
        try:
            self._stay_entries = list(entries)
            visible_entries = self._stay_entries[:8]
            visible_ids: set[int] = set()
            existing_ids = set(self.stay_card_widgets.keys())
            for item in visible_entries:
                track_id = int(item[0])
                stay_mins_int = max(0, int(float(item[1])))
                visible_ids.add(track_id)
                card = self.stay_card_widgets.get(track_id)
                if card is None:
                    card = self._create_stay_card()
                    self.stay_card_widgets[track_id] = card
                self._update_stay_card(card, track_id, stay_mins_int)

            added_ids = visible_ids - existing_ids
            removed_ids = set(self.stay_card_widgets.keys()) - visible_ids
            for track_id in removed_ids:
                card = self.stay_card_widgets.pop(track_id, None)
                if card is None:
                    continue
                self.stay_grid.removeWidget(card)
                card.deleteLater()

            if self.stay_empty_label is None:
                self.stay_empty_label = QtWidgets.QLabel("該当なし")
                self.stay_empty_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                self.stay_empty_label.setStyleSheet("font-size:11px;font-weight:700;color:#7dc7df;padding:0px;")
                self.stay_empty_label.setFixedHeight(36)
            self.stay_empty_label.setVisible(not visible_entries)

            if len(visible_entries) != self._last_stay_visible_count or added_ids or removed_ids:
                logging.debug(
                    "[DEBUG] cam%s stay cards visible=%d added=%s removed=%s",
                    self.camera_id,
                    len(visible_entries),
                    sorted(added_ids),
                    sorted(removed_ids),
                )
                self._last_stay_visible_count = len(visible_entries)

            self._relayout_stay_cards()
        except Exception as exc:
            logging.exception("[ERROR] cam%s stay card render failed: %s", self.camera_id, exc)

    def _create_stay_card(self) -> QtWidgets.QWidget:
        card = QtWidgets.QLabel("")
        card.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        card.setTextFormat(QtCore.Qt.TextFormat.RichText)
        card.setFixedHeight(36)
        return card

    def _update_stay_card(self, card: QtWidgets.QLabel, track_id: int, stay_mins_int: int) -> None:
        border_color = "#16b8d8"
        if stay_mins_int >= 20:
            border_color = "#ff4d4d"
        elif stay_mins_int >= 10:
            border_color = "#ffe066"
        card.setText(
            f"<div style='text-align:center;line-height:1.05em;'><span style='font-size:11px;font-weight:800;'>ID{track_id:03d}</span><br><span style='font-size:13px;font-weight:900;'>{stay_mins_int:d}分</span></div>"
        )
        card.setStyleSheet(
            "font-size:10px;"
            "background:#061726;"
            f"border:1px solid {border_color};"
            "border-radius:4px;"
            "color:#95f6ff;"
            "padding:0px;"
            "font-weight:bold;"
            "line-height:1.05em;"
        )

    def _relayout_stay_cards(self) -> None:
        cols = 8
        spacing = max(0, self.stay_grid.horizontalSpacing())
        contents = self.stay_grid.contentsMargins()
        box_w = max(self.stay_box.width(), self.stay_box.sizeHint().width(), 528)
        available_w = max(
            8,
            box_w - contents.left() - contents.right() - spacing * (cols - 1),
        )
        card_w = max(50, available_w // cols)
        extra_px = max(0, available_w - card_w * cols)
        col_widths = [card_w + (1 if i < extra_px else 0) for i in range(cols)]

        visible_entries = self._stay_entries[:8]
        if not visible_entries:
            if self.stay_empty_label is not None:
                self.stay_grid.removeWidget(self.stay_empty_label)
                self.stay_empty_label.setFixedWidth(col_widths[0])
                self.stay_grid.addWidget(self.stay_empty_label, 0, 0)
                self.stay_empty_label.show()
            self.stay_box.setFixedHeight(36)
            return

        if self.stay_empty_label is not None:
            self.stay_grid.removeWidget(self.stay_empty_label)
            self.stay_empty_label.hide()

        for idx, item in enumerate(visible_entries):
            track_id = int(item[0])
            card = self.stay_card_widgets.get(track_id)
            if card is None:
                continue
            col = idx % cols
            self.stay_grid.removeWidget(card)
            card.setFixedWidth(col_widths[col])
            self.stay_grid.addWidget(card, idx // cols, col)

        self.stay_box.setFixedHeight(36)

    def _build_stay_signature(self, entries: list[list[float] | tuple[int, float]]) -> tuple[tuple[int, int], ...]:
        signature: list[tuple[int, int]] = []
        for item in entries[:8]:
            signature.append((int(item[0]), max(0, int(float(item[1])))))
        return tuple(signature)

    def _update_wakimura_card(self, payload: dict[str, Any], wak_alpha: float, wak_mode: bool) -> None:
        if not self.is_king:
            return
        nt = int(payload.get("wak_total_tracks", 0))
        n_out = int(payload.get("wakimura_n_out", 0))
        avg_stay_sec = max(0, int(round(float(payload.get("wakimura_avg_stay_sec", 0.0)))))
        base_stay = int(round(float(self.camera_cfg.get("wakimura_base_stay_time_sec", 60.0))))
        mode_text = "HL" if wak_mode else "WIN"
        for key, frame in self.wakimura_frames.items():
            if key in ("alpha", "time") and wak_mode:
                card_border = "#ff944d"
            else:
                card_border = "#6a4cff"
            frame.setStyleSheet(f"background:#081225;border:1px solid {card_border};border-radius:6px;")
        for label in self.wakimura_cards.values():
            label.setStyleSheet("font-size:11px;color:#dfd4ff;font-weight:500;")

        self.wakimura_cards["alpha"].setText(
            f"α＝{wak_alpha:.2f}[{mode_text}]\n判定基準：α＜0.85\n評価期間：5分"
        )
        self.wakimura_cards["stay"].setText(
            f"現在滞在台数：{nt}台\n判定基準：7台以上\n判定：[WIN] or [HL]"
        )
        self.wakimura_cards["exit"].setText(
            f"基準退出台数(5min)：50台\n現在退出状況(5min)：{n_out}台\nα[WIN]={n_out}/50={n_out/50.0:.2f}"
        )
        self.wakimura_cards["time"].setText(
            f"基準滞在時間：{base_stay}秒\n平均滞在時間：{avg_stay_sec}秒\nα[HL]={base_stay}/{max(1, avg_stay_sec)}={base_stay/max(1, avg_stay_sec):.2f}"
        )

    def _clear_layout(self, layout: QtWidgets.QLayout) -> None:
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            child_layout = item.layout()
            if child_layout is not None:
                self._clear_layout(child_layout)
            if widget is not None:
                widget.deleteLater()


def resolve_model_path(camera_cfg: dict[str, Any], system_cfg: dict[str, Any], root_dir: Path) -> Path:
    raw = system_cfg.get("model_path") or system_cfg.get("yolo_model") or system_cfg.get("YOLO_MODEL") or camera_cfg.get("yolo_model") or "yolo11m.pt"
    if not raw:
        raise FileNotFoundError("YOLO model is not configured. Set yolo_model / model_path in config.")
    p = Path(str(raw))
    if not p.is_absolute():
        candidates = [root_dir / p, root_dir / "models" / p, Path.cwd() / p]
        for c in candidates:
            if c.exists():
                return c
        raise FileNotFoundError(
            f"YOLO model file not found: {raw}. "
            "Place the model locally under 04_ai_monitor or 04_ai_monitor/models."
        )
    if not p.exists():
        raise FileNotFoundError(f"YOLO model file not found: {p}")
    return p


# =========================================
# Camera Worker
# =========================================
class CameraWorker(QtCore.QObject):
    frame_ready = QtCore.pyqtSignal(dict)
    status_changed = QtCore.pyqtSignal(int, str)
    error_occurred = QtCore.pyqtSignal(int, str)
    finished = QtCore.pyqtSignal(int)

    def __init__(self, camera_cfg: dict[str, Any], system_cfg: dict[str, Any], root_dir: Path):
        super().__init__()
        self.camera_cfg = camera_cfg
        self.system_cfg = system_cfg
        self.root_dir = root_dir
        self.camera_id = int(camera_cfg["camera_id"])
        self.camera_name = camera_cfg["camera_name"]

        self.device = self._resolve_device(system_cfg.get("device_preference", "auto"))
        self.gpu_name = torch.cuda.get_device_name(0) if self.device.startswith("cuda") else "CPU"

        self.model_path = resolve_model_path(self.camera_cfg, self.system_cfg, self.root_dir)
        try:
            self.model = YOLO(str(self.model_path))
        except Exception as exc:
            raise RuntimeError(f"[ERROR] cam{self.camera_id} model load failed: {exc}") from exc
        self.target_classes = set(int(x) for x in self.camera_cfg.get("target_classes", [2, 3, 5, 7]))

        line = [self.camera_cfg.get("line_start", [0, 0]), self.camera_cfg.get("line_end", [100, 0])]
        self.counter = LineCounter(line)
        self.congestion = CongestionScorer(
            int(self.camera_cfg.get("congestion_calculation_interval", 3)),
            smoothing_window=int(self.system_cfg.get("congestion_smoothing_window", 6)),
        )
        # 脇村指標 α は参考表示専用で、既存 LEVEL 判定ロジックには一切使用しない。
        # 算出ロジックは app/90_sample/20_rotary_efficiency_analysis.py の考え方を監視向けに簡略適用。
        self.wakimura_alpha = WakimuraAlphaCalculator(
            rotary_capacity=int(self.camera_cfg.get("wakimura_rotary_capacity", 10)),
            base_stay_time_sec=float(self.camera_cfg.get("wakimura_base_stay_time_sec", 60.0)),
            window_seconds=int(self.camera_cfg.get("wakimura_window_seconds", 300)),
            high_load_vehicle_threshold=int(self.camera_cfg.get("wakimura_high_load_vehicle_threshold", 7)),
        )
        self.track_state = TrackState()
        self.track_class_memory: dict[int, dict[str, Any]] = {}
        self.display_id_map: dict[int, int] = {}
        self.display_id_counter = 1
        self.cross_flash_frames: dict[int, int] = {}

        self.cap = None
        self.last_frame = None
        self.last_raw_frame = None
        self.fps = 0.0
        self.frame_index = 0
        self._running = True
        self._next_reconnect_time = 0.0
        self._next_infer_retry_time = 0.0
        self._last_warn_emit = 0.0
        self._status_text = "INIT"
        self.display_scale = float(self.camera_cfg.get("display_scale", 0.8))
        self.read_fail_count = 0
        self.max_read_fail_before_reconnect = 5
        self.last_reconnect_at = 0.0
        self.reconnect_fail_count = 0
        self.max_graph_points = 1500
        self.last_wakimura_update_ts = 0.0
        self.wakimura_update_interval_sec = 1.0
        self.cached_wakimura_payload = self._default_wakimura_payload()
        # old版寄りに戻すため、推論の時間間引きを停止（frame_skip のみ有効）
        self.infer_interval_sec = 0.0
        self.last_infer_ts = 0.0
        self.last_graph_revision_ts = 0.0

        self.today = datetime.now().date()
        self.metrics_dir = root_dir / "data" / "metrics" / f"cam{self.camera_id}"
        self.metrics_dir.mkdir(parents=True, exist_ok=True)
        self.daily_report_dir = root_dir / "data" / "reports" / "daily"
        self.previous_day_hist_ltor, self.previous_day_hist_rtol = self._load_previous_day_histogram()
        self.previous_day_congestion_points = self._load_previous_day_congestion_points()
        self.boot_today_hist_ltor, self.boot_today_hist_rtol = self._load_today_histogram()
        self.boot_today_congestion_points = self._load_today_congestion_points()
        self.counter.state.pass_bins_ltor = list(self.boot_today_hist_ltor)
        self.counter.state.pass_bins_rtol = list(self.boot_today_hist_rtol)
        if self.boot_today_congestion_points:
            self.last_graph_revision_ts = time.time()

    def get_latest_raw_frame(self):
        return None if self.last_raw_frame is None else self.last_raw_frame.copy()

    def update_camera_config(self, new_cfg: dict[str, Any]) -> None:
        prev_cfg = dict(self.camera_cfg)
        old_model_path = self.model_path
        self.camera_cfg.update(new_cfg)
        self.camera_name = new_cfg.get("camera_name", self.camera_name)
        self.target_classes = set(int(x) for x in new_cfg.get("target_classes", [2, 3, 5, 7]))
        self.counter.update_line([new_cfg.get("line_start", [0, 0]), new_cfg.get("line_end", [100, 0])])
        self.congestion.update_interval(int(new_cfg.get("congestion_calculation_interval", 3)))
        self.congestion.update_smoothing_window(int(self.system_cfg.get("congestion_smoothing_window", 6)))
        self.display_scale = float(self.camera_cfg.get("display_scale", 0.8))

        new_model_path = resolve_model_path(self.camera_cfg, self.system_cfg, self.root_dir)
        if new_model_path != old_model_path:
            try:
                new_model = YOLO(str(new_model_path))
                self.model = new_model
                self.model_path = new_model_path
            except Exception as exc:
                self.camera_cfg = prev_cfg
                raise RuntimeError(f"[ERROR] cam{self.camera_id} model load failed: {exc}") from exc

    def _resolve_device(self, preference: str) -> str:
        if preference == "cpu":
            return "cpu"
        if torch.cuda.is_available():
            return "cuda:0"
        return "cpu"

    def _stream_source(self):
        stream_url = str(self.camera_cfg.get("stream_url", "0"))
        return int(stream_url) if stream_url.isdigit() else stream_url

    def connect(self) -> None:
        self._release_capture()
        src = self._stream_source()
        try:
            self.cap = cv2.VideoCapture(src, cv2.CAP_FFMPEG)
        except Exception:
            self.cap = cv2.VideoCapture(src)
        if self.cap is not None:
            try:
                self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            except Exception:
                pass
        self.read_fail_count = 0

    def _reconnect_capture(self, reason: str = "") -> None:
        now_ts = time.time()
        reconnect_sec = float(self.camera_cfg.get("reconnect_sec", 3))
        if now_ts - self.last_reconnect_at < reconnect_sec:
            return
        self.last_reconnect_at = now_ts
        self._set_status("RECONNECTING")
        self.error_occurred.emit(self.camera_id, f"[INFO] cam{self.camera_id} reconnect start: {reason or '-'}")
        self._release_capture()
        time.sleep(min(1.0, reconnect_sec))
        try:
            self.connect()
            self._set_status("RUNNING")
            self.reconnect_fail_count = 0
            self.error_occurred.emit(self.camera_id, f"[INFO] cam{self.camera_id} reconnect success")
        except Exception as exc:
            self.reconnect_fail_count += 1
            if self.reconnect_fail_count <= 3 or self.reconnect_fail_count % 10 == 0:
                self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} reconnect failed({self.reconnect_fail_count}): {exc}")

    def _release_capture(self) -> None:
        if self.cap is not None:
            try:
                self.cap.release()
            except Exception:
                pass
            finally:
                self.cap = None

    def _set_status(self, text: str) -> None:
        if self._status_text == text:
            return
        self._status_text = text
        self.status_changed.emit(self.camera_id, text)

    def _in_polygon(self, point: tuple[float, float], polygon_points: list[list[int]]) -> bool:
        if not polygon_points:
            return False
        poly = np.array(polygon_points, np.int32)
        return cv2.pointPolygonTest(poly, point, False) >= 0

    def _load_daily_data_sheet(self, target_date: date) -> pd.DataFrame:
        path = self.daily_report_dir / f"{target_date.isoformat()}.xlsx"
        if not path.exists():
            return pd.DataFrame()
        try:
            df = pd.read_excel(path, sheet_name="Data")
        except Exception:
            return pd.DataFrame()
        return df if not df.empty else pd.DataFrame()

    def _default_wakimura_payload(self) -> dict[str, Any]:
        return {
            "wakimura_alpha": 0.0,
            "wakimura_n_out": 0,
            "wakimura_avg_stay_sec": 0.0,
            "wakimura_high_load_mode": False,
            "wak_total_tracks": 0,
            "wak_max_stay_min": 0.0,
        }

    def _load_previous_day_histogram(self) -> tuple[list[int], list[int]]:
        prev = self.today.fromordinal(self.today.toordinal() - 1)
        ltor = [0] * 144
        rtol = [0] * 144
        df = self._load_daily_data_sheet(prev)
        if df.empty:
            return ltor, rtol
        for _, row in df.iterrows():
            try:
                hhmm = str(row.get("時刻（10分単位）", "")).strip()
                dt = datetime.strptime(hhmm, "%H:%M")
                idx = (dt.hour * 60 + dt.minute) // 10
                if 0 <= idx < 144:
                    ltor[idx] = int(row.get(f"Camera{self.camera_id} LtoR", 0) or 0)
                    rtol[idx] = int(row.get(f"Camera{self.camera_id} RtoL", 0) or 0)
            except Exception:
                continue
        return ltor, rtol

    def _load_previous_day_congestion_points(self) -> list[tuple[datetime, float]]:
        prev = self.today.fromordinal(self.today.toordinal() - 1)
        points: list[tuple[datetime, float]] = []
        df = self._load_daily_data_sheet(prev)
        if df.empty:
            return points
        for _, row in df.iterrows():
            try:
                hhmm = str(row.get("時刻（10分単位）", "")).strip()
                ts = datetime.combine(prev, datetime.strptime(hhmm, "%H:%M").time())
                score = float(row.get(f"Camera{self.camera_id} 渋滞指数", 0.0) or 0.0)
                points.append((ts, score))
            except Exception:
                continue
        return points

    @staticmethod
    def _parse_hhmm_to_bin_index(value: Any) -> int | None:
        try:
            hhmm = str(value).strip()
            dt = datetime.strptime(hhmm, "%H:%M")
            idx = (dt.hour * 60 + dt.minute) // 10
            return idx if 0 <= idx < 144 else None
        except Exception:
            return None

    def _load_today_histogram(self) -> tuple[list[int], list[int]]:
        ltor = [0] * 144
        rtol = [0] * 144
        df = self._load_daily_data_sheet(self.today)
        if df.empty:
            return ltor, rtol
        ltor_col = f"Camera{self.camera_id} LtoR"
        rtol_col = f"Camera{self.camera_id} RtoL"
        for _, row in df.iterrows():
            idx = self._parse_hhmm_to_bin_index(row.get("時刻（10分単位）", ""))
            if idx is None:
                continue
            try:
                ltor[idx] = int(row.get(ltor_col, 0) or 0)
                rtol[idx] = int(row.get(rtol_col, 0) or 0)
            except Exception:
                continue
        return ltor, rtol

    def _load_today_congestion_points(self) -> list[tuple[datetime, float]]:
        points: list[tuple[datetime, float]] = []
        df = self._load_daily_data_sheet(self.today)
        if df.empty:
            return points
        score_col = f"Camera{self.camera_id} 渋滞指数"
        for _, row in df.iterrows():
            try:
                hhmm = str(row.get("時刻（10分単位）", "")).strip()
                ts = datetime.combine(self.today, datetime.strptime(hhmm, "%H:%M").time())
                score = float(row.get(score_col, 0.0) or 0.0)
                points.append((ts, score))
            except Exception:
                continue
        return points

    def _rollover_if_needed(self, now: datetime) -> None:
        if now.date() == self.today:
            return
        self.today = now.date()
        self.display_id_map.clear()
        self.display_id_counter = 1
        self.cross_flash_frames.clear()
        self.previous_day_hist_ltor, self.previous_day_hist_rtol = self._load_previous_day_histogram()
        self.previous_day_congestion_points = self._load_previous_day_congestion_points()
        self.boot_today_hist_ltor, self.boot_today_hist_rtol = self._load_today_histogram()
        self.boot_today_congestion_points = self._load_today_congestion_points()
        self.counter.state.pass_bins_ltor, self.counter.state.pass_bins_rtol = (
            list(self.boot_today_hist_ltor),
            list(self.boot_today_hist_rtol),
        )
        self.congestion.state.frame_time_stamps = []
        self.congestion.state.frame_motion_scores = []
        self.congestion.state.smoothed_motion_scores = []
        self.congestion.smoother = CongestionSmoother(int(self.system_cfg.get("congestion_smoothing_window", 6)))
        self.congestion.state.current_congestion_index = 0.0
        self.congestion.state.current_smoothed_index = 0.0
        self.last_graph_revision_ts = time.time() if self.boot_today_congestion_points else 0.0

    def _get_display_id(self, track_id: int) -> int:
        if track_id not in self.display_id_map:
            self.display_id_map[track_id] = self.display_id_counter
            self.display_id_counter += 1
        return self.display_id_map[track_id]

    @staticmethod
    def _aggregate_points_by_minute(points: list[tuple[datetime, float]]) -> list[tuple[datetime, float]]:
        if not points:
            return []
        minute_buckets: dict[datetime, list[float]] = {}
        for ts, value in points:
            minute_ts = ts.replace(second=0, microsecond=0)
            minute_buckets.setdefault(minute_ts, []).append(float(value))
        return sorted(
            [(minute_ts, float(np.mean(values))) for minute_ts, values in minute_buckets.items()],
            key=lambda x: x[0],
        )

    @staticmethod
    def _bbox_iou(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> float:
        ax1, ay1, ax2, ay2 = a
        bx1, by1, bx2, by2 = b
        inter_x1, inter_y1 = max(ax1, bx1), max(ay1, by1)
        inter_x2, inter_y2 = min(ax2, bx2), min(ay2, by2)
        iw = max(0, inter_x2 - inter_x1)
        ih = max(0, inter_y2 - inter_y1)
        inter = iw * ih
        area_a = max(1, (ax2 - ax1) * (ay2 - ay1))
        area_b = max(1, (bx2 - bx1) * (by2 - by1))
        union = max(1, area_a + area_b - inter)
        return inter / union

    def _deduplicate_overlapping_detections(self, raw_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        priority = {"truck": 0, "bus": 1, "car": 2, "motorcycle": 3, "bicycle": 4, "person": 5}
        used = [False] * len(raw_items)
        kept: list[dict[str, Any]] = []
        conf_eps = 0.05
        for i, item in enumerate(raw_items):
            if used[i]:
                continue
            cluster = [item]
            used[i] = True
            x1, y1, x2, y2 = item["bbox"]
            base_w = max(1.0, float(x2 - x1))
            base_h = max(1.0, float(y2 - y1))
            for j in range(i + 1, len(raw_items)):
                if used[j]:
                    continue
                cand = raw_items[j]
                iou = self._bbox_iou(item["bbox"], cand["bbox"])
                cx1, cy1 = item["center"]
                cx2, cy2 = cand["center"]
                dist = ((cx1 - cx2) ** 2 + (cy1 - cy2) ** 2) ** 0.5
                if iou >= 0.65 or dist <= min(base_w, base_h) * 0.35:
                    cluster.append(cand)
                    used[j] = True
            cluster.sort(key=lambda x: (-float(x.get("conf", 0.0)), priority.get(str(x.get("cls_name", "")).lower(), 999)))
            best = cluster[0]
            tid = int(best.get("track_id", -1))
            if tid >= 0:
                mem = self.track_class_memory.get(tid)
                if mem is not None:
                    prev_cls = str(mem.get("cls_name", best["cls_name"]))
                    prev_ts = mem.get("ts", datetime.min)
                    prev_area = float(mem.get("area", 1.0))
                    new_area = max(1.0, (best["bbox"][2] - best["bbox"][0]) * (best["bbox"][3] - best["bbox"][1]))
                    if (datetime.now() - prev_ts).total_seconds() <= 3.0 and abs(float(best.get("conf", 0)) - float(mem.get("conf", 0))) <= conf_eps:
                        if prev_cls != best["cls_name"] and 0.55 <= (new_area / max(1.0, prev_area)) <= 1.8:
                            best["cls_name"] = prev_cls
                self.track_class_memory[tid] = {"cls_name": best["cls_name"], "ts": datetime.now(), "area": max(1.0, (best["bbox"][2] - best["bbox"][0]) * (best["bbox"][3] - best["bbox"][1])), "conf": float(best.get("conf", 0.0))}
            kept.append(best)
        return kept

    @QtCore.pyqtSlot()
    def run(self) -> None:
        self._set_status("RUNNING")
        while self._running:
            try:
                payload = self.process_once_nonblocking()
                if payload is not None:
                    # old版寄りに戻すため、UI送信間引きを廃止して即時反映
                    self.frame_ready.emit(payload)
            except Exception as exc:
                self.error_occurred.emit(self.camera_id, str(exc))
            QtCore.QThread.msleep(5)
        self._release_capture()
        self.finished.emit(self.camera_id)

    @QtCore.pyqtSlot()
    def stop(self) -> None:
        self._running = False
        self._release_capture()

    def process_once_nonblocking(self) -> dict[str, Any] | None:
        try:
            start = time.time()
            now_ts = time.time()
            if now_ts < self._next_reconnect_time or now_ts < self._next_infer_retry_time:
                return None

            try:
                if self.cap is None or not self.cap.isOpened():
                    self._reconnect_capture("cap not opened")
                    return None

                ok, frame = self.cap.read()
                if not ok or frame is None:
                    self.read_fail_count += 1
                    if self.read_fail_count >= self.max_read_fail_before_reconnect:
                        self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} frame read failed repeatedly. reconnecting.")
                        self._reconnect_capture("read failed")
                    return None
                self.read_fail_count = 0
            except cv2.error as exc:
                self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} OpenCV error: {exc}")
                self._reconnect_capture("cv2.error")
                return None
            except Exception as exc:
                self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} capture exception: {exc}")
                self._reconnect_capture("generic exception")
                return None

            self._set_status("RUNNING")
            src_h, src_w = frame.shape[:2]
            self.last_raw_frame = self._resize_for_display(frame.copy())
            now = datetime.now()
            self._rollover_if_needed(now)

            self.frame_index += 1
            frame_skip = max(1, int(self.camera_cfg.get("frame_skip", 1)))
            if self.frame_index % frame_skip != 0:
                return None

            now_ts = time.time()
            if self.infer_interval_sec > 0 and (now_ts - self.last_infer_ts < self.infer_interval_sec):
                return None
            self.last_infer_ts = now_ts

            try:
                result = self.model.track(
                    source=frame,
                    persist=True,
                    tracker="bytetrack.yaml",
                    verbose=False,
                    device=self.device,
                    conf=float(self.camera_cfg.get("confidence_threshold", 0.25)),
                    iou=float(self.camera_cfg.get("iou_threshold", 0.5)),
                    imgsz=int(self.camera_cfg.get("imgsz", 640)),
                    classes=sorted(self.target_classes) if self.target_classes else None,
                )[0]
            except Exception as exc:
                self._next_infer_retry_time = time.time() + 3.0
                self._set_status("MODEL ERROR")
                self.error_occurred.emit(self.camera_id, f"[ERROR] cam{self.camera_id} tracker failed: {exc}")
                return None

            boxes = result.boxes
            tracks: list[dict[str, Any]] = []
            raw_items: list[dict[str, Any]] = []
            pass_events: list[dict[str, Any]] = []
            long_stay_events: list[dict[str, Any]] = []
            long_stay_list: list[tuple[int, float]] = []
            exclude_polygon = self.camera_cfg.get("exclude_polygon", [])
            stay_zone = self.camera_cfg.get("stay_zone_polygon", [])

            if boxes is not None and boxes.id is not None:
                cls_array = boxes.cls.cpu().numpy() if boxes.cls is not None else []
                conf_array = boxes.conf.cpu().numpy() if boxes.conf is not None else []
                for i, (box, tid) in enumerate(zip(boxes.xyxy.cpu().numpy(), boxes.id.cpu().numpy())):
                    cls_idx = int(cls_array[i]) if len(cls_array) > i else -1
                    if self.target_classes and cls_idx not in self.target_classes:
                        continue
                    cls_name = self.model.names.get(cls_idx, str(cls_idx)) if isinstance(self.model.names, dict) else str(cls_idx)
                    conf = float(conf_array[i]) if len(conf_array) > i else 0.0

                    x1, y1, x2, y2 = box.tolist()
                    bbox_h = max(1.0, y2 - y1)
                    judge_x = (x1 + x2) / 2
                    judge_y = y2 - (bbox_h * 0.2)
                    if self._in_polygon((judge_x, judge_y), exclude_polygon):
                        continue

                    track_id = int(tid)
                    display_id = self._get_display_id(track_id)
                    raw_items.append({
                        "track_id": track_id,
                        "display_id": display_id,
                        "cls_idx": cls_idx,
                        "cls_name": cls_name,
                        "conf": conf,
                        "bbox": (int(x1), int(y1), int(x2), int(y2)),
                        "center": (judge_x, judge_y),
                        "judge_point": (judge_x, judge_y),
                    })

                for item in self._deduplicate_overlapping_detections(raw_items):
                    track_id = int(item["track_id"])
                    display_id = int(item.get("display_id", track_id))
                    cls_name = item["cls_name"]
                    judge_x, judge_y = item["center"]
                    tracks.append(
                        {
                            "track_id": track_id,
                            "display_id": display_id,
                            "center": (judge_x, judge_y),
                            "judge_point": (judge_x, judge_y),
                            "bbox": item["bbox"],
                            "class_name": cls_name,
                            "crossed": False,
                        }
                    )

                    event = self.counter.update(track_id, (judge_x, judge_y), cls_name, now)
                    if event:
                        pass_events.append(event)
                        self.wakimura_alpha.record_exit(now)
                        self.cross_flash_frames[track_id] = 12
                        tracks[-1]["crossed"] = True

                    if track_id not in self.track_state.first_seen:
                        self.track_state.first_seen[track_id] = now

                    in_stay_zone = self._in_polygon((judge_x, judge_y), stay_zone) if stay_zone else True
                    if in_stay_zone:
                        stay_mins = (now - self.track_state.first_seen[track_id]).total_seconds() / 60.0
                        if stay_mins >= float(self.camera_cfg.get("long_stay_minutes", 15)):
                            long_stay_list.append((display_id, stay_mins))
                            if track_id not in self.track_state.long_stay_emitted:
                                self.track_state.long_stay_emitted.add(track_id)
                                long_stay_events.append({
                                    "first_seen": self.track_state.first_seen[track_id].strftime("%Y-%m-%d %H:%M:%S"),
                                    "detected_at": now.strftime("%Y-%m-%d %H:%M:%S"),
                                    "camera_id": self.camera_id,
                                    "track_id": track_id,
                                    "stay_minutes": round(stay_mins, 1),
                                    "class_name": cls_name,
                                })
                    elif self.cross_flash_frames.get(track_id, 0) > 0:
                        tracks[-1]["crossed"] = True

            for tid in list(self.cross_flash_frames):
                self.cross_flash_frames[tid] -= 1
                if self.cross_flash_frames[tid] <= 0:
                    self.cross_flash_frames.pop(tid, None)

            frame_width = int(frame.shape[1]) if frame is not None else 1920
            prev_points_len = len(self.congestion.state.frame_time_stamps)
            congestion_score = self.congestion.update(tracks, now, frame_width) if self.camera_cfg.get("enable_congestion", True) else 0.0
            threshold = float(self.camera_cfg.get("congestion_threshold", 5))
            threshold_over = congestion_score >= threshold
            count_ltor = int(sum(self.counter.state.pass_bins_ltor))
            count_rtol = int(sum(self.counter.state.pass_bins_rtol))
            max_stay_min = max((mins for _, mins in long_stay_list), default=0.0)
            wakimura = self.cached_wakimura_payload
            if self.camera_id == 2 and (time.time() - self.last_wakimura_update_ts >= self.wakimura_update_interval_sec):
                avg_stay_sec = 0.0
                if tracks:
                    stay_secs = []
                    for tr in tracks:
                        seen = self.track_state.first_seen.get(int(tr["track_id"]))
                        if seen is not None:
                            stay_secs.append(max(0.0, (now - seen).total_seconds()))
                    avg_stay_sec = float(np.mean(stay_secs)) if stay_secs else 0.0
                try:
                    updated = self.wakimura_alpha.update(now=now, vehicle_count=len(tracks), avg_stay_sec=avg_stay_sec)
                    wakimura = {
                        "wakimura_alpha": float(updated.get("wakimura_alpha", 0.0)),
                        "wakimura_n_out": int(updated.get("wakimura_n_out", 0)),
                        "wakimura_avg_stay_sec": float(updated.get("wakimura_avg_stay_sec", 0.0)),
                        "wakimura_high_load_mode": bool(updated.get("wakimura_high_load_mode", False)),
                        "wak_total_tracks": int(len(tracks)),
                        "wak_max_stay_min": float(max_stay_min),
                    }
                except Exception as exc:
                    self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} 脇村指標 α update skipped: {exc}")
                    wakimura = self.cached_wakimura_payload
                self.cached_wakimura_payload = wakimura
                self.last_wakimura_update_ts = time.time()
            elif self.camera_id != 2:
                wakimura = self._default_wakimura_payload()

            elapsed = max(1e-6, time.time() - start)
            self.fps = 1.0 / elapsed

            if len(self.congestion.state.frame_time_stamps) > prev_points_len:
                self.last_graph_revision_ts = time.time()

            self.last_frame = self._resize_for_display(self._draw_overlay(frame.copy(), tracks))
            try:
                long_stay_list.sort(key=lambda x: x[1], reverse=True)
            except Exception as exc:
                self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} long_stay list sort skipped: {exc}")
                long_stay_list = []
            smoothed_score = float(self.congestion.state.current_smoothed_index)
            current_raw_points = list(zip(self.congestion.state.frame_time_stamps, self.congestion.state.frame_motion_scores))
            merged_points_by_ts: dict[datetime, float] = {}
            for ts, score in self.boot_today_congestion_points:
                merged_points_by_ts[ts] = float(score)
            for ts, score in current_raw_points:
                merged_points_by_ts[ts] = float(score)
            merged_today_points = sorted(merged_points_by_ts.items(), key=lambda x: x[0])
            curr_points = self._aggregate_points_by_minute(merged_today_points)[-self.max_graph_points:]
            prev_points = self._aggregate_points_by_minute(self.previous_day_congestion_points)[-self.max_graph_points:]

            payload = {
                "camera_id": self.camera_id,
                "camera_name": self.camera_name,
                "stream_name": self.camera_cfg.get("stream_name", self.camera_cfg.get("stream_url", "")),
                "frame": self.last_frame,
                "source_frame_width": int(src_w),
                "source_frame_height": int(src_h),
                "congestion_score": congestion_score,
                "smoothed_congestion_score": smoothed_score,
                "congestion_level": 1,
                "threshold": threshold,
                "threshold_over": threshold_over,
                "congestion_points": curr_points,
                "prev_congestion_points": prev_points,
                "pass_bins_ltor": self.counter.state.pass_bins_ltor,
                "pass_bins_rtol": self.counter.state.pass_bins_rtol,
                "count_ltor": count_ltor,
                "count_rtol": count_rtol,
                "hist_prev_ltor": self.previous_day_hist_ltor,
                "hist_prev_rtol": self.previous_day_hist_rtol,
                "graph_revision_ts": float(self.last_graph_revision_ts),
                "long_stay_count": len(long_stay_list),
                "long_stay_list": [[int(tid), float(minutes)] for tid, minutes in long_stay_list[:10]],
                **wakimura,
                "debug_metrics": {
                    "raw_congestion_index": round(float(congestion_score), 3),
                    "smoothed_congestion_index": round(smoothed_score, 3),
                    "level": 1,
                    "last_update_timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
                },
                "fps": self.fps,
                "status": self._status_text,
                "device": self.device,
                "gpu_name": self.gpu_name,
            }
            return payload
        except Exception as exc:
            self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} loop exception: {exc}")
            self._reconnect_capture("loop exception")
            return None

    def _resize_for_display(self, frame: np.ndarray) -> np.ndarray:
        if frame is None:
            return frame
        if self.display_scale >= 0.999:
            return frame
        h, w = frame.shape[:2]
        new_w = max(1, int(w * self.display_scale))
        new_h = max(1, int(h * self.display_scale))
        return cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_AREA)

    def _draw_overlay(self, frame: np.ndarray, tracks: list[dict[str, Any]]) -> np.ndarray:
        line = [self.camera_cfg.get("line_start", [10, 10]), self.camera_cfg.get("line_end", [100, 10])]
        cv2.line(frame, tuple(line[0]), tuple(line[1]), (0, 255, 255), 2)

        poly = self.camera_cfg.get("exclude_polygon", [])
        if len(poly) >= 3:
            cv2.polylines(frame, [np.array(poly, np.int32)], isClosed=True, color=(255, 255, 0), thickness=2)
            cv2.putText(frame, "EXCLUDE", tuple(np.array(poly[0], dtype=int)), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 0), 2, cv2.LINE_AA)

        for tr in tracks:
            x1, y1, x2, y2 = tr["bbox"]
            box_color = (0, 0, 255) if tr.get("crossed", False) or self.cross_flash_frames.get(int(tr["track_id"]), 0) > 0 else (0, 255, 0)
            cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)
            px, py = tr.get("judge_point", tr["center"])
            cv2.circle(frame, (int(px), int(py)), 3, (255, 255, 255), -1)
            cv2.putText(
                frame,
                f"ID:{int(tr.get('display_id', tr['track_id'])):03d} {tr['class_name']}",
                (x1, max(20, y1 - 8)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.55,
                box_color,
                2,
                cv2.LINE_AA,
            )

        return frame

# =========================================
# Main Window
# =========================================
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, root_dir: Path):
        super().__init__()
        self.root_dir = root_dir
        self.setWindowTitle("AI Congestion Monitor")
        self.setStyleSheet("background:#02060a;")
        self.resize(1220, 1600)
        self.setMinimumSize(1180, 1300)

        self.cfg_mgr = ConfigManager(root_dir)
        self.app_cfg = self.cfg_mgr.load()
        self.reporter = ReportWriter(root_dir / "data")
        self.ten_min_writer = TenMinuteRecordWriter(root_dir / "data")
        raw_ai_status = Path(self.app_cfg.system.get("ai_status_json_path", "app/11_config/ai_status.json"))
        script_base = Path(__file__).resolve().parents[2]
        if raw_ai_status.is_absolute():
            ai_status_path = raw_ai_status
        else:
            ai_status_path = script_base / raw_ai_status
        ai_status_path.parent.mkdir(parents=True, exist_ok=True)
        self.status_mgr = StatusManager(ai_status_path, self.app_cfg.system)

        self.threads: dict[int, QtCore.QThread] = {}
        self.workers: dict[int, CameraWorker] = {}
        self.panels: dict[int, CameraPanel] = {}
        self.latest_payloads: dict[int, dict[str, Any]] = {}
        self.ten_min_buffer: list[dict[str, Any]] = []
        now_dt = datetime.now()
        self.current_10min_bin_start = self.ten_min_writer._floor_to_10min(now_dt)
        self.current_10min_date = now_dt.date()
        self.system_level_history_today: list[tuple[datetime, float]] = []
        self.system_level_history_yesterday: list[tuple[datetime, float]] = []
        self.system_level_history_date = datetime.now().date()
        self.last_system_level_record_ts = 0.0

        toolbar = QtWidgets.QToolBar("Main", self)
        toolbar.setMovable(False)
        self.addToolBar(QtCore.Qt.ToolBarArea.TopToolBarArea, toolbar)
        action_setting = toolbar.addAction("解析条件")
        action_setting.triggered.connect(self.open_settings)
        action_save = toolbar.addAction("保存")
        action_save.triggered.connect(self.save_current_settings)
        action_load = toolbar.addAction("読込")
        action_load.triggered.connect(self.load_settings_from_json)
        action_daily = toolbar.addAction("日次Excel出力")
        action_daily.triggered.connect(self.export_daily)
        action_monthly = toolbar.addAction("月次Excel出力")
        action_monthly.triggered.connect(self.export_monthly)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.setCentralWidget(scroll)
        content = QtWidgets.QWidget()
        content.setMinimumWidth(1080)
        content.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred)
        layout = QtWidgets.QVBoxLayout(content)
        layout.setContentsMargins(6, 4, 6, 4)
        layout.setSpacing(14)
        scroll.setWidget(content)

        top_info_layout = QtWidgets.QVBoxLayout()
        top_info_layout.setContentsMargins(0, 0, 0, 0)
        top_info_layout.setSpacing(4)
        self.congestion_formula_label = QtWidgets.QLabel(
            "【渋滞指数】車両の停滞傾向を表す指標。移動量が小さい車ほど値が高くなり、流れが悪い状態を表す。\n"
            "例えば渋滞指標＝３であれば、停止している車が３台という目安になる。"
        )
        self.congestion_formula_label.setWordWrap(True)
        self.congestion_formula_label.setStyleSheet(
            "color:#9af2ff;background:#08121b;border:1px solid #1f4f7a;padding:4px;font-size:10px;"
        )
        self.wakimura_formula_label = QtWidgets.QLabel(
            "【運用効率指標α】どの程度効率的に車が流れているかを簡略化して表す指標。滞在台数と滞在時間を用いて算出する。\n"
            "α=1(100%)を理想状態とし、α<0.85(85%以下)となる状態を「効率低下」として判定する。（慶応SFC考案の指標）"
        )
        self.wakimura_formula_label.setWordWrap(True)
        self.wakimura_formula_label.setStyleSheet(
            "color:#d6cbff;background:#0d1020;border:1px solid #4b3cb0;padding:4px;font-size:10px;"
        )
        self.congestion_formula_label.setMinimumWidth(500)
        self.wakimura_formula_label.setMinimumWidth(500)
        self.congestion_formula_label.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred)
        self.wakimura_formula_label.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred)

        top_left_box = QtWidgets.QVBoxLayout()
        top_left_box.setSpacing(2)
        top_left_box.setContentsMargins(4, 4, 4, 4)
        top_left_widget = QtWidgets.QWidget()
        top_left_widget.setLayout(top_left_box)
        top_left_widget.setMinimumWidth(500)
        top_left_widget.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred)

        formula_stack = QtWidgets.QVBoxLayout()
        formula_stack.setContentsMargins(0, 0, 0, 0)
        formula_stack.setSpacing(3)
        formula_widget = QtWidgets.QWidget()
        formula_widget.setLayout(formula_stack)
        formula_stack.addWidget(self.congestion_formula_label)
        formula_stack.addWidget(self.wakimura_formula_label)

        level_block = QtWidgets.QVBoxLayout()
        level_block.setContentsMargins(0, 0, 0, 0)
        level_block.setSpacing(3)
        level_block_widget = QtWidgets.QWidget()
        level_block_widget.setLayout(level_block)
        level_block_widget.setMinimumWidth(520)
        level_block_widget.setMaximumWidth(520)
        level_block_widget.setSizePolicy(QtWidgets.QSizePolicy.Policy.Preferred, QtWidgets.QSizePolicy.Policy.Preferred)
        self.level_badge = QtWidgets.QLabel("🟢 渋滞LEVEL1")
        self.level_badge.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.level_badge.setMinimumWidth(500)
        self.level_badge.setMaximumWidth(500)
        self.level_badge.setMinimumHeight(72)
        self.level_badge.setSizePolicy(QtWidgets.QSizePolicy.Policy.Preferred, QtWidgets.QSizePolicy.Policy.Fixed)
        self.level_badge.setStyleSheet("background:#7fd0ff;color:#000000;border-radius:8px;font-weight:900;font-size:28px;padding:2px 4px;")
        self.system_title_ja = QtWidgets.QLabel("AI渋滞判定システム")
        self.system_title_ja.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.system_title_ja.setStyleSheet("font-size:30px;font-weight:900;color:#9fe8ff;")
        self.system_runtime_label = QtWidgets.QLabel("GPU: n/a ｜ model: n/a ｜ tracker: ByteTrack")
        self.system_runtime_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.system_runtime_label.setStyleSheet("font-size:11px;color:#9abed0;")
        top_left_box.addWidget(self.system_title_ja)
        top_left_box.addWidget(self.system_runtime_label)
        self.level_rule_label = QtWidgets.QLabel(
            "<b>LEVEL1：</b>通常時<br>"
            "<b>LEVEL2：</b>[KING]渋滞指数5以上 + 5分以上滞在台数3台以上<br>"
            "<b>LEVEL3：</b>[QUEEN]渋滞指数3以上<br>"
            "<b>LEVEL4：</b>[QUEEN]渋滞指数3以上 + [JACK]渋滞指数3以上"
        )
        self.level_rule_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignTop)
        self.level_rule_label.setWordWrap(True)
        self.level_rule_label.setTextFormat(QtCore.Qt.TextFormat.RichText)
        self.level_rule_label.setStyleSheet("color:#b7dbff;background:#0a1420;border:1px solid #1f4f7a;padding:4px;font-size:12px;line-height:1.25em;")
        self.level_rule_label.setMinimumWidth(520)
        self.level_rule_label.setMaximumWidth(520)
        self.level_rule_label.setSizePolicy(QtWidgets.QSizePolicy.Policy.Preferred, QtWidgets.QSizePolicy.Policy.Preferred)
        level_block.addWidget(self.level_badge)

        top_header_row = QtWidgets.QHBoxLayout()
        top_header_row.setContentsMargins(0, 0, 0, 0)
        top_header_row.setSpacing(6)
        top_header_row.addWidget(top_left_widget, 1)
        top_header_row.addWidget(level_block_widget, 0)
        top_info_layout.addLayout(top_header_row)

        top_detail_row = QtWidgets.QHBoxLayout()
        top_detail_row.setContentsMargins(0, 0, 0, 0)
        top_detail_row.setSpacing(6)
        top_detail_row.addWidget(formula_widget, 1)
        top_detail_row.addWidget(self.level_rule_label, 0)
        top_info_layout.addLayout(top_detail_row)
        layout.addLayout(top_info_layout)

        self.system_level_graph = CombinedTimelineGraph("line")
        self.system_level_graph.setFixedHeight(78)
        self.system_level_graph.set_y_axis_config(
            y_min=1.0,
            y_max=4.0,
            labels={1.0: "LEVEL1", 2.0: "LEVEL2", 3.0: "LEVEL3", 4.0: "LEVEL4"},
        )
        layout.addWidget(self.system_level_graph)
        layout.addSpacing(12)
        self.system_level_history_yesterday = self._load_system_level_history(self.system_level_history_date - timedelta(days=1))
        self.system_level_history_today = self._load_system_level_history(self.system_level_history_date)
        self.update_level_rule_text()
        self._refresh_system_level_graph()

        for cam in self.app_cfg.cameras:
            if not cam.get("enabled", True):
                continue
            panel = CameraPanel(cam)
            panel.line_setting_requested.connect(lambda cid, m="line": self.open_settings_for_camera(cid, m))
            panel.exclude_setting_requested.connect(lambda cid, m="poly": self.open_settings_for_camera(cid, m))
            panel.camera_setting_requested.connect(lambda cid, m="basic": self.open_settings_for_camera(cid, m))
            panel.threshold_changed.connect(self.on_threshold_changed)
            layout.addWidget(panel, 1)
            self.panels[cam["camera_id"]] = panel
            try:
                worker = CameraWorker(cam, self.app_cfg.system, self.root_dir)
            except Exception as exc:
                panel.set_status("MODEL ERROR")
                logging.exception("cam%s worker init failed: %s", cam.get("camera_id"), exc)
                continue
            thread = QtCore.QThread(self)
            worker.moveToThread(thread)
            thread.started.connect(worker.run)
            worker.frame_ready.connect(self.on_camera_payload)
            worker.error_occurred.connect(self.on_camera_error)
            worker.status_changed.connect(self.on_camera_status_changed)
            worker.finished.connect(thread.quit)
            worker.finished.connect(worker.deleteLater)
            thread.finished.connect(thread.deleteLater)

            cid = cam["camera_id"]
            self.workers[cid] = worker
            self.threads[cid] = thread
            thread.start()

        self.timer = QtCore.QTimer(self)
        self.timer.timeout.connect(self.tick)
        self.timer.start(int(self.app_cfg.system.get("display_update_interval_ms", 200)))
        QtCore.QTimer.singleShot(0, self._show_on_target_screen)

    def tick(self) -> None:
        now = datetime.now()
        self._rollover_system_level_history_if_needed(now)
        self._update_status_level()
        self._update_global_status()
        interval_sec = max(1, int(self.app_cfg.system.get("graph_update_interval_sec", 10)))
        now_ts = time.time()
        if now_ts - self.last_system_level_record_ts >= interval_sec:
            self._record_system_level(self.compute_system_level(), now)
            self.last_system_level_record_ts = now_ts
            self._refresh_system_level_graph()
        self._collect_and_flush_10min_records(now)

    @QtCore.pyqtSlot(dict)
    def on_camera_payload(self, payload: dict[str, Any]) -> None:
        cid = int(payload.get("camera_id", -1))
        if cid in self.panels:
            try:
                self.panels[cid].update_view(payload)
            except Exception as exc:
                logging.exception("[ERROR] cam%s update_view failed: %s", cid, exc)
        self.latest_payloads[cid] = payload
        try:
            self._update_status_level()
        except Exception as exc:
            logging.warning("on_camera_payload status update skipped cam%s: %s", cid, exc)

    def _show_on_target_screen(self) -> None:
        try:
            screens = QtGui.QGuiApplication.screens()
            if not screens:
                self.showMaximized()
                return
            target_index = 2 if len(screens) >= 3 else len(screens) - 1
            target_screen = screens[target_index]
            geom = target_screen.availableGeometry()
            self.setGeometry(geom)
            self.move(geom.topLeft())
        except Exception as exc:
            logging.warning("show_on_target_screen fallback to maximized: %s", exc)
        self.showMaximized()

    @QtCore.pyqtSlot(int, str)
    def on_camera_error(self, camera_id: int, message: str) -> None:
        logging.warning("camera %s: %s", camera_id, message)

    @QtCore.pyqtSlot(int, str)
    def on_camera_status_changed(self, camera_id: int, status_text: str) -> None:
        panel = self.panels.get(camera_id)
        if panel is not None:
            panel.set_status(status_text)

    def _update_status_level(self) -> None:
        overall_level = self.compute_system_level()
        try:
            self.status_mgr.update_if_needed(int(overall_level))
        except Exception as exc:
            logging.warning("ai_status update skipped: %s", exc)

    def _find_payload_by_camera_name(self, camera_name: str) -> dict[str, Any] | None:
        for payload in self.latest_payloads.values():
            if str(payload.get("camera_name", "")).strip() == camera_name:
                return payload
        return None

    def get_camera_congestion_score(self, camera_name: str) -> float:
        payload = self._find_payload_by_camera_name(camera_name)
        if payload is None:
            return 0.0
        return float(payload.get("congestion_score", 0.0))

    def get_camera_long_stay_count(self, camera_name: str, minutes: int = 5) -> int:
        payload = self._find_payload_by_camera_name(camera_name)
        if payload is None:
            return 0
        long_stay_list = payload.get("long_stay_list", [])
        count = 0
        for entry in long_stay_list:
            try:
                if float(entry[1]) >= float(minutes):
                    count += 1
            except (TypeError, ValueError, IndexError):
                continue
        return count

    def get_camera_threshold(self, camera_name: str) -> float:
        for cam in self.app_cfg.cameras:
            if str(cam.get("camera_name", "")).strip() == camera_name:
                return float(cam.get("congestion_threshold", 5.0))
        return 5.0

    def compute_system_level(self) -> int:
        cam1_score = self.get_camera_congestion_score("Camera1")
        cam2_score = self.get_camera_congestion_score("Camera2")
        cam3_score = self.get_camera_congestion_score("Camera3")
        cam1_th = self.get_camera_threshold("Camera1")
        cam2_th = self.get_camera_threshold("Camera2")
        cam3_th = self.get_camera_threshold("Camera3")
        cam2_long_stay_5min = self.get_camera_long_stay_count("Camera2", minutes=5)
        if cam1_score >= cam1_th and cam3_score >= cam3_th:
            return 4
        if cam1_score >= cam1_th:
            return 3
        if cam2_score >= cam2_th and cam2_long_stay_5min >= 3:
            return 2
        return 1

    def update_level_rule_text(self) -> None:
        cam1_th = self.get_camera_threshold("Camera1")
        cam2_th = self.get_camera_threshold("Camera2")
        cam3_th = self.get_camera_threshold("Camera3")
        self.level_rule_label.setText(
            "<b>LEVEL1：</b>通常時<br>"
            f"<b>LEVEL2：</b>[KING]渋滞指数{cam2_th:.1f}以上 + 5分以上滞在台数3台以上<br>"
            f"<b>LEVEL3：</b>[QUEEN]渋滞指数{cam1_th:.1f}以上<br>"
            f"<b>LEVEL4：</b>[QUEEN]渋滞指数{cam1_th:.1f}以上 + [JACK]渋滞指数{cam3_th:.1f}以上"
        )

    @QtCore.pyqtSlot(int, float)
    def on_threshold_changed(self, camera_id: int, value: float) -> None:
        try:
            target = next((cam for cam in self.app_cfg.cameras if int(cam.get("camera_id", -1)) == int(camera_id)), None)
            if target is None:
                return
            target["congestion_threshold"] = float(value)
            self.cfg_mgr.save_camera_settings(self.app_cfg.cameras)
            worker = self.workers.get(camera_id)
            if worker is not None:
                worker.update_camera_config({"congestion_threshold": float(value)})
            self.update_level_rule_text()
            self._update_status_level()
            self._update_global_status()
        except Exception as exc:
            logging.warning("cam%s threshold update failed: %s", camera_id, exc)

    def _update_global_status(self) -> None:
        model_name = (
            self.app_cfg.system.get("yolo_model")
            or self.app_cfg.system.get("YOLO_MODEL")
            or self.app_cfg.system.get("model_path")
            or "n/a"
        )
        gpu = next(iter(self.latest_payloads.values()), {}).get("gpu_name", "n/a")
        self.system_runtime_label.setText(
            f"GPU: {gpu} ｜ model: {model_name} ｜ tracker: ByteTrack"
        )
        level = self.compute_system_level()
        style = level_style(level)
        self.level_badge.setText(f"{style['icon']} 渋滞LEVEL{level}")
        self.level_badge.setStyleSheet(
            f"background:{style['bg']};color:{style['fg']};border-radius:8px;font-weight:900;font-size:38px;padding:4px 12px;"
        )

    def _system_level_metrics_dir(self) -> Path:
        metrics_dir = self.root_dir / "data" / "metrics" / "system"
        metrics_dir.mkdir(parents=True, exist_ok=True)
        return metrics_dir

    def _system_level_history_csv_path(self, target_date: date) -> Path:
        return self._system_level_metrics_dir() / f"system_level_history_{target_date.isoformat()}.csv"

    def _ensure_system_level_history_csv(self, target_date: date) -> Path:
        csv_path = self._system_level_history_csv_path(target_date)
        if csv_path.exists():
            return csv_path
        with csv_path.open("w", newline="", encoding="utf-8") as fh:
            csv.writer(fh).writerow(["timestamp", "system_level"])
        return csv_path

    def _load_system_level_history(self, target_date: date) -> list[tuple[datetime, float]]:
        csv_path = self._system_level_history_csv_path(target_date)
        points: list[tuple[datetime, float]] = []
        if not csv_path.exists():
            return points
        with csv_path.open("r", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                try:
                    ts = datetime.strptime(str(row.get("timestamp", "")), "%Y-%m-%d %H:%M:%S")
                    level = float(row.get("system_level", "0"))
                    if 1.0 <= level <= 4.0:
                        points.append((ts, level))
                except Exception:
                    continue
        return points

    def _rollover_system_level_history_if_needed(self, now: datetime) -> None:
        if now.date() == self.system_level_history_date:
            return
        self.system_level_history_date = now.date()
        self.system_level_history_yesterday = self._load_system_level_history(now.date() - timedelta(days=1))
        self.system_level_history_today = self._load_system_level_history(now.date())

    def _record_system_level(self, level: int, now: datetime) -> None:
        level_val = float(max(1, min(4, int(level))))
        if self.system_level_history_today and (now - self.system_level_history_today[-1][0]).total_seconds() < 1.0:
            self.system_level_history_today[-1] = (now, level_val)
            return
        self.system_level_history_today.append((now, level_val))
        csv_path = self._ensure_system_level_history_csv(now.date())
        with csv_path.open("a", newline="", encoding="utf-8") as fh:
            csv.writer(fh).writerow([now.strftime("%Y-%m-%d %H:%M:%S"), int(level_val)])

    def _refresh_system_level_graph(self) -> None:
        self.system_level_graph.set_line_data(
            self.system_level_history_yesterday,
            self.system_level_history_today,
            "全体渋滞レベル履歴（0-24時）",
            threshold=None,
            show_threshold=False,
        )

    def open_settings(self) -> None:
        cams = self.app_cfg.cameras
        names = [f"{c['camera_id']}: {c['camera_name']}" for c in cams]
        selected, ok = QtWidgets.QInputDialog.getItem(self, "対象カメラ", "設定するカメラ", names, 0, False)
        if not ok:
            return
        camera_id = int(selected.split(":", 1)[0])
        self.open_settings_for_camera(camera_id, "basic")

    def open_settings_for_camera(self, camera_id: int, mode: str = "basic") -> None:
        cams = self.app_cfg.cameras
        idx = next(i for i, c in enumerate(cams) if c["camera_id"] == camera_id)
        live = self.workers[camera_id].get_latest_raw_frame() if camera_id in self.workers else None
        initial_mode = "line" if mode in {"line", "basic"} else "poly"
        dlg = CameraSettingsDialog(cams[idx], live, self, initial_mode=initial_mode)
        if dlg.exec() != QtWidgets.QDialog.DialogCode.Accepted:
            return
        cams[idx] = dlg.get_updated_config()
        self.cfg_mgr.save_camera_settings(cams)
        try:
            self.workers[camera_id].update_camera_config(cams[idx])
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "モデル更新失敗", str(exc))
            return
        QtWidgets.QMessageBox.information(self, "保存", "設定を保存し、次フレームから反映しました。")

    def save_current_settings(self) -> None:
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "設定保存", str(self.root_dir / "config" / "monitor_config.json"), "JSON (*.json)")
        if not path:
            return
        data = {"system": self.app_cfg.system, "cameras": self.app_cfg.cameras}
        Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        QtWidgets.QMessageBox.information(self, "保存", f"保存しました: {path}")

    def load_settings_from_json(self) -> None:
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "設定読込", str(self.root_dir / "config"), "JSON (*.json)")
        if not path:
            return
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        self.app_cfg.system.update(data.get("system", {}))
        self.app_cfg.cameras = data.get("cameras", self.app_cfg.cameras)
        self.cfg_mgr.save_system_settings(self.app_cfg.system)
        self.cfg_mgr.save_camera_settings(self.app_cfg.cameras)
        for cam in self.app_cfg.cameras:
            cid = cam["camera_id"]
            if cid in self.workers:
                try:
                    self.workers[cid].update_camera_config(cam)
                except Exception as exc:
                    QtWidgets.QMessageBox.critical(self, "設定反映失敗", str(exc))
        QtWidgets.QMessageBox.information(self, "読込", "設定を反映しました。")

    def export_daily(self) -> None:
        path = self.reporter.write_daily_report(date.today(), self.app_cfg.cameras, self.root_dir / "data" / "metrics")
        QtWidgets.QMessageBox.information(self, "日次", f"出力完了: {path}")

    def export_monthly(self) -> None:
        month = datetime.now().strftime("%Y-%m")
        path = self.reporter.write_monthly_report(month, self.root_dir / "data" / "metrics", self.app_cfg.cameras)
        QtWidgets.QMessageBox.information(self, "月次", f"出力完了: {path}")

    def _collect_and_flush_10min_records(self, now: datetime) -> None:
        level = self.compute_system_level()
        self.ten_min_buffer.append(self.ten_min_writer.collect_sample(now, level, self.latest_payloads))
        active_bin = self.ten_min_writer._floor_to_10min(now)
        if active_bin == self.current_10min_bin_start:
            return
        self._flush_current_10min_bin()
        prev_date = self.current_10min_date
        self.current_10min_bin_start = active_bin
        self.current_10min_date = active_bin.date()
        if self.current_10min_date != prev_date:
            self.ten_min_writer.finalize_day_graph(prev_date)

    def _flush_current_10min_bin(self) -> None:
        if not self.ten_min_buffer:
            return
        row = self.ten_min_writer.aggregate_10min(self.current_10min_bin_start, self.ten_min_buffer)
        self.ten_min_writer.append_row(self.current_10min_bin_start.date(), row)
        self.ten_min_buffer.clear()

    def closeEvent(self, event: QtGui.QCloseEvent) -> None:
        self._flush_current_10min_bin()
        for worker in self.workers.values():
            try:
                worker.stop()
            except Exception:
                pass

        for thread in self.threads.values():
            thread.quit()
            thread.wait(3000)

        super().closeEvent(event)


MonitorMainWindow = MainWindow


# =========================================
# parse_args / main
# =========================================
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="3-camera AI congestion monitor")
    p.add_argument("--root", default=str(Path(__file__).resolve().parent), help="04_ai_monitor root directory")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    root_dir = Path(args.root)
    setup_logging(root_dir / "logs")
    logging.info("AI monitor starting. root=%s", root_dir)
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow(root_dir)
    win.show()
    try:
        return app.exec()
    except Exception:
        err_path = root_dir / "logs" / "ai_monitor_last_traceback.log"
        err_path.write_text(traceback.format_exc(), encoding="utf-8")
        logging.exception("fatal exception in QApplication loop")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
