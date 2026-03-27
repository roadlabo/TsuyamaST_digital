from __future__ import annotations

# =========================================
# Imports
# =========================================
import argparse
import csv
import json
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

os.environ.setdefault("OPENCV_FFMPEG_CAPTURE_OPTIONS", "rtsp_transport;tcp|fflags;nobuffer|flags;low_delay")
os.environ.setdefault("OPENCV_VIDEOIO_DEBUG", "0")
os.environ.setdefault("OPENCV_LOG_LEVEL", "ERROR")

import cv2
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import torch
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from PyQt6 import QtCore, QtGui, QtWidgets
from ultralytics import YOLO

# =========================================
# Constants / CLASS_MAP
# =========================================
CLASS_MAP = {
    0: "人",
    1: "自転車",
    2: "車",
    3: "オートバイ",
    4: "飛行機",
    5: "バス",
    6: "電車",
    7: "トラック",
}

LEVEL_STYLE_MAP: dict[str, dict[str, str]] = {
    "LEVEL0": {"label": "渋滞LEVEL1", "bg": "#7fd0ff", "fg": "#000000", "icon": "🟢"},
    "LEVEL1": {"label": "渋滞LEVEL2", "bg": "#ffb347", "fg": "#000000", "icon": "🟠"},
    "LEVEL2": {"label": "渋滞LEVEL3", "bg": "#e53935", "fg": "#ffffff", "icon": "🔴"},
    "LEVEL3": {"label": "渋滞LEVEL4", "bg": "#000000", "fg": "#ffffff", "icon": "⚫"},
}

# =========================================
# Default Settings
# =========================================
DEFAULT_SYSTEM_CONFIG: dict[str, Any] = {
    "model_path": "yolo11m.pt",
    "device_preference": "auto",
    "metrics_save_interval_sec": 5,
    "ui_refresh_interval_ms": 500,
    "ai_status_json_path": "app/config/ai_status.json",
    "status_update_interval_sec": 3,
    "output_root": "app/ai_monitor/data",
    "display_update_interval_ms": 800,
    "graph_update_interval_sec": 10,
}

DEFAULT_CAMERA_SETTINGS: dict[str, Any] = {
    "cameras": [
        {
            "camera_id": 1,
            "camera_name": "Camera1",
            "stream_url": "0",
            "enabled": True,
            "line_start": [100, 300],
            "line_end": [1000, 300],
            "exclude_polygon": [],
            "congestion_threshold": 5,
            "long_stay_minutes": 15,
            "long_stay_trigger_count": 1,
            "stay_zone_polygon": [],
            "reconnect_sec": 3,
            "tracking_method": "bytetrack",
            "yolo_model": "yolo11m.pt",
            "confidence_threshold": 0.25,
            "iou_threshold": 0.5,
            "frame_skip": 1,
            "imgsz": 640,
            "target_classes": [2, 3, 5, 7],
            "bt_track_high_thresh": 0.3,
            "bt_track_low_thresh": 0.1,
            "bt_match_thresh": 0.8,
            "bt_track_buffer": 30,
            "crossing_judgment_pattern": "line_cross",
            "distance_threshold": 25.0,
            "congestion_calculation_interval": 10,
            "enable_congestion": True,
            "line_direction_mode": "line_vector",
            "display_scale": 0.85,
        },
        {
            "camera_id": 2,
            "camera_name": "Camera2",
            "stream_url": "1",
            "enabled": True,
            "line_start": [100, 300],
            "line_end": [1000, 300],
            "exclude_polygon": [],
            "congestion_threshold": 5,
            "long_stay_minutes": 15,
            "long_stay_trigger_count": 1,
            "stay_zone_polygon": [],
            "reconnect_sec": 3,
            "tracking_method": "bytetrack",
            "yolo_model": "yolo11m.pt",
            "confidence_threshold": 0.25,
            "iou_threshold": 0.5,
            "frame_skip": 1,
            "imgsz": 640,
            "target_classes": [2, 3, 5, 7],
            "bt_track_high_thresh": 0.3,
            "bt_track_low_thresh": 0.1,
            "bt_match_thresh": 0.8,
            "bt_track_buffer": 30,
            "crossing_judgment_pattern": "line_cross",
            "distance_threshold": 25.0,
            "congestion_calculation_interval": 10,
            "enable_congestion": True,
            "line_direction_mode": "line_vector",
            "display_scale": 0.85,
        },
        {
            "camera_id": 3,
            "camera_name": "Camera3",
            "stream_url": "2",
            "enabled": True,
            "line_start": [100, 300],
            "line_end": [1000, 300],
            "exclude_polygon": [],
            "congestion_threshold": 5,
            "long_stay_minutes": 15,
            "long_stay_trigger_count": 1,
            "stay_zone_polygon": [],
            "reconnect_sec": 3,
            "tracking_method": "bytetrack",
            "yolo_model": "yolo11m.pt",
            "confidence_threshold": 0.25,
            "iou_threshold": 0.5,
            "frame_skip": 1,
            "imgsz": 640,
            "target_classes": [2, 3, 5, 7],
            "bt_track_high_thresh": 0.3,
            "bt_track_low_thresh": 0.1,
            "bt_match_thresh": 0.8,
            "bt_track_buffer": 30,
            "crossing_judgment_pattern": "line_cross",
            "distance_threshold": 25.0,
            "congestion_calculation_interval": 10,
            "enable_congestion": True,
            "line_direction_mode": "line_vector",
            "display_scale": 0.85,
        },
    ]
}

KEEP_CAMERA_KEYS = [
    "camera_id",
    "camera_name",
    "stream_url",
    "enabled",
    "line_start",
    "line_end",
    "exclude_polygon",
    "stay_zone_polygon",
    "congestion_threshold",
    "long_stay_minutes",
]

# =========================================
# Dataclasses
# =========================================
@dataclass
class AppConfig:
    root_dir: Path
    system_config_path: Path
    camera_settings_path: Path
    system: dict[str, Any]
    cameras: list[dict[str, Any]]


@dataclass
class CongestionState:
    frame_inverse_distances: list[float] = field(default_factory=list)
    frame_time_stamps: list[datetime] = field(default_factory=list)
    frame_cumulative_inverse_distance: float = 0.0
    current_congestion_index: float = 0.0
    window_start: datetime | None = None


@dataclass
class CounterState:
    counted_track_ids: set[int] = field(default_factory=set)
    previous_side: dict[int, float] = field(default_factory=dict)
    pass_bins_ltor: list[int] = field(default_factory=lambda: [0] * 144)
    pass_bins_rtol: list[int] = field(default_factory=lambda: [0] * 144)


@dataclass
class TrackState:
    first_seen: dict[int, datetime] = field(default_factory=dict)
    long_stay_emitted: set[int] = field(default_factory=set)


# =========================================
# Config Manager
# =========================================
class ConfigManager:
    def __init__(self, root_dir: Path):
        self.root_dir = root_dir
        self.config_dir = root_dir / "config"
        self.config_dir.mkdir(parents=True, exist_ok=True)
        self.system_config_path = self.config_dir / "system_config.json"
        self.camera_settings_path = self.config_dir / "camera_settings.json"

    def ensure_defaults(self) -> None:
        if not self.system_config_path.exists():
            self._atomic_write_json(self.system_config_path, DEFAULT_SYSTEM_CONFIG)
        if not self.camera_settings_path.exists():
            slim_defaults = [self._to_slim_camera_settings(cam) for cam in DEFAULT_CAMERA_SETTINGS.get("cameras", [])]
            self._atomic_write_json(self.camera_settings_path, {"cameras": slim_defaults})

    def load(self) -> AppConfig:
        self.ensure_defaults()
        system = json.loads(self.system_config_path.read_text(encoding="utf-8"))
        camera_dict = json.loads(self.camera_settings_path.read_text(encoding="utf-8"))
        loaded_cameras = camera_dict.get("cameras", [])
        default_map = {int(c["camera_id"]): dict(c) for c in DEFAULT_CAMERA_SETTINGS.get("cameras", [])}
        cameras: list[dict[str, Any]] = []
        for cam in loaded_cameras:
            raw = dict(cam)
            if "line_points" in raw and "line_start" not in raw and "line_end" not in raw:
                points = raw.get("line_points") or []
                if isinstance(points, list) and len(points) >= 2:
                    raw["line_start"] = points[0]
                    raw["line_end"] = points[1]
            cid = int(raw.get("camera_id", -1))
            base = dict(default_map.get(cid, {}))
            base.update(raw)
            base.pop("line_points", None)
            base.pop("direction", None)
            cameras.append(base)
        return AppConfig(self.root_dir, self.system_config_path, self.camera_settings_path, system, cameras)

    def save_camera_settings(self, cameras: list[dict[str, Any]]) -> None:
        slim = [self._to_slim_camera_settings(cam) for cam in cameras]
        self._atomic_write_json(self.camera_settings_path, {"cameras": slim})

    def save_system_settings(self, system: dict[str, Any]) -> None:
        self._atomic_write_json(self.system_config_path, system)

    @staticmethod
    def _atomic_write_json(path: Path, data: dict[str, Any]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(tmp, path)

    @staticmethod
    def _to_slim_camera_settings(camera: dict[str, Any]) -> dict[str, Any]:
        return {k: camera[k] for k in KEEP_CAMERA_KEYS if k in camera}


# =========================================
# Congestion Logic
# =========================================
class CongestionScorer:
    """AICount11.py の congestion 算出式を監視向けに時間窓化して適用。"""

    def __init__(self, interval_sec: int = 10, day_keep: int = 1):
        self.interval_sec = max(1, int(interval_sec))
        self.day_keep = max(1, day_keep)
        self.state = CongestionState()

    def update_interval(self, interval_sec: int) -> None:
        self.interval_sec = max(1, int(interval_sec))

    def _compute_frame_inverse_distance(self, tracks: list[dict], frame_width: int) -> float:
        if len(tracks) < 2 or frame_width <= 0:
            return 0.0
        total = 0.0
        for i in range(len(tracks)):
            x1, y1 = tracks[i]["center"]
            for j in range(i + 1, len(tracks)):
                x2, y2 = tracks[j]["center"]
                distance = ((x1 - x2) ** 2 + (y1 - y2) ** 2) ** 0.5
                total += 1 / (1 + (distance / frame_width) * 500)
        return total

    def update(self, tracks: list[dict], now: datetime, frame_width: int) -> float:
        if self.state.window_start is None:
            self.state.window_start = now
        self.state.frame_cumulative_inverse_distance += self._compute_frame_inverse_distance(tracks, frame_width)
        elapsed = (now - self.state.window_start).total_seconds()
        if elapsed < self.interval_sec:
            return self.state.current_congestion_index

        value = round(self.state.frame_cumulative_inverse_distance / self.interval_sec, 3)
        self.state.frame_inverse_distances.append(value)
        self.state.frame_time_stamps.append(now)
        self.state.current_congestion_index = value
        self.state.frame_cumulative_inverse_distance = 0.0
        self.state.window_start = now

        day_ago = now - timedelta(days=self.day_keep)
        while self.state.frame_time_stamps and self.state.frame_time_stamps[0] < day_ago:
            self.state.frame_time_stamps.pop(0)
            self.state.frame_inverse_distances.pop(0)
        return value


# =========================================
# Line Counter Logic
# =========================================
class LineCounter:
    def __init__(self, line_points: list[list[int]]):
        self.line_points = line_points
        self.state = CounterState()

    def update_line(self, line_points: list[list[int]]) -> None:
        self.line_points = line_points
        self.state.previous_side.clear()
        self.state.counted_track_ids.clear()

    def _signed_side(self, point: tuple[float, float]) -> float:
        p1, p2 = self.line_points
        vx, vy = p2[0] - p1[0], p2[1] - p1[1]
        wx, wy = point[0] - p1[0], point[1] - p1[1]
        return vx * wy - vy * wx

    def update(self, track_id: int, point: tuple[float, float], class_name: str, now: datetime):
        if len(self.line_points) != 2:
            return None
        current_side = self._signed_side(point)
        prev_side = self.state.previous_side.get(track_id)
        self.state.previous_side[track_id] = current_side
        if prev_side is None or track_id in self.state.counted_track_ids:
            return None

        crossed = (prev_side < 0 <= current_side) or (prev_side > 0 >= current_side)
        if not crossed:
            return None

        direction = "LtoR" if prev_side < current_side else "RtoL"
        self.state.counted_track_ids.add(track_id)
        bin_index = (now.hour * 60 + now.minute) // 10
        if 0 <= bin_index < 144:
            if direction == "LtoR":
                self.state.pass_bins_ltor[bin_index] += 1
            else:
                self.state.pass_bins_rtol[bin_index] += 1

        return {
            "timestamp": now.strftime("%Y-%m-%d %H:%M:%S"),
            "track_id": track_id,
            "class_name": class_name,
            "direction": direction,
        }


# =========================================
# Status Manager
# =========================================
class StatusManager:
    def __init__(self, ai_status_path: Path):
        self.ai_status_path = ai_status_path
        self.last_level = None

    def decide_level(self, cam1_over: bool, cam2_long_stay_count: int, cam2_long_stay_trigger_count: int, cam3_over: bool) -> str:
        if cam3_over:
            return "LEVEL3"
        if cam1_over:
            return "LEVEL1"
        if cam2_long_stay_count >= cam2_long_stay_trigger_count:
            return "LEVEL2"
        return "LEVEL0"

    def update_if_needed(self, level: str) -> None:
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        payload = {"congestion_level": level, "updated_at": now_str}
        if self.last_level == level and self.ai_status_path.exists():
            return
        self._atomic_write_json(self.ai_status_path, payload)
        self.last_level = level

    @staticmethod
    def _atomic_write_json(path: Path, data: dict) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = path.with_suffix(path.suffix + ".tmp")
        tmp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(tmp_path, path)


# =========================================
# Report Writer
# =========================================
class ReportWriter:
    def __init__(self, output_root: Path):
        self.output_root = output_root

    def write_daily_report(self, target_date: date, cameras: list[dict], metrics_root: Path) -> Path:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "summary"
        self._style_sheet(ws_summary)

        ws_summary["A1"] = f"Daily Report {target_date.isoformat()}"
        ws_summary["A1"].font = Font(bold=True, size=16, color="00D7FF")
        headers = ["camera", "total_pass", "max_congestion", "over_threshold_points", "long_stay_events"]
        ws_summary.append(headers)

        total_pass_all = 0
        for cam in cameras:
            cid = cam["camera_id"]
            date_str = target_date.isoformat()
            cam_dir = metrics_root / f"cam{cid}"
            pass_df = self._safe_read(cam_dir / f"pass_events_{date_str}.csv")
            metric_df = self._safe_read(cam_dir / f"congestion_timeseries_{date_str}.csv")
            long_df = self._safe_read(cam_dir / f"long_stay_events_{date_str}.csv")

            total_pass = len(pass_df)
            total_pass_all += total_pass
            max_cong = float(metric_df["congestion_score"].max()) if not metric_df.empty else 0.0
            over_count = int((metric_df["congestion_score"] > metric_df["congestion_threshold"]).sum()) if not metric_df.empty else 0
            long_count = len(long_df)
            ws_summary.append([cam["camera_name"], total_pass, round(max_cong, 2), over_count, long_count])
            self._add_camera_sheet(wb, cam, pass_df, metric_df, long_df)

        ws_summary["A2"] = "date"
        ws_summary["B2"] = target_date.isoformat()
        ws_summary["A3"] = "all_cameras_total_pass"
        ws_summary["B3"] = total_pass_all

        out_path = self.output_root / "reports" / "daily" / f"daily_report_{target_date.isoformat()}.xlsx"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def write_monthly_report(self, target_month: str, metrics_root: Path, cameras: list[dict]) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "monthly"
        self._style_sheet(ws)
        ws.append(["date", "total_pass", "max_congestion", "long_stay_events"])

        daily_map: dict[str, dict[str, float]] = {}
        for cam in cameras:
            cam_dir = metrics_root / f"cam{cam['camera_id']}"
            for file in cam_dir.glob(f"congestion_timeseries_{target_month}-*.csv"):
                date_str = file.stem.split("_")[-1]
                d = daily_map.setdefault(date_str, {"pass": 0, "max": 0, "long": 0})
                metric_df = self._safe_read(file)
                pass_df = self._safe_read(cam_dir / f"pass_events_{date_str}.csv")
                long_df = self._safe_read(cam_dir / f"long_stay_events_{date_str}.csv")
                d["pass"] += len(pass_df)
                if not metric_df.empty:
                    d["max"] = max(d["max"], float(metric_df["congestion_score"].max()))
                d["long"] += len(long_df)

        for key in sorted(daily_map.keys()):
            d = daily_map[key]
            ws.append([key, int(d["pass"]), round(d["max"], 2), int(d["long"])])

        out_path = self.output_root / "reports" / "monthly" / f"monthly_report_{target_month}.xlsx"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def _add_camera_sheet(self, wb: Workbook, cam: dict, pass_df: pd.DataFrame, metric_df: pd.DataFrame, long_df: pd.DataFrame) -> None:
        ws = wb.create_sheet(f"cam{cam['camera_id']}")
        self._style_sheet(ws)
        ws["A1"] = f"{cam['camera_name']} Detail"
        ws["A1"].font = Font(bold=True, size=14, color="00D7FF")

        hist = [0] * 144
        if not pass_df.empty:
            pass_df["timestamp"] = pd.to_datetime(pass_df["timestamp"])
            for ts in pass_df["timestamp"]:
                hist[(ts.hour * 60 + ts.minute) // 10] += 1

        ws.append(["bin", "pass_count"])
        for i, v in enumerate(hist):
            ws.append([i, v])

        bar = BarChart()
        bar.title = "Pass Histogram"
        data = Reference(ws, min_col=2, min_row=3, max_row=146)
        cats = Reference(ws, min_col=1, min_row=3, max_row=146)
        bar.add_data(data, titles_from_data=False)
        bar.set_categories(cats)
        bar.height = 6
        bar.width = 13
        ws.add_chart(bar, "D3")

        start_row = 150
        ws[f"A{start_row}"] = "congestion_time_series"
        ws.append(["timestamp", "congestion_score"])
        for _, r in metric_df[["timestamp", "congestion_score"]].iterrows() if not metric_df.empty else []:
            ws.append([r["timestamp"], float(r["congestion_score"])])

        if len(metric_df) > 2:
            line = LineChart()
            line.title = "Congestion Score"
            dref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + len(metric_df))
            cref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(metric_df))
            line.add_data(dref, titles_from_data=False)
            line.set_categories(cref)
            line.height = 5
            line.width = 13
            ws.add_chart(line, "D18")

        ls_row = 18
        ws[f"A{ls_row}"] = "long_stay_list"
        ws[f"A{ls_row+1}"] = "track_id"
        ws[f"B{ls_row+1}"] = "stay_minutes"
        for i, (_, r) in enumerate(long_df.iterrows() if not long_df.empty else []):
            ws[f"A{ls_row+2+i}"] = int(r["track_id"])
            ws[f"B{ls_row+2+i}"] = float(r["stay_minutes"])

    def _style_sheet(self, ws):
        fill = PatternFill("solid", fgColor="101922")
        thin = Side(style="thin", color="1A9FB6")
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 22
        for row in range(1, 220):
            for col in range(1, 8):
                c = ws.cell(row=row, column=col)
                c.fill = fill
                c.font = Font(color="FFFFFF", size=10)
                c.alignment = Alignment(vertical="center")
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    @staticmethod
    def _safe_read(path: Path) -> pd.DataFrame:
        if not path.exists():
            return pd.DataFrame()
        try:
            return pd.read_csv(path)
        except Exception:
            return pd.DataFrame()


# =========================================
# Graph Utilities
# =========================================
def _normalize_time(ts: pd.Timestamp) -> pd.Timestamp:
    return pd.Timestamp(year=1900, month=1, day=1, hour=ts.hour, minute=ts.minute, second=ts.second)


def build_multi_day_trend(metrics_files: list[Path], metric_col: str) -> pd.DataFrame:
    frames = []
    for path in metrics_files:
        if not path.exists():
            continue
        df = pd.read_csv(path)
        if "timestamp" not in df.columns or metric_col not in df.columns:
            continue
        df["timestamp"] = pd.to_datetime(df["timestamp"])
        df["time_normalized"] = df["timestamp"].apply(_normalize_time)
        frames.append(df[["time_normalized", metric_col]])

    if not frames:
        return pd.DataFrame(columns=["time_normalized", "avg", "median"])

    merged = pd.concat(frames, ignore_index=True)
    merged["time_rounded"] = merged["time_normalized"].dt.floor("1min")
    grouped = merged.groupby("time_rounded")[metric_col]
    out = grouped.mean().rename("avg").to_frame()
    out["median"] = grouped.median()
    return out.reset_index().rename(columns={"time_rounded": "time_normalized"})


def save_multi_day_trend_plot(metrics_files: list[Path], metric_col: str, output_png: Path) -> None:
    trend = build_multi_day_trend(metrics_files, metric_col)
    if trend.empty:
        return
    output_png.parent.mkdir(parents=True, exist_ok=True)

    plt.figure(figsize=(12, 4.5))
    plt.plot(trend["time_normalized"], trend["avg"], label="Average", linewidth=2)
    plt.plot(trend["time_normalized"], trend["median"], label="Median", linewidth=2)
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    plt.gca().xaxis.set_major_locator(mdates.HourLocator(interval=1))
    plt.xticks(rotation=45)
    plt.grid(alpha=0.3)
    plt.title(f"Multi-day trend: {metric_col}")
    plt.tight_layout()
    plt.legend()
    plt.savefig(output_png, dpi=150)
    plt.close()


# =========================================
# UI Panels
# =========================================
class ClickableImageLabel(QtWidgets.QLabel):
    point_clicked = QtCore.pyqtSignal(int, int)

    def __init__(self, text: str = "", parent=None):
        super().__init__(text, parent)
        self.source_size = (1, 1)

    def set_source_size(self, width: int, height: int) -> None:
        self.source_size = (max(1, width), max(1, height))

    def mousePressEvent(self, event: QtGui.QMouseEvent) -> None:
        pix = self.pixmap()
        if pix is None:
            return
        scaled = pix.size()
        off_x = (self.width() - scaled.width()) // 2
        off_y = (self.height() - scaled.height()) // 2
        local_x = event.pos().x() - off_x
        local_y = event.pos().y() - off_y
        if local_x < 0 or local_y < 0 or local_x >= scaled.width() or local_y >= scaled.height():
            return
        src_w, src_h = self.source_size
        x = int(local_x * src_w / max(1, scaled.width()))
        y = int(local_y * src_h / max(1, scaled.height()))
        self.point_clicked.emit(x, y)
        super().mousePressEvent(event)


def _segments_intersect(p1, p2, p3, p4) -> bool:
    def ccw(a, b, c):
        return (c[1] - a[1]) * (b[0] - a[0]) > (b[1] - a[1]) * (c[0] - a[0])
    return ccw(p1, p3, p4) != ccw(p2, p3, p4) and ccw(p1, p2, p3) != ccw(p1, p2, p4)


def has_self_intersection(points: list[list[int]]) -> bool:
    if len(points) < 4:
        return False
    n = len(points)
    for i in range(n):
        a1 = points[i]
        a2 = points[(i + 1) % n]
        for j in range(i + 1, n):
            if abs(i - j) <= 1 or (i == 0 and j == n - 1):
                continue
            b1 = points[j]
            b2 = points[(j + 1) % n]
            if _segments_intersect(a1, a2, b1, b2):
                return True
    return False


class CameraSettingsDialog(QtWidgets.QDialog):
    def __init__(self, camera_cfg: dict[str, Any], latest_frame=None, parent=None, initial_mode: str = "line"):
        super().__init__(parent)
        self.setWindowTitle(f"設定: {camera_cfg['camera_name']}")
        self.resize(1000, 780)
        self.camera_cfg = dict(camera_cfg)
        self.line_points = [camera_cfg.get("line_start", [100, 100])[:], camera_cfg.get("line_end", [400, 100])[:]]
        self.exclude_polygon: list[list[int]] = [p[:] for p in camera_cfg.get("exclude_polygon", [])]
        self.mode = initial_mode if initial_mode in {"line", "poly"} else "line"

        root = QtWidgets.QVBoxLayout(self)
        tabs = QtWidgets.QTabWidget()
        root.addWidget(tabs)

        basic = QtWidgets.QWidget()
        tabs.addTab(basic, "基本")
        form = QtWidgets.QFormLayout(basic)
        self.name_edit = QtWidgets.QLineEdit(camera_cfg.get("camera_name", ""))
        self.url_edit = QtWidgets.QLineEdit(str(camera_cfg.get("stream_url", "")))
        form.addRow("カメラ名", self.name_edit)
        form.addRow("stream_url", self.url_edit)

        ai_tab = QtWidgets.QWidget()
        tabs.addTab(ai_tab, "解析条件")
        ai_form = QtWidgets.QFormLayout(ai_tab)

        self.yolo_model = QtWidgets.QLineEdit(str(camera_cfg.get("yolo_model", "yolo11m.pt")))
        self.conf = QtWidgets.QDoubleSpinBox(); self.conf.setRange(0, 1); self.conf.setSingleStep(0.01); self.conf.setValue(float(camera_cfg.get("confidence_threshold", 0.25)))
        self.iou = QtWidgets.QDoubleSpinBox(); self.iou.setRange(0, 1); self.iou.setSingleStep(0.01); self.iou.setValue(float(camera_cfg.get("iou_threshold", 0.5)))
        self.frame_skip = QtWidgets.QSpinBox(); self.frame_skip.setRange(1, 30); self.frame_skip.setValue(int(camera_cfg.get("frame_skip", 1)))
        self.imgsz = QtWidgets.QSpinBox(); self.imgsz.setRange(320, 2048); self.imgsz.setSingleStep(32); self.imgsz.setValue(int(camera_cfg.get("imgsz", 640)))
        self.bt_hi = QtWidgets.QDoubleSpinBox(); self.bt_hi.setRange(0, 1); self.bt_hi.setValue(float(camera_cfg.get("bt_track_high_thresh", 0.3)))
        self.bt_lo = QtWidgets.QDoubleSpinBox(); self.bt_lo.setRange(0, 1); self.bt_lo.setValue(float(camera_cfg.get("bt_track_low_thresh", 0.1)))
        self.bt_match = QtWidgets.QDoubleSpinBox(); self.bt_match.setRange(0, 1); self.bt_match.setValue(float(camera_cfg.get("bt_match_thresh", 0.8)))
        self.bt_buffer = QtWidgets.QSpinBox(); self.bt_buffer.setRange(1, 1000); self.bt_buffer.setValue(int(camera_cfg.get("bt_track_buffer", 30)))
        self.crossing = QtWidgets.QLineEdit(str(camera_cfg.get("crossing_judgment_pattern", "line_cross")))
        self.distance_th = QtWidgets.QDoubleSpinBox(); self.distance_th.setRange(1, 1000); self.distance_th.setValue(float(camera_cfg.get("distance_threshold", 25.0)))
        self.cong_interval = QtWidgets.QSpinBox(); self.cong_interval.setRange(1, 60); self.cong_interval.setValue(int(camera_cfg.get("congestion_calculation_interval", 10)))
        self.enable_cong = QtWidgets.QCheckBox("enable_congestion"); self.enable_cong.setChecked(bool(camera_cfg.get("enable_congestion", True)))
        self.spin_threshold = QtWidgets.QDoubleSpinBox(); self.spin_threshold.setRange(0.0, 50.0); self.spin_threshold.setDecimals(2); self.spin_threshold.setSingleStep(0.1); self.spin_threshold.setValue(float(camera_cfg.get("congestion_threshold", 5)))
        self.spin_stay = QtWidgets.QSpinBox(); self.spin_stay.setRange(1, 120); self.spin_stay.setValue(int(camera_cfg.get("long_stay_minutes", 15)))

        ai_form.addRow("yolo_model（未設定時のみ使用）", self.yolo_model)
        ai_form.addRow("confidence_threshold", self.conf)
        ai_form.addRow("iou_threshold", self.iou)
        ai_form.addRow("frame_skip", self.frame_skip)
        ai_form.addRow("imgsz", self.imgsz)
        ai_form.addRow("bt_track_high_thresh", self.bt_hi)
        ai_form.addRow("bt_track_low_thresh", self.bt_lo)
        ai_form.addRow("bt_match_thresh", self.bt_match)
        ai_form.addRow("bt_track_buffer", self.bt_buffer)
        ai_form.addRow("crossing_judgment_pattern", self.crossing)
        ai_form.addRow("distance_threshold", self.distance_th)
        ai_form.addRow("congestion_calculation_interval", self.cong_interval)
        ai_form.addRow(self.enable_cong)
        ai_form.addRow("渋滞指数閾値", self.spin_threshold)
        ai_form.addRow("長時間滞在(分)", self.spin_stay)

        class_group = QtWidgets.QGroupBox("対象クラス (0-7)")
        class_layout = QtWidgets.QGridLayout(class_group)
        self.class_checks: dict[int, QtWidgets.QCheckBox] = {}
        selected = set(int(x) for x in camera_cfg.get("target_classes", [2, 3, 5, 7]))
        for i in range(8):
            cb = QtWidgets.QCheckBox(f"{CLASS_MAP[i]}({i})")
            cb.setChecked(i in selected)
            self.class_checks[i] = cb
            class_layout.addWidget(cb, i // 4, i % 4)
        ai_form.addRow(class_group)

        self.image = ClickableImageLabel("snapshot")
        self.image.setMinimumHeight(340)
        self.image.setStyleSheet("background:#0c0f16;border:1px solid #00D7FF;")
        self.image.point_clicked.connect(self._on_click)
        root.addWidget(self.image)

        self.mode_label = QtWidgets.QLabel("")
        self.mode_label.setStyleSheet("color:#d4e6ff;font-weight:bold;")
        root.addWidget(self.mode_label)

        row = QtWidgets.QHBoxLayout()
        btn_line = QtWidgets.QPushButton("ライン設定")
        btn_line.clicked.connect(lambda: self._set_mode("line"))
        btn_poly = QtWidgets.QPushButton("除外エリア設定")
        btn_poly.clicked.connect(lambda: self._set_mode("poly"))
        btn_finish_poly = QtWidgets.QPushButton("指定終了")
        btn_finish_poly.clicked.connect(self._finish_polygon)
        btn_reset = QtWidgets.QPushButton("やり直し")
        btn_reset.clicked.connect(self._reset_mode)
        row.addWidget(btn_line); row.addWidget(btn_poly); row.addWidget(btn_finish_poly); row.addWidget(btn_reset)
        root.addLayout(row)

        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Ok | QtWidgets.QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._validate_and_accept)
        bb.rejected.connect(self.reject)
        root.addWidget(bb)

        self.snapshot = latest_frame
        if self.snapshot is None:
            self.snapshot = np.zeros((360, 640, 3), dtype=np.uint8)
            cv2.putText(self.snapshot, "No live frame", (30, 180), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 0, 255), 2)
        self._update_mode_label()
        self._render_snapshot()

    def _set_mode(self, mode: str) -> None:
        self.mode = mode
        if mode == "line":
            QtWidgets.QMessageBox.information(self, "ライン設定", "ライン設定は2点を取り直します。")
        self._update_mode_label()

    def _update_mode_label(self) -> None:
        txt = "現在モード: ライン設定" if self.mode == "line" else "現在モード: 除外エリア設定"
        self.mode_label.setText(txt)

    def _on_click(self, x: int, y: int) -> None:
        if self.mode == "line":
            if len(self.line_points) >= 2:
                self.line_points = []
            self.line_points.append([x, y])
        else:
            self.exclude_polygon.append([x, y])
        self._render_snapshot()

    def _finish_polygon(self) -> None:
        if len(self.exclude_polygon) < 3:
            QtWidgets.QMessageBox.warning(self, "警告", "除外エリアは3点以上必要です。")
            return
        if has_self_intersection(self.exclude_polygon):
            QtWidgets.QMessageBox.warning(self, "警告", "自己交差ポリゴンは設定できません。やり直してください。")
            self.exclude_polygon = []
        self._render_snapshot()

    def _reset_mode(self) -> None:
        if self.mode == "line":
            self.line_points = []
        else:
            self.exclude_polygon = []
        self._render_snapshot()

    def _render_snapshot(self) -> None:
        frame = self.snapshot.copy()
        if len(self.line_points) == 2:
            cv2.line(frame, tuple(self.line_points[0]), tuple(self.line_points[1]), (0, 255, 255), 2)
            for p in self.line_points:
                cv2.circle(frame, tuple(p), 4, (0, 255, 255), -1)
                cv2.putText(frame, f"{p[0]},{p[1]}", (p[0] + 5, p[1] - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)
        for i, p in enumerate(self.exclude_polygon):
            cv2.circle(frame, tuple(p), 4, (150, 150, 150), -1)
            cv2.putText(frame, str(i + 1), (p[0] + 5, p[1] - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (255, 255, 255), 1)
        if len(self.exclude_polygon) >= 2:
            cv2.polylines(frame, [np.array(self.exclude_polygon, np.int32)], len(self.exclude_polygon) >= 3, (120, 120, 120), 2)

        h, w, _ = frame.shape
        self.image.set_source_size(w, h)
        qimg = QtGui.QImage(frame.data, w, h, frame.strides[0], QtGui.QImage.Format.Format_BGR888)
        self.image.setPixmap(QtGui.QPixmap.fromImage(qimg).scaled(self.image.size(), QtCore.Qt.AspectRatioMode.KeepAspectRatio))

    def resizeEvent(self, event: QtGui.QResizeEvent) -> None:
        super().resizeEvent(event)
        self._render_snapshot()

    def _validate_and_accept(self) -> None:
        selected_classes = [i for i, cb in self.class_checks.items() if cb.isChecked()]
        if not selected_classes:
            QtWidgets.QMessageBox.warning(self, "警告", "対象クラスを1つ以上選択してください。")
            return
        if len(self.line_points) != 2:
            QtWidgets.QMessageBox.warning(self, "警告", "ラインは2点指定してください。")
            return
        self.accept()

    def get_updated_config(self) -> dict[str, Any]:
        cfg = dict(self.camera_cfg)
        cfg["camera_name"] = self.name_edit.text().strip() or cfg["camera_name"]
        cfg["stream_url"] = self.url_edit.text().strip()
        cfg["line_start"] = self.line_points[0]
        cfg["line_end"] = self.line_points[1]
        cfg["exclude_polygon"] = self.exclude_polygon
        cfg["yolo_model"] = self.yolo_model.text().strip() or "yolo11m.pt"
        cfg["confidence_threshold"] = float(self.conf.value())
        cfg["iou_threshold"] = float(self.iou.value())
        cfg["frame_skip"] = int(self.frame_skip.value())
        cfg["imgsz"] = int(self.imgsz.value())
        cfg["target_classes"] = [i for i, cb in self.class_checks.items() if cb.isChecked()]
        cfg["bt_track_high_thresh"] = float(self.bt_hi.value())
        cfg["bt_track_low_thresh"] = float(self.bt_lo.value())
        cfg["bt_match_thresh"] = float(self.bt_match.value())
        cfg["bt_track_buffer"] = int(self.bt_buffer.value())
        cfg["crossing_judgment_pattern"] = self.crossing.text().strip() or "line_cross"
        cfg["distance_threshold"] = float(self.distance_th.value())
        cfg["congestion_calculation_interval"] = int(self.cong_interval.value())
        cfg["enable_congestion"] = bool(self.enable_cong.isChecked())
        cfg["congestion_threshold"] = float(self.spin_threshold.value())
        cfg["long_stay_minutes"] = int(self.spin_stay.value())
        return cfg


class CombinedTimelineGraph(QtWidgets.QWidget):
    def __init__(self, mode: str = "line", parent=None):
        super().__init__(parent)
        self.mode = mode
        self.title = ""
        self.today_points: list[tuple[datetime, float]] = []
        self.prev_points: list[tuple[datetime, float]] = []
        self.today_values: list[float] = []
        self.prev_values: list[float] = []
        self.threshold: float | None = None
        self.show_threshold = False
        self.setFixedHeight(72)

    def set_line_data(self, prev_points: list[tuple[datetime, float]], today_points: list[tuple[datetime, float]], title: str, threshold: float | None = None, show_threshold: bool = True) -> None:
        self.mode = "line"
        self.prev_points = prev_points
        self.today_points = today_points
        self.prev_values = []
        self.today_values = []
        self.title = title
        self.threshold = threshold
        self.show_threshold = show_threshold
        self.update()

    def set_bar_data(self, prev_values: list[int], today_values: list[int], title: str) -> None:
        self.mode = "bar"
        self.prev_values = [float(v) for v in prev_values]
        self.today_values = [float(v) for v in today_values]
        self.prev_points = []
        self.today_points = []
        self.title = title
        self.threshold = None
        self.show_threshold = False
        self.update()

    def paintEvent(self, event: QtGui.QPaintEvent) -> None:
        super().paintEvent(event)
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        painter.fillRect(self.rect(), QtGui.QColor("#0f1620"))

        y_axis_w = 34
        right_margin = 8
        top_margin = 20
        bottom_margin = 22
        plot = QtCore.QRectF(y_axis_w, top_margin, max(10, self.width() - y_axis_w - right_margin), max(10, self.height() - top_margin - bottom_margin))

        painter.setPen(QtGui.QPen(QtGui.QColor("#1d6f8b"), 1))
        painter.drawRoundedRect(self.rect().adjusted(0, 0, -1, -1), 3, 3)
        painter.drawRect(plot)
        painter.setPen(QtGui.QColor("#cfefff"))
        painter.drawText(8, 14, self.title)

        if self.mode == "line":
            ys = [v for _, v in self.prev_points] + [v for _, v in self.today_points]
        else:
            ys = self.prev_values[:] + self.today_values[:]
        if self.show_threshold and self.threshold is not None:
            ys.append(float(self.threshold))
        y_max = max(1.0, max(ys) if ys else 1.0)
        y_min = 0.0

        for i in range(5):
            ratio = i / 4.0
            y = plot.bottom() - ratio * plot.height()
            painter.setPen(QtGui.QPen(QtGui.QColor("#274457"), 1))
            painter.drawLine(QtCore.QPointF(plot.left(), y), QtCore.QPointF(plot.right(), y))
            painter.setPen(QtGui.QColor("#8db6c7"))
            value = y_min + (y_max - y_min) * ratio
            painter.drawText(2, int(y) + 4, f"{value:.1f}")

        x_ticks = list(range(25))
        for h in x_ticks:
            x = plot.left() + (h / 24.0) * plot.width()
            painter.setPen(QtGui.QPen(QtGui.QColor("#274457"), 1))
            painter.drawLine(QtCore.QPointF(x, plot.top()), QtCore.QPointF(x, plot.bottom()))
            painter.setPen(QtGui.QColor("#8db6c7"))
            text = f"{h}"
            text_w = painter.fontMetrics().horizontalAdvance(text)
            painter.drawText(int(x - text_w / 2), self.height() - 6, text)

        if self.mode == "line" and (self.prev_points or self.today_points):
            for points, color in [(self.prev_points, QtGui.QColor("#2f7dff")), (self.today_points, QtGui.QColor("#ff3b3b"))]:
                if not points:
                    continue
                path = QtGui.QPainterPath()
                for i, (ts, value) in enumerate(points):
                    sec = ts.hour * 3600 + ts.minute * 60 + ts.second
                    x = plot.left() + (sec / 86400.0) * plot.width()
                    y = plot.bottom() - ((value - y_min) / (y_max - y_min)) * plot.height()
                    if i == 0:
                        path.moveTo(x, y)
                    else:
                        path.lineTo(x, y)
                painter.setPen(QtGui.QPen(color, 2.0))
                painter.drawPath(path)
            if self.show_threshold and self.threshold is not None:
                th_y = plot.bottom() - ((self.threshold - y_min) / (y_max - y_min)) * plot.height()
                painter.setPen(QtGui.QPen(QtGui.QColor("#ffd400"), 2.0))
                painter.drawLine(QtCore.QPointF(plot.left(), th_y), QtCore.QPointF(plot.right(), th_y))
                painter.setPen(QtGui.QColor("#ffd400"))
                painter.drawText(int(plot.right()) - 82, int(th_y) - 4, f"TH={self.threshold:.2f}")
        elif self.mode == "bar" and (self.prev_values or self.today_values):
            n = max(1, len(self.prev_values), len(self.today_values))
            slot_w = plot.width() / n
            bar_w = max(1.5, slot_w * 0.42)
            for i in range(n):
                prev_val = self.prev_values[i] if i < len(self.prev_values) else 0.0
                today_val = self.today_values[i] if i < len(self.today_values) else 0.0
                base_x = plot.left() + i * slot_w
                prev_h = ((prev_val - y_min) / (y_max - y_min)) * plot.height()
                today_h = ((today_val - y_min) / (y_max - y_min)) * plot.height()
                painter.fillRect(QtCore.QRectF(base_x, plot.bottom() - prev_h, bar_w, prev_h), QtGui.QColor("#2f7dff"))
                painter.fillRect(QtCore.QRectF(base_x + bar_w + 0.5, plot.bottom() - today_h, bar_w, today_h), QtGui.QColor("#ff3b3b"))

        self._draw_legend(painter, plot)

    def _draw_legend(self, painter: QtGui.QPainter, plot: QtCore.QRectF) -> None:
        legend = [("前日", QtGui.QColor("#2f7dff")), ("当日", QtGui.QColor("#ff3b3b"))]
        if self.mode == "line" and self.show_threshold:
            legend.append(("閾値", QtGui.QColor("#ffd400")))
        x = int(plot.right()) - 80
        y = int(plot.top()) + 10
        for label, color in legend:
            painter.setPen(QtGui.QPen(color, 2))
            painter.drawLine(x, y, x + 12, y)
            painter.setPen(QtGui.QColor("#d9ecff"))
            painter.drawText(x + 15, y + 4, label)
            y += 12


class CongestionIndexBar(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.score = 0.0
        self.threshold = 5.0
        self.setMinimumHeight(30)
        self.setMaximumHeight(30)

    def set_values(self, score: float, threshold: float) -> None:
        self.score = float(score)
        self.threshold = float(threshold)
        self.update()

    def paintEvent(self, event: QtGui.QPaintEvent) -> None:
        super().paintEvent(event)
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
        rect = self.rect().adjusted(1, 1, -1, -1)
        painter.setPen(QtGui.QPen(QtGui.QColor("#1f4f7a"), 1))
        painter.setBrush(QtGui.QColor("#0c1a24"))
        painter.drawRoundedRect(rect, 4, 4)

        score_limited = max(0.0, min(20.0, self.score))
        threshold_limited = max(0.0, min(20.0, self.threshold))
        ratio = score_limited / 20.0
        fill = QtCore.QRectF(rect.left(), rect.top(), rect.width() * ratio, rect.height())
        bar_color = QtGui.QColor("#ff3b3b" if self.score >= self.threshold else "#2f7dff")
        painter.fillRect(fill, bar_color)

        th_x = rect.left() + rect.width() * (threshold_limited / 20.0)
        painter.setPen(QtGui.QPen(QtGui.QColor("#ffd400"), 2))
        painter.drawLine(QtCore.QPointF(th_x, rect.top()), QtCore.QPointF(th_x, rect.bottom()))

        painter.setPen(QtGui.QColor("#ffffff"))
        painter.drawText(rect, QtCore.Qt.AlignmentFlag.AlignCenter, f"渋滞指数 {self.score:.1f}")


class CameraPanel(QtWidgets.QFrame):
    line_setting_requested = QtCore.pyqtSignal(int)
    exclude_setting_requested = QtCore.pyqtSignal(int)
    camera_setting_requested = QtCore.pyqtSignal(int)

    def __init__(self, camera_cfg: dict[str, Any], parent=None):
        super().__init__(parent)
        self.camera_id = camera_cfg["camera_id"]
        self.long_stay_minutes = int(camera_cfg.get("long_stay_minutes", 15))
        self.setStyleSheet("QFrame{background:#0a0e13;border:1px solid #169db8;border-radius:6px;} QLabel{color:#cfefff;}")
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(6, 6, 6, 6)
        root.setSpacing(6)

        top_row = QtWidgets.QHBoxLayout()
        top_row.setSpacing(6)

        self.video = QtWidgets.QLabel("video")
        self.video.setMinimumSize(660, 370)
        self.video.setMaximumHeight(370)
        self.video.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.video.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)
        self.video.setStyleSheet("background:#010203;border:1px solid #00a6d6;")
        top_row.addWidget(self.video, 6)

        right_box = QtWidgets.QWidget()
        right_box.setFixedWidth(350)
        right = QtWidgets.QVBoxLayout(right_box)
        right.setContentsMargins(4, 4, 4, 4)
        right.setSpacing(4)
        self.title = QtWidgets.QLabel(camera_cfg["camera_name"])
        self.title.setStyleSheet("font-size:14px;color:#00D7FF;font-weight:bold;")
        right.addWidget(self.title)

        btn_col = QtWidgets.QVBoxLayout()
        btn_col.setSpacing(6)
        self.btn_line = QtWidgets.QPushButton("ライン設定")
        self.btn_line.clicked.connect(lambda: self.line_setting_requested.emit(self.camera_id))
        self.btn_exclude = QtWidgets.QPushButton("除外エリア")
        self.btn_exclude.clicked.connect(lambda: self.exclude_setting_requested.emit(self.camera_id))
        self.btn_ai = QtWidgets.QPushButton("解析条件")
        self.btn_ai.clicked.connect(lambda: self.camera_setting_requested.emit(self.camera_id))
        for btn in (self.btn_line, self.btn_exclude, self.btn_ai):
            btn.setFixedHeight(28)
            btn.setStyleSheet("font-size:12px;padding:4px;")
            btn.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Fixed)
            btn_col.addWidget(btn)

        self.status_label = QtWidgets.QLabel("IDLE")
        self.status_label.setStyleSheet("font-size:13px;font-weight:bold;color:#ffd166;padding:0px;margin:0px;")
        right.addWidget(self.status_label)

        self.congestion_bar = CongestionIndexBar()
        right.addWidget(self.congestion_bar)

        self.label_threshold = QtWidgets.QLabel("渋滞判定閾値：0.0")
        self.label_threshold.setStyleSheet("font-size:12px;color:#ffd400;font-weight:bold;")
        right.addWidget(self.label_threshold)

        self.summary = QtWidgets.QLabel("交通量合計：0（L→R:0 / R→L:0）")
        self.summary.setStyleSheet("font-size:12px;")
        right.addWidget(self.summary)

        self.label_stay = QtWidgets.QLabel(f"滞在時間閾値：{self.long_stay_minutes}分以上")
        self.label_stay.setStyleSheet("font-size:12px;color:#ff9fb0;font-weight:bold;")
        right.addWidget(self.label_stay)

        self.logic_desc = QtWidgets.QLabel(
            "渋滞指数ロジック\n"
            "・検出対象（車・トラック・バス等）の中心点どうしの距離を使用\n"
            "・近い組み合わせほど大きく加点\n"
            "・各フレームの近接度を10秒ごとに平均化して渋滞指数を算出\n"
            "・渋滞指数が設定閾値以上で「渋滞」と判定"
        )
        self.logic_desc.setWordWrap(True)
        self.logic_desc.setStyleSheet("font-size:11px;color:#c7def5;line-height:1.3;")
        right.addWidget(self.logic_desc)

        self.long_stay_title = QtWidgets.QLabel("滞在時間閾値以上")
        long_stay_title = self.long_stay_title
        long_stay_title.setStyleSheet("color:#ff8893;font-weight:bold;")
        right.addWidget(long_stay_title)
        self.long_stay_scroll = QtWidgets.QScrollArea()
        self.long_stay_scroll.setWidgetResizable(True)
        self.long_stay_scroll.setFixedHeight(118)
        self.long_stay_container = QtWidgets.QWidget()
        self.long_stay_layout = QtWidgets.QVBoxLayout(self.long_stay_container)
        self.long_stay_layout.setContentsMargins(0, 0, 0, 0)
        self.long_stay_layout.setSpacing(4)
        self.long_stay_scroll.setWidget(self.long_stay_container)
        right.addWidget(self.long_stay_scroll, 1)
        right.addLayout(btn_col)

        top_row.addWidget(right_box, 4)
        root.addLayout(top_row)

        self.graphs: list[CombinedTimelineGraph] = []
        graphs_box = QtWidgets.QWidget()
        graphs_layout = QtWidgets.QVBoxLayout(graphs_box)
        graphs_layout.setContentsMargins(2, 0, 2, 0)
        graphs_layout.setSpacing(4)
        for _ in range(3):
            g = CombinedTimelineGraph("line")
            self.graphs.append(g)
            graphs_layout.addWidget(g)
        root.addWidget(graphs_box)

    def update_view(self, payload: dict[str, Any]) -> None:
        frame = payload.get("frame")
        if frame is not None:
            h, w, _ = frame.shape
            qimg = QtGui.QImage(frame.data, w, h, frame.strides[0], QtGui.QImage.Format.Format_BGR888)
            self.video.setPixmap(QtGui.QPixmap.fromImage(qimg).scaled(self.video.size(), QtCore.Qt.AspectRatioMode.KeepAspectRatio))

        score = float(payload.get("congestion_score", 0))
        threshold = float(payload.get("threshold", 5))
        self.status_label.setText(payload.get("status", "RUNNING"))

        ltor = payload.get("pass_bins_ltor", [0] * 144)
        rtol = payload.get("pass_bins_rtol", [0] * 144)
        total_ltor = sum(ltor)
        total_rtol = sum(rtol)
        self.summary.setText(f"交通量合計：{total_ltor + total_rtol}（L→R:{total_ltor} / R→L:{total_rtol}）")
        self.label_threshold.setText(f"渋滞判定閾値：{threshold:.2f}")
        self.long_stay_minutes = int(payload.get("long_stay_minutes", self.long_stay_minutes))
        self.label_stay.setText(f"滞在時間閾値：{self.long_stay_minutes}分以上")
        self.long_stay_title.setText("滞在時間閾値以上")
        self._update_congestion_bar(score, threshold)
        self.graphs[0].set_line_data(payload.get("prev_congestion_points", []), payload.get("congestion_points", []), "渋滞指数", threshold=threshold, show_threshold=True)
        self.graphs[1].set_bar_data(payload.get("hist_prev_ltor", [0] * 144), ltor, "LtoR")
        self.graphs[2].set_bar_data(payload.get("hist_prev_rtol", [0] * 144), rtol, "RtoL")
        self._rebuild_long_stay_cards(payload.get("long_stays", []))

    def _update_congestion_bar(self, score: float, threshold: float) -> None:
        self.congestion_bar.set_values(score, threshold)

    def _rebuild_long_stay_cards(self, long_stays: list[tuple[int, float]]) -> None:
        while self.long_stay_layout.count():
            child = self.long_stay_layout.takeAt(0)
            widget = child.widget()
            if widget is not None:
                widget.deleteLater()

        if not long_stays:
            self.long_stay_layout.addStretch()
            return

        for tid, mins in long_stays:
            card = QtWidgets.QFrame()
            card.setStyleSheet("background:#1a1014;border:1px solid #ff5a6e;border-radius:6px;")
            card_layout = QtWidgets.QVBoxLayout(card)
            card_layout.setContentsMargins(8, 6, 8, 6)
            id_label = QtWidgets.QLabel(f"ID={tid:03d}")
            id_label.setStyleSheet("color:#ffd3d9;font-weight:bold;")
            min_label = QtWidgets.QLabel(f"{mins:.0f}min")
            min_label.setStyleSheet("color:#ff6678;")
            card_layout.addWidget(id_label)
            card_layout.addWidget(min_label)
            self.long_stay_layout.addWidget(card)
        self.long_stay_layout.addStretch()

    def set_status(self, status_text: str) -> None:
        self.status_label.setText(status_text)


def resolve_model_path(camera_cfg: dict[str, Any], system_cfg: dict[str, Any], root_dir: Path) -> Path:
    raw = system_cfg.get("model_path") or system_cfg.get("yolo_model") or system_cfg.get("YOLO_MODEL") or camera_cfg.get("yolo_model") or "yolo11m.pt"
    if not raw:
        raise FileNotFoundError("YOLO model is not configured. Set yolo_model / model_path in config.")
    p = Path(str(raw))
    if not p.is_absolute():
        candidates = [root_dir / p, root_dir / "models" / p, Path.cwd() / p]
        for c in candidates:
            if c.exists():
                return c
        raise FileNotFoundError(
            f"YOLO model file not found: {raw}. "
            "Place the model locally under ai_monitor or ai_monitor/models."
        )
    if not p.exists():
        raise FileNotFoundError(f"YOLO model file not found: {p}")
    return p


# =========================================
# Camera Worker
# =========================================
class CameraWorker(QtCore.QObject):
    frame_ready = QtCore.pyqtSignal(dict)
    status_changed = QtCore.pyqtSignal(int, str)
    error_occurred = QtCore.pyqtSignal(int, str)
    finished = QtCore.pyqtSignal(int)

    def __init__(self, camera_cfg: dict[str, Any], system_cfg: dict[str, Any], root_dir: Path):
        super().__init__()
        self.camera_cfg = camera_cfg
        self.system_cfg = system_cfg
        self.root_dir = root_dir
        self.camera_id = int(camera_cfg["camera_id"])
        self.camera_name = camera_cfg["camera_name"]

        self.device = self._resolve_device(system_cfg.get("device_preference", "auto"))
        self.gpu_name = torch.cuda.get_device_name(0) if self.device.startswith("cuda") else "CPU"

        self.model_path = resolve_model_path(self.camera_cfg, self.system_cfg, self.root_dir)
        try:
            self.model = YOLO(str(self.model_path))
        except Exception as exc:
            raise RuntimeError(f"[ERROR] cam{self.camera_id} model load failed: {exc}") from exc
        self.target_classes = set(int(x) for x in self.camera_cfg.get("target_classes", [2, 3, 5, 7]))

        line = [self.camera_cfg.get("line_start", [0, 0]), self.camera_cfg.get("line_end", [100, 0])]
        self.counter = LineCounter(line)
        self.congestion = CongestionScorer(int(self.camera_cfg.get("congestion_calculation_interval", 10)))
        self.track_state = TrackState()
        self.track_class_memory: dict[int, dict[str, Any]] = {}
        self.display_id_map: dict[int, int] = {}
        self.display_id_counter = 1
        self.cross_flash_frames: dict[int, int] = {}

        self.cap = None
        self.last_frame = None
        self.last_raw_frame = None
        self.fps = 0.0
        self.frame_index = 0
        self._running = True
        self._next_reconnect_time = 0.0
        self._next_infer_retry_time = 0.0
        self._last_warn_emit = 0.0
        self._status_text = "INIT"
        self.display_scale = 0.85
        self.csv_error_state = {"congestion": False, "pass": False, "long_stay": False}
        self.read_fail_count = 0
        self.max_read_fail_before_reconnect = 5
        self.last_reconnect_at = 0.0

        self.today = datetime.now().date()
        self.metrics_dir = root_dir / "data" / "metrics" / f"cam{self.camera_id}"
        self.metrics_dir.mkdir(parents=True, exist_ok=True)
        self.congestion_csv, self.pass_csv, self.long_stay_csv = self._ensure_daily_csvs()
        self.previous_day_hist_ltor, self.previous_day_hist_rtol = self._load_previous_day_histogram()
        self.previous_day_congestion_points = self._load_previous_day_congestion_points()
        today_ltor, today_rtol = self._load_today_pass_histogram()
        self.counter.state.pass_bins_ltor = today_ltor
        self.counter.state.pass_bins_rtol = today_rtol
        today_points = self._load_today_congestion_points()
        self.congestion.state.frame_time_stamps = [ts for ts, _ in today_points]
        self.congestion.state.frame_inverse_distances = [v for _, v in today_points]
        if today_points:
            self.congestion.state.current_congestion_index = today_points[-1][1]

    def get_latest_raw_frame(self):
        return None if self.last_raw_frame is None else self.last_raw_frame.copy()

    def update_camera_config(self, new_cfg: dict[str, Any]) -> None:
        prev_cfg = dict(self.camera_cfg)
        old_model_path = self.model_path
        self.camera_cfg.update(new_cfg)
        self.camera_name = new_cfg.get("camera_name", self.camera_name)
        self.target_classes = set(int(x) for x in new_cfg.get("target_classes", [2, 3, 5, 7]))
        self.counter.update_line([new_cfg.get("line_start", [0, 0]), new_cfg.get("line_end", [100, 0])])
        self.congestion.update_interval(int(new_cfg.get("congestion_calculation_interval", 10)))
        self.display_scale = 0.85

        new_model_path = resolve_model_path(self.camera_cfg, self.system_cfg, self.root_dir)
        if new_model_path != old_model_path:
            try:
                new_model = YOLO(str(new_model_path))
                self.model = new_model
                self.model_path = new_model_path
            except Exception as exc:
                self.camera_cfg = prev_cfg
                raise RuntimeError(f"[ERROR] cam{self.camera_id} model load failed: {exc}") from exc

    def _resolve_device(self, preference: str) -> str:
        if preference == "cpu":
            return "cpu"
        if torch.cuda.is_available():
            return "cuda:0"
        return "cpu"

    def _stream_source(self):
        stream_url = str(self.camera_cfg.get("stream_url", "0"))
        return int(stream_url) if stream_url.isdigit() else stream_url

    def connect(self) -> None:
        self._release_capture()
        src = self._stream_source()
        try:
            self.cap = cv2.VideoCapture(src, cv2.CAP_FFMPEG)
        except Exception:
            self.cap = cv2.VideoCapture(src)
        if self.cap is not None:
            try:
                self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            except Exception:
                pass
        self.read_fail_count = 0

    def _reconnect_capture(self, reason: str = "") -> None:
        now_ts = time.time()
        reconnect_sec = float(self.camera_cfg.get("reconnect_sec", 3))
        if now_ts - self.last_reconnect_at < reconnect_sec:
            return
        self.last_reconnect_at = now_ts
        self._set_status("RECONNECTING")
        self.error_occurred.emit(self.camera_id, f"{self.camera_name} 再接続中… ({reason})" if reason else f"{self.camera_name} 再接続中…")
        self._release_capture()
        time.sleep(min(1.0, reconnect_sec))
        try:
            self.connect()
            self._set_status("RUNNING")
            self.error_occurred.emit(self.camera_id, f"{self.camera_name} RUNNING")
        except Exception as exc:
            self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} reconnect failed: {exc}")

    def _release_capture(self) -> None:
        if self.cap is not None:
            try:
                self.cap.release()
            except Exception:
                pass
            finally:
                self.cap = None

    def _set_status(self, text: str) -> None:
        if self._status_text == text:
            return
        self._status_text = text
        self.status_changed.emit(self.camera_id, text)

    def _in_polygon(self, point: tuple[float, float], polygon_points: list[list[int]]) -> bool:
        if not polygon_points:
            return False
        poly = np.array(polygon_points, np.int32)
        return cv2.pointPolygonTest(poly, point, False) >= 0

    def _ensure_daily_csvs(self):
        date_str = self.today.strftime("%Y-%m-%d")
        congestion_ts = self.metrics_dir / f"congestion_timeseries_{date_str}.csv"
        pass_events = self.metrics_dir / f"pass_events_{date_str}.csv"
        long_stay = self.metrics_dir / f"long_stay_events_{date_str}.csv"

        self._ensure_csv_header(
            congestion_ts,
            ["timestamp", "camera_id", "camera_name", "congestion_score", "congestion_threshold", "threshold_over", "fps"],
        )
        self._ensure_csv_header(pass_events, ["timestamp", "camera_id", "track_id", "class_name", "direction"])
        self._ensure_csv_header(long_stay, ["first_seen", "detected_at", "camera_id", "track_id", "stay_minutes", "class_name"])
        return congestion_ts, pass_events, long_stay

    @staticmethod
    def _ensure_csv_header(path: Path, header: list[str]) -> None:
        if path.exists():
            return
        with path.open("w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(header)

    def _append_csv(self, path: Path, row: list[Any]) -> None:
        with path.open("a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(row)

    def _append_csv_safe(self, path: Path, row: list[Any], kind: str) -> None:
        try:
            self._append_csv(path, row)
            self.csv_error_state[kind] = False
        except PermissionError as exc:
            if not self.csv_error_state.get(kind, False):
                self.csv_error_state[kind] = True
                self.error_occurred.emit(
                    self.camera_id,
                    f"[WARN] cam{self.camera_id} {kind} csv is locked. Excel等でCSVが開かれているため書き込めません。閉じると自動で再開します。: {exc}",
                )
        except Exception as exc:
            if not self.csv_error_state.get(kind, False):
                self.csv_error_state[kind] = True
                self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} {kind} csv write failed: {exc}")

    def _load_previous_day_histogram(self) -> tuple[list[int], list[int]]:
        prev = self.today.fromordinal(self.today.toordinal() - 1)
        prev_file = self.metrics_dir / f"pass_events_{prev.strftime('%Y-%m-%d')}.csv"
        ltor = [0] * 144
        rtol = [0] * 144
        if not prev_file.exists():
            return ltor, rtol
        with prev_file.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                dt = datetime.strptime(row["timestamp"], "%Y-%m-%d %H:%M:%S")
                idx = (dt.hour * 60 + dt.minute) // 10
                if 0 <= idx < 144:
                    if row.get("direction") == "LtoR":
                        ltor[idx] += 1
                    else:
                        rtol[idx] += 1
        return ltor, rtol

    def _load_previous_day_congestion_points(self) -> list[tuple[datetime, float]]:
        prev = self.today.fromordinal(self.today.toordinal() - 1)
        prev_file = self.metrics_dir / f"congestion_timeseries_{prev.strftime('%Y-%m-%d')}.csv"
        points: list[tuple[datetime, float]] = []
        if not prev_file.exists():
            return points
        with prev_file.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    ts = datetime.strptime(row["timestamp"], "%Y-%m-%d %H:%M:%S")
                    score = float(row.get("congestion_score", 0.0))
                    points.append((ts, score))
                except Exception:
                    continue
        return points

    def _load_today_pass_histogram(self) -> tuple[list[int], list[int]]:
        today_file = self.metrics_dir / f"pass_events_{self.today.strftime('%Y-%m-%d')}.csv"
        ltor = [0] * 144
        rtol = [0] * 144
        if not today_file.exists():
            return ltor, rtol
        with today_file.open("r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                try:
                    dt = datetime.strptime(row["timestamp"], "%Y-%m-%d %H:%M:%S")
                except Exception:
                    continue
                idx = (dt.hour * 60 + dt.minute) // 10
                if 0 <= idx < 144:
                    if row.get("direction") == "LtoR":
                        ltor[idx] += 1
                    else:
                        rtol[idx] += 1
        return ltor, rtol

    def _load_today_congestion_points(self) -> list[tuple[datetime, float]]:
        today_file = self.metrics_dir / f"congestion_timeseries_{self.today.strftime('%Y-%m-%d')}.csv"
        points: list[tuple[datetime, float]] = []
        if not today_file.exists():
            return points
        with today_file.open("r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                try:
                    points.append((datetime.strptime(row["timestamp"], "%Y-%m-%d %H:%M:%S"), float(row["congestion_score"])))
                except Exception:
                    continue
        return points

    def _rollover_if_needed(self, now: datetime) -> None:
        if now.date() == self.today:
            return
        self.today = now.date()
        self.display_id_map.clear()
        self.display_id_counter = 1
        self.cross_flash_frames.clear()
        self.congestion_csv, self.pass_csv, self.long_stay_csv = self._ensure_daily_csvs()
        self.previous_day_hist_ltor, self.previous_day_hist_rtol = self._load_previous_day_histogram()
        self.previous_day_congestion_points = self._load_previous_day_congestion_points()
        self.counter.state.pass_bins_ltor, self.counter.state.pass_bins_rtol = self._load_today_pass_histogram()
        today_points = self._load_today_congestion_points()
        self.congestion.state.frame_time_stamps = [ts for ts, _ in today_points]
        self.congestion.state.frame_inverse_distances = [v for _, v in today_points]
        self.congestion.state.current_congestion_index = today_points[-1][1] if today_points else 0.0

    def _get_display_id(self, track_id: int) -> int:
        if track_id not in self.display_id_map:
            self.display_id_map[track_id] = self.display_id_counter
            self.display_id_counter += 1
        return self.display_id_map[track_id]

    @staticmethod
    def _bbox_iou(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> float:
        ax1, ay1, ax2, ay2 = a
        bx1, by1, bx2, by2 = b
        inter_x1, inter_y1 = max(ax1, bx1), max(ay1, by1)
        inter_x2, inter_y2 = min(ax2, bx2), min(ay2, by2)
        iw = max(0, inter_x2 - inter_x1)
        ih = max(0, inter_y2 - inter_y1)
        inter = iw * ih
        area_a = max(1, (ax2 - ax1) * (ay2 - ay1))
        area_b = max(1, (bx2 - bx1) * (by2 - by1))
        union = max(1, area_a + area_b - inter)
        return inter / union

    def _deduplicate_overlapping_detections(self, raw_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        priority = {"truck": 0, "bus": 1, "car": 2, "motorcycle": 3, "bicycle": 4, "person": 5}
        used = [False] * len(raw_items)
        kept: list[dict[str, Any]] = []
        conf_eps = 0.05
        for i, item in enumerate(raw_items):
            if used[i]:
                continue
            cluster = [item]
            used[i] = True
            x1, y1, x2, y2 = item["bbox"]
            base_w = max(1.0, float(x2 - x1))
            base_h = max(1.0, float(y2 - y1))
            for j in range(i + 1, len(raw_items)):
                if used[j]:
                    continue
                cand = raw_items[j]
                iou = self._bbox_iou(item["bbox"], cand["bbox"])
                cx1, cy1 = item["center"]
                cx2, cy2 = cand["center"]
                dist = ((cx1 - cx2) ** 2 + (cy1 - cy2) ** 2) ** 0.5
                if iou >= 0.65 or dist <= min(base_w, base_h) * 0.35:
                    cluster.append(cand)
                    used[j] = True
            cluster.sort(key=lambda x: (-float(x.get("conf", 0.0)), priority.get(str(x.get("cls_name", "")).lower(), 999)))
            best = cluster[0]
            tid = int(best.get("track_id", -1))
            if tid >= 0:
                mem = self.track_class_memory.get(tid)
                if mem is not None:
                    prev_cls = str(mem.get("cls_name", best["cls_name"]))
                    prev_ts = mem.get("ts", datetime.min)
                    prev_area = float(mem.get("area", 1.0))
                    new_area = max(1.0, (best["bbox"][2] - best["bbox"][0]) * (best["bbox"][3] - best["bbox"][1]))
                    if (datetime.now() - prev_ts).total_seconds() <= 3.0 and abs(float(best.get("conf", 0)) - float(mem.get("conf", 0))) <= conf_eps:
                        if prev_cls != best["cls_name"] and 0.55 <= (new_area / max(1.0, prev_area)) <= 1.8:
                            best["cls_name"] = prev_cls
                self.track_class_memory[tid] = {"cls_name": best["cls_name"], "ts": datetime.now(), "area": max(1.0, (best["bbox"][2] - best["bbox"][0]) * (best["bbox"][3] - best["bbox"][1])), "conf": float(best.get("conf", 0.0))}
            kept.append(best)
        return kept

    @QtCore.pyqtSlot()
    def run(self) -> None:
        self._set_status("RUNNING")
        while self._running:
            try:
                payload = self.process_once_nonblocking()
                if payload is not None:
                    self.frame_ready.emit(payload)
            except Exception as exc:
                self.error_occurred.emit(self.camera_id, str(exc))
            QtCore.QThread.msleep(5)
        self._release_capture()
        self.finished.emit(self.camera_id)

    @QtCore.pyqtSlot()
    def stop(self) -> None:
        self._running = False
        self._release_capture()

    def process_once_nonblocking(self) -> dict[str, Any] | None:
        try:
            start = time.time()
            now_ts = time.time()
            if now_ts < self._next_reconnect_time or now_ts < self._next_infer_retry_time:
                return None

            try:
                if self.cap is None or not self.cap.isOpened():
                    self._reconnect_capture("cap not opened")
                    return None

                ok, frame = self.cap.read()
                if not ok or frame is None:
                    self.read_fail_count += 1
                    if self.read_fail_count >= self.max_read_fail_before_reconnect:
                        self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} frame read failed repeatedly. reconnecting.")
                        self._reconnect_capture("read failed")
                    return None
                self.read_fail_count = 0
            except cv2.error as exc:
                self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} OpenCV error: {exc}")
                self._reconnect_capture("cv2.error")
                return None
            except Exception as exc:
                self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} capture exception: {exc}")
                self._reconnect_capture("generic exception")
                return None

            self._set_status("RUNNING")
            self.last_raw_frame = self._resize_for_display(frame.copy())
            now = datetime.now()
            self._rollover_if_needed(now)

            self.frame_index += 1
            frame_skip = max(1, int(self.camera_cfg.get("frame_skip", 1)))
            if self.frame_index % frame_skip != 0:
                return None

            try:
                result = self.model.track(
                    source=frame,
                    persist=True,
                    tracker="bytetrack.yaml",
                    verbose=False,
                    device=self.device,
                    conf=float(self.camera_cfg.get("confidence_threshold", 0.25)),
                    iou=float(self.camera_cfg.get("iou_threshold", 0.5)),
                    imgsz=int(self.camera_cfg.get("imgsz", 640)),
                    classes=sorted(self.target_classes) if self.target_classes else None,
                )[0]
            except Exception as exc:
                self._next_infer_retry_time = time.time() + 3.0
                self._set_status("MODEL ERROR")
                self.error_occurred.emit(self.camera_id, f"[ERROR] cam{self.camera_id} tracker failed: {exc}")
                return None

            boxes = result.boxes
            tracks: list[dict[str, Any]] = []
            raw_items: list[dict[str, Any]] = []
            pass_events: list[dict[str, Any]] = []
            long_stay_events: list[dict[str, Any]] = []
            long_stay_list: list[tuple[int, float]] = []
            exclude_polygon = self.camera_cfg.get("exclude_polygon", [])
            stay_zone = self.camera_cfg.get("stay_zone_polygon", [])

            if boxes is not None and boxes.id is not None:
                cls_array = boxes.cls.cpu().numpy() if boxes.cls is not None else []
                conf_array = boxes.conf.cpu().numpy() if boxes.conf is not None else []
                for i, (box, tid) in enumerate(zip(boxes.xyxy.cpu().numpy(), boxes.id.cpu().numpy())):
                    cls_idx = int(cls_array[i]) if len(cls_array) > i else -1
                    if self.target_classes and cls_idx not in self.target_classes:
                        continue
                    cls_name = self.model.names.get(cls_idx, str(cls_idx)) if isinstance(self.model.names, dict) else str(cls_idx)
                    conf = float(conf_array[i]) if len(conf_array) > i else 0.0

                    x1, y1, x2, y2 = box.tolist()
                    bbox_h = max(1.0, y2 - y1)
                    judge_x = (x1 + x2) / 2
                    judge_y = y2 - (bbox_h * 0.2)
                    if self._in_polygon((judge_x, judge_y), exclude_polygon):
                        continue

                    track_id = int(tid)
                    display_id = self._get_display_id(track_id)
                    raw_items.append({
                        "track_id": track_id,
                        "display_id": display_id,
                        "cls_idx": cls_idx,
                        "cls_name": cls_name,
                        "conf": conf,
                        "bbox": (int(x1), int(y1), int(x2), int(y2)),
                        "center": (judge_x, judge_y),
                        "judge_point": (judge_x, judge_y),
                    })

                for item in self._deduplicate_overlapping_detections(raw_items):
                    track_id = int(item["track_id"])
                    display_id = int(item.get("display_id", track_id))
                    cls_name = item["cls_name"]
                    judge_x, judge_y = item["center"]
                    tracks.append(
                        {
                            "track_id": track_id,
                            "display_id": display_id,
                            "center": (judge_x, judge_y),
                            "judge_point": (judge_x, judge_y),
                            "bbox": item["bbox"],
                            "class_name": cls_name,
                            "crossed": False,
                        }
                    )

                    event = self.counter.update(track_id, (judge_x, judge_y), cls_name, now)
                    if event:
                        pass_events.append(event)
                        self.cross_flash_frames[track_id] = 12
                        tracks[-1]["crossed"] = True

                    if track_id not in self.track_state.first_seen:
                        self.track_state.first_seen[track_id] = now

                    in_stay_zone = self._in_polygon((judge_x, judge_y), stay_zone) if stay_zone else True
                    if in_stay_zone:
                        stay_mins = (now - self.track_state.first_seen[track_id]).total_seconds() / 60.0
                        if stay_mins >= float(self.camera_cfg.get("long_stay_minutes", 15)):
                            long_stay_list.append((display_id, stay_mins))
                            if track_id not in self.track_state.long_stay_emitted:
                                self.track_state.long_stay_emitted.add(track_id)
                                long_stay_events.append({
                                    "first_seen": self.track_state.first_seen[track_id].strftime("%Y-%m-%d %H:%M:%S"),
                                    "detected_at": now.strftime("%Y-%m-%d %H:%M:%S"),
                                    "camera_id": self.camera_id,
                                    "track_id": track_id,
                                    "stay_minutes": round(stay_mins, 1),
                                    "class_name": cls_name,
                                })
                    elif self.cross_flash_frames.get(track_id, 0) > 0:
                        tracks[-1]["crossed"] = True

            for tid in list(self.cross_flash_frames):
                self.cross_flash_frames[tid] -= 1
                if self.cross_flash_frames[tid] <= 0:
                    self.cross_flash_frames.pop(tid, None)

            frame_width = int(frame.shape[1]) if frame is not None else 1920
            prev_points_len = len(self.congestion.state.frame_time_stamps)
            congestion_score = self.congestion.update(tracks, now, frame_width) if self.camera_cfg.get("enable_congestion", True) else 0.0
            threshold = float(self.camera_cfg.get("congestion_threshold", 5))
            threshold_over = congestion_score >= threshold

            elapsed = max(1e-6, time.time() - start)
            self.fps = 1.0 / elapsed

            for pe in pass_events:
                self._append_csv_safe(self.pass_csv, [pe["timestamp"], self.camera_id, pe["track_id"], pe["class_name"], pe["direction"]], kind="pass")
            for le in long_stay_events:
                self._append_csv_safe(
                    self.long_stay_csv,
                    [le["first_seen"], le["detected_at"], le["camera_id"], le["track_id"], le["stay_minutes"], le["class_name"]],
                    kind="long_stay",
                )

            if len(self.congestion.state.frame_time_stamps) > prev_points_len:
                self._append_csv_safe(
                    self.congestion_csv,
                    [
                        now.strftime("%Y-%m-%d %H:%M:%S"),
                        self.camera_id,
                        self.camera_name,
                        round(congestion_score, 2),
                        round(threshold, 2),
                        bool(threshold_over),
                        round(self.fps, 2),
                    ],
                    kind="congestion",
                )

            self.last_frame = self._resize_for_display(self._draw_overlay(frame.copy(), tracks, long_stay_list))
            long_stay_list.sort(key=lambda x: x[1], reverse=True)

            return {
                "camera_id": self.camera_id,
                "camera_name": self.camera_name,
                "frame": self.last_frame,
                "congestion_score": congestion_score,
                "threshold": threshold,
                "threshold_over": threshold_over,
                "congestion_points": list(zip(self.congestion.state.frame_time_stamps, self.congestion.state.frame_inverse_distances)),
                "prev_congestion_points": self.previous_day_congestion_points,
                "pass_bins_ltor": self.counter.state.pass_bins_ltor,
                "pass_bins_rtol": self.counter.state.pass_bins_rtol,
                "hist_prev_ltor": self.previous_day_hist_ltor,
                "hist_prev_rtol": self.previous_day_hist_rtol,
                "long_stays": long_stay_list[:10],
                "long_stay_count": len(long_stay_list),
                "long_stay_minutes": int(self.camera_cfg.get("long_stay_minutes", 15)),
                "fps": self.fps,
                "status": self._status_text,
                "device": self.device,
                "gpu_name": self.gpu_name,
            }
        except Exception as exc:
            self.error_occurred.emit(self.camera_id, f"[WARN] cam{self.camera_id} loop exception: {exc}")
            self._reconnect_capture("loop exception")
            return None

    def _resize_for_display(self, frame: np.ndarray) -> np.ndarray:
        if frame is None:
            return frame
        if self.display_scale >= 0.999:
            return frame
        h, w = frame.shape[:2]
        new_w = max(1, int(w * self.display_scale))
        new_h = max(1, int(h * self.display_scale))
        return cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_AREA)

    def _draw_overlay(self, frame: np.ndarray, tracks: list[dict[str, Any]], long_stays: list[tuple[int, float]]) -> np.ndarray:
        line = [self.camera_cfg.get("line_start", [10, 10]), self.camera_cfg.get("line_end", [100, 10])]
        cv2.line(frame, tuple(line[0]), tuple(line[1]), (0, 255, 255), 2)

        poly = self.camera_cfg.get("exclude_polygon", [])
        if len(poly) >= 3:
            cv2.polylines(frame, [np.array(poly, np.int32)], isClosed=True, color=(255, 255, 0), thickness=2)
            cv2.putText(frame, "EXCLUDE", tuple(np.array(poly[0], dtype=int)), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 0), 2, cv2.LINE_AA)

        for tr in tracks:
            x1, y1, x2, y2 = tr["bbox"]
            box_color = (0, 0, 255) if tr.get("crossed", False) or self.cross_flash_frames.get(int(tr["track_id"]), 0) > 0 else (0, 255, 0)
            cv2.rectangle(frame, (x1, y1), (x2, y2), box_color, 2)
            px, py = tr.get("judge_point", tr["center"])
            cv2.circle(frame, (int(px), int(py)), 3, (255, 255, 255), -1)
            cv2.putText(
                frame,
                f"ID:{int(tr.get('display_id', tr['track_id'])):03d} {tr['class_name']}",
                (x1, max(20, y1 - 8)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.55,
                box_color,
                2,
                cv2.LINE_AA,
            )

        for i, (tid, mins) in enumerate(long_stays[:5]):
            cv2.putText(frame, f"LONG STAY ID:{tid:03d} {mins:.1f}m", (20, 40 + 22 * i), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
        return frame

# =========================================
# Main Window
# =========================================
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, root_dir: Path):
        super().__init__()
        self.root_dir = root_dir
        self.setWindowTitle("AI Congestion Monitor")
        self.setStyleSheet("background:#02060a;")
        self.resize(1080, 1860)
        self.setMinimumSize(980, 1600)

        self.cfg_mgr = ConfigManager(root_dir)
        self.app_cfg = self.cfg_mgr.load()
        self.reporter = ReportWriter(root_dir / "data")
        raw_ai_status = Path(self.app_cfg.system.get("ai_status_json_path", "app/config/ai_status.json"))
        script_base = Path(__file__).resolve().parents[2]
        if raw_ai_status.is_absolute():
            ai_status_path = raw_ai_status
        else:
            ai_status_path = script_base / raw_ai_status
        ai_status_path.parent.mkdir(parents=True, exist_ok=True)
        self.status_mgr = StatusManager(ai_status_path)

        self.threads: dict[int, QtCore.QThread] = {}
        self.workers: dict[int, CameraWorker] = {}
        self.panels: dict[int, CameraPanel] = {}
        self.latest_payloads: dict[int, dict[str, Any]] = {}
        self.pending_report_update = False
        self.report_warning_shown = False

        toolbar = QtWidgets.QToolBar("Main", self)
        toolbar.setMovable(False)
        self.addToolBar(QtCore.Qt.ToolBarArea.TopToolBarArea, toolbar)
        action_setting = toolbar.addAction("解析条件")
        action_setting.triggered.connect(self.open_settings)
        action_save = toolbar.addAction("保存")
        action_save.triggered.connect(self.save_current_settings)
        action_load = toolbar.addAction("読込")
        action_load.triggered.connect(self.load_settings_from_json)
        action_daily = toolbar.addAction("日次Excel出力")
        action_daily.triggered.connect(self.export_daily)
        action_monthly = toolbar.addAction("月次Excel出力")
        action_monthly.triggered.connect(self.export_monthly)

        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        self.setCentralWidget(scroll)
        content = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(content)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(6)
        scroll.setWidget(content)

        top_status_row = QtWidgets.QHBoxLayout()
        self.global_status = QtWidgets.QLabel("時刻 | device | GPU | model | output")
        self.global_status.setStyleSheet("color:#b7dbff;background:#0a1420;border:1px solid #1f4f7a;padding:6px;font-size:12px;")
        top_status_row.addWidget(self.global_status, 4)
        self.level_badge = QtWidgets.QLabel("🟢 渋滞LEVEL1")
        self.level_badge.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.level_badge.setMinimumHeight(46)
        self.level_badge.setStyleSheet("background:#7fd0ff;color:#000000;border-radius:8px;font-weight:900;font-size:24px;padding:4px 14px;")
        top_status_row.addWidget(self.level_badge, 2)
        layout.addLayout(top_status_row)

        for cam in self.app_cfg.cameras:
            if not cam.get("enabled", True):
                continue
            panel = CameraPanel(cam)
            panel.line_setting_requested.connect(lambda cid, m="line": self.open_settings_for_camera(cid, m))
            panel.exclude_setting_requested.connect(lambda cid, m="poly": self.open_settings_for_camera(cid, m))
            panel.camera_setting_requested.connect(lambda cid, m="basic": self.open_settings_for_camera(cid, m))
            layout.addWidget(panel, 1)
            self.panels[cam["camera_id"]] = panel
            try:
                worker = CameraWorker(cam, self.app_cfg.system, self.root_dir)
            except Exception as exc:
                panel.set_status("MODEL ERROR")
                print(exc)
                continue
            thread = QtCore.QThread(self)
            worker.moveToThread(thread)
            thread.started.connect(worker.run)
            worker.frame_ready.connect(self.on_camera_payload)
            worker.error_occurred.connect(self.on_camera_error)
            worker.status_changed.connect(self.on_camera_status_changed)
            worker.finished.connect(thread.quit)
            worker.finished.connect(worker.deleteLater)
            thread.finished.connect(thread.deleteLater)

            cid = cam["camera_id"]
            self.workers[cid] = worker
            self.threads[cid] = thread
            thread.start()

        self.timer = QtCore.QTimer(self)
        self.timer.timeout.connect(self.tick)
        self.timer.start(int(self.app_cfg.system.get("display_update_interval_ms", 800)))
        self.report_update_timer = QtCore.QTimer(self)
        self.report_update_timer.timeout.connect(self._flush_report_update_if_needed)
        self.report_update_timer.start(300000)
        self.report_retry_timer = QtCore.QTimer(self)
        self.report_retry_timer.timeout.connect(self._retry_pending_report_update)
        self.report_retry_timer.start(30000)
        QtCore.QTimer.singleShot(0, self._show_on_target_screen)

    def tick(self) -> None:
        self._update_status_level()
        self._update_global_status()

    @QtCore.pyqtSlot(dict)
    def on_camera_payload(self, payload: dict[str, Any]) -> None:
        cid = int(payload.get("camera_id", -1))
        if cid in self.panels:
            self.panels[cid].update_view(payload)
        self.latest_payloads[cid] = payload
        self._update_status_level()
        self.pending_report_update = True

    def _show_on_target_screen(self) -> None:
        screens = QtGui.QGuiApplication.screens()
        if not screens:
            self.showMaximized()
            return

        target_index = 2 if len(screens) >= 3 else len(screens) - 1
        target_screen = screens[target_index]
        geom = target_screen.availableGeometry()
        self.setGeometry(geom)
        self.move(geom.topLeft())
        self.showMaximized()

    @QtCore.pyqtSlot(int, str)
    def on_camera_error(self, camera_id: int, message: str) -> None:
        print(f"[WARN] camera {camera_id}: {message}")

    @QtCore.pyqtSlot(int, str)
    def on_camera_status_changed(self, camera_id: int, status_text: str) -> None:
        panel = self.panels.get(camera_id)
        if panel is not None:
            panel.set_status(status_text)

    def _update_status_level(self) -> None:
        cam1 = self.latest_payloads.get(1, {})
        cam2 = self.latest_payloads.get(2, {})
        cam3 = self.latest_payloads.get(3, {})
        cam2_cfg = next((c for c in self.app_cfg.cameras if c.get("camera_id") == 2), {})
        level = self.status_mgr.decide_level(
            cam1_over=bool(cam1.get("threshold_over", False)),
            cam2_long_stay_count=int(cam2.get("long_stay_count", 0)),
            cam2_long_stay_trigger_count=int(cam2_cfg.get("long_stay_trigger_count", 1)),
            cam3_over=bool(cam3.get("threshold_over", False)),
        )
        self.status_mgr.update_if_needed(level)

    def _update_global_status(self) -> None:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        model_name = (
            self.app_cfg.system.get("yolo_model")
            or self.app_cfg.system.get("YOLO_MODEL")
            or self.app_cfg.system.get("model_path")
            or "n/a"
        )
        device = next(iter(self.latest_payloads.values()), {}).get("device", "n/a")
        gpu = next(iter(self.latest_payloads.values()), {}).get("gpu_name", "n/a")
        cam1 = self.latest_payloads.get(1, {})
        cam2 = self.latest_payloads.get(2, {})
        cam3 = self.latest_payloads.get(3, {})
        cam2_cfg = next((c for c in self.app_cfg.cameras if c.get("camera_id") == 2), {})
        level = self.status_mgr.decide_level(bool(cam1.get("threshold_over", False)), int(cam2.get("long_stay_count", 0)), int(cam2_cfg.get("long_stay_trigger_count", 1)), bool(cam3.get("threshold_over", False)))
        self.global_status.setText(
            f"{now} | device={device} | GPU={gpu} | model={model_name} | output={self.root_dir / 'data'}"
        )
        style = LEVEL_STYLE_MAP.get(level, LEVEL_STYLE_MAP["LEVEL0"])
        self.level_badge.setText(f"{style['icon']} {style['label']}")
        self.level_badge.setStyleSheet(
            f"background:{style['bg']};color:{style['fg']};border-radius:8px;font-weight:900;font-size:24px;padding:4px 14px;"
        )

    def open_settings(self) -> None:
        cams = self.app_cfg.cameras
        names = [f"{c['camera_id']}: {c['camera_name']}" for c in cams]
        selected, ok = QtWidgets.QInputDialog.getItem(self, "対象カメラ", "設定するカメラ", names, 0, False)
        if not ok:
            return
        camera_id = int(selected.split(":", 1)[0])
        self.open_settings_for_camera(camera_id, "basic")

    def open_settings_for_camera(self, camera_id: int, mode: str = "basic") -> None:
        cams = self.app_cfg.cameras
        idx = next(i for i, c in enumerate(cams) if c["camera_id"] == camera_id)
        live = self.workers[camera_id].get_latest_raw_frame() if camera_id in self.workers else None
        initial_mode = "line" if mode in {"line", "basic"} else "poly"
        dlg = CameraSettingsDialog(cams[idx], live, self, initial_mode=initial_mode)
        if dlg.exec() != QtWidgets.QDialog.DialogCode.Accepted:
            return
        cams[idx] = dlg.get_updated_config()
        self.cfg_mgr.save_camera_settings(cams)
        try:
            self.workers[camera_id].update_camera_config(cams[idx])
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "モデル更新失敗", str(exc))
            return
        QtWidgets.QMessageBox.information(self, "保存", "設定を保存し、次フレームから反映しました。")

    def save_current_settings(self) -> None:
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "設定保存", str(self.root_dir / "config" / "monitor_config.json"), "JSON (*.json)")
        if not path:
            return
        data = {"system": self.app_cfg.system, "cameras": self.app_cfg.cameras}
        Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        QtWidgets.QMessageBox.information(self, "保存", f"保存しました: {path}")

    def load_settings_from_json(self) -> None:
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "設定読込", str(self.root_dir / "config"), "JSON (*.json)")
        if not path:
            return
        data = json.loads(Path(path).read_text(encoding="utf-8"))
        self.app_cfg.system.update(data.get("system", {}))
        self.app_cfg.cameras = data.get("cameras", self.app_cfg.cameras)
        self.cfg_mgr.save_system_settings(self.app_cfg.system)
        self.cfg_mgr.save_camera_settings(self.app_cfg.cameras)
        for cam in self.app_cfg.cameras:
            cid = cam["camera_id"]
            if cid in self.workers:
                try:
                    self.workers[cid].update_camera_config(cam)
                except Exception as exc:
                    QtWidgets.QMessageBox.critical(self, "設定反映失敗", str(exc))
        QtWidgets.QMessageBox.information(self, "読込", "設定を反映しました。")

    def export_daily(self) -> None:
        path = self.reporter.write_daily_report(date.today(), self.app_cfg.cameras, self.root_dir / "data" / "metrics")
        self._export_multi_day_plot()
        QtWidgets.QMessageBox.information(self, "日次", f"出力完了: {path}")

    def export_monthly(self) -> None:
        month = datetime.now().strftime("%Y-%m")
        path = self.reporter.write_monthly_report(month, self.root_dir / "data" / "metrics", self.app_cfg.cameras)
        QtWidgets.QMessageBox.information(self, "月次", f"出力完了: {path}")

    def _export_multi_day_plot(self) -> None:
        metrics_files = []
        for cam in self.app_cfg.cameras:
            cid = cam["camera_id"]
            cam_dir = self.root_dir / "data" / "metrics" / f"cam{cid}"
            metrics_files.extend(sorted(cam_dir.glob("congestion_timeseries_*.csv"))[-7:])
        out = self.root_dir / "data" / "reports" / "daily" / f"multi_day_trend_{date.today().isoformat()}.png"
        save_multi_day_trend_plot(metrics_files, "congestion_score", out)

    def _request_report_update(self) -> None:
        try:
            self.reporter.write_daily_report(date.today(), self.app_cfg.cameras, self.root_dir / "data" / "metrics")
            self.pending_report_update = False
            self.report_warning_shown = False
        except PermissionError:
            self.pending_report_update = True
            if not self.report_warning_shown:
                self.report_warning_shown = True
                QtWidgets.QMessageBox.warning(
                    self,
                    "Excel更新保留",
                    "Excelレポートが開かれているため更新できません。Excelを閉じてください。閉じられ次第、自動で再更新します。",
                )
        except Exception as exc:
            self.pending_report_update = True
            print(f"[WARN] report update skipped: {exc}")

    def _flush_report_update_if_needed(self) -> None:
        if not self.pending_report_update:
            return
        self._request_report_update()

    def _retry_pending_report_update(self) -> None:
        if self.pending_report_update:
            self._request_report_update()

    def closeEvent(self, event: QtGui.QCloseEvent) -> None:
        for worker in self.workers.values():
            try:
                worker.stop()
            except Exception:
                pass

        for thread in self.threads.values():
            thread.quit()
            thread.wait(3000)

        super().closeEvent(event)


MonitorMainWindow = MainWindow


# =========================================
# parse_args / main
# =========================================
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="3-camera AI congestion monitor")
    p.add_argument("--root", default=str(Path(__file__).resolve().parent), help="ai_monitor root directory")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    root_dir = Path(args.root)
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow(root_dir)
    win.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
