from __future__ import annotations

# ================== imports ==================
import argparse
import csv
import importlib
import importlib.metadata
import inspect
import json
import os
import platform
import sys
import tempfile
import time
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PyQt6 import QtCore, QtGui, QtWidgets


# ================== safe runtime helpers ==================
def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}")


def package_version(name: str) -> str:
    try:
        return importlib.metadata.version(name)
    except Exception:
        return "not-installed"


def lazy_import_torch() -> Any:
    return importlib.import_module("torch")


def probe_runtime_environment(skip_torch: bool = False) -> dict[str, str]:
    info = {
        "python_version": platform.python_version(),
        "platform": platform.platform(),
        "opencv_version": getattr(cv2, "__version__", "unknown"),
        "ultralytics_version": package_version("ultralytics"),
        "boxmot_version": package_version("boxmot"),
        "torch_version": "not-loaded",
        "cuda_available": "unknown",
        "gpu_name": "N/A",
        "compute_capability": "N/A",
        "runtime_device": "cpu",
        "probe_error": "",
    }
    if skip_torch:
        return info
    try:
        torch = lazy_import_torch()
        info["torch_version"] = getattr(torch, "__version__", "unknown")
        cuda_available = bool(torch.cuda.is_available())
        info["cuda_available"] = str(cuda_available)
        if cuda_available:
            info["gpu_name"] = torch.cuda.get_device_name(0)
            cc = torch.cuda.get_device_capability(0)
            info["compute_capability"] = f"{cc[0]}.{cc[1]}"
            info["runtime_device"] = "cuda:0"
    except Exception as exc:
        info["probe_error"] = f"torch probe failed: {exc}"
    return info


def choose_device(preference: str, safe_mode: bool = False) -> tuple[str, str, str]:
    """Return (device, gpu_name, reason)."""
    if preference == "cpu" or safe_mode:
        return "cpu", "CPU", "forced cpu"
    try:
        torch = lazy_import_torch()
        if not torch.cuda.is_available():
            return "cpu", "CPU", "cuda unavailable"
        gpu_name = torch.cuda.get_device_name(0)
        major, minor = torch.cuda.get_device_capability(0)
        if major * 10 + minor < 50:
            return "cpu", gpu_name, "legacy gpu compute capability"
        return "cuda:0", gpu_name, "cuda available"
    except Exception as exc:
        return "cpu", "CPU", f"cuda probe failed: {exc}"


# ================== config ==================
DEFAULT_CONFIG = {
    "system_config": {
        "model_path": "yolo11n.pt",
        "device_preference": "auto",
        "target_classes": ["car", "bus", "truck", "motorcycle"],
        "metrics_save_interval_sec": 5,
        "ui_refresh_interval_ms": 500,
        "ai_status_json_path": "app/config/ai_status.json",
    },
    "camera_settings": {
        "cameras": [
            {
                "camera_id": 1,
                "camera_name": "Camera1",
                "stream_url": "0",
                "enabled": True,
                "direction": "LtoR",
                "line_points": [[80, 200], [520, 200]],
                "exclude_polygon": [],
                "congestion_threshold": 60,
                "long_stay_minutes": 15,
                "long_stay_trigger_count": 1,
            },
            {
                "camera_id": 2,
                "camera_name": "Camera2",
                "stream_url": "0",
                "enabled": True,
                "direction": "LtoR",
                "line_points": [[80, 200], [520, 200]],
                "exclude_polygon": [],
                "congestion_threshold": 60,
                "long_stay_minutes": 15,
                "long_stay_trigger_count": 1,
            },
            {
                "camera_id": 3,
                "camera_name": "Camera3",
                "stream_url": "0",
                "enabled": True,
                "direction": "LtoR",
                "line_points": [[80, 200], [520, 200]],
                "exclude_polygon": [],
                "congestion_threshold": 60,
                "long_stay_minutes": 15,
                "long_stay_trigger_count": 1,
            },
        ]
    },
}


class ConfigStore:
    def __init__(self, config_path: Path):
        self.config_path = config_path

    def ensure(self) -> None:
        if self.config_path.exists():
            return
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        self.save(DEFAULT_CONFIG)

    def load(self) -> dict[str, Any]:
        self.ensure()
        return json.loads(self.config_path.read_text(encoding="utf-8"))

    def save(self, payload: dict[str, Any]) -> None:
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        self.config_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def atomic_update_ai_status(ai_status_path: Path, level: int) -> None:
    payload = {
        "congestion_level": int(level),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    ai_status_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=ai_status_path.parent, delete=False) as tmp:
        json.dump(payload, tmp, ensure_ascii=False, indent=2)
        tmp.flush()
        os.fsync(tmp.fileno())
        tmp_path = Path(tmp.name)
    tmp_path.replace(ai_status_path)


# ================== AI backend loader ==================
class AIBackendError(RuntimeError):
    pass


class AIBackendLoader:
    def __init__(self, model_path: str, device_preference: str, safe_mode: bool = False):
        self.model_path = model_path
        self.device_preference = device_preference
        self.safe_mode = safe_mode
        self.model: Any = None
        self.tracker_factory: Any = None
        self.device = "cpu"
        self.gpu_name = "CPU"
        self.reason = "not loaded"

    def load(self) -> None:
        self.device, self.gpu_name, self.reason = choose_device(self.device_preference, self.safe_mode)
        try:
            ultralytics = importlib.import_module("ultralytics")
            yolo_cls = getattr(ultralytics, "YOLO")
            self.model = yolo_cls(self.model_path)
            self.tracker_factory = self._build_bytetrack_factory()
        except Exception as exc:
            raise AIBackendError(f"YOLO initialization failed: {exc}") from exc

    def _build_bytetrack_factory(self):
        try:
            boxmot = importlib.import_module("boxmot")
            bt_cls = getattr(boxmot, "ByteTrack", None)
            if bt_cls is None:
                bt_module = importlib.import_module("boxmot.trackers.byte_tracker.byte_tracker")
                bt_cls = getattr(bt_module, "ByteTrack")
            sig = inspect.signature(bt_cls.__init__)
            valid_keys = set(sig.parameters.keys())

            def create(frame_rate: float = 30.0) -> Any:
                kwargs = {
                    "track_thresh": 0.5,
                    "match_thresh": 0.8,
                    "track_buffer": 30,
                    "min_conf": 0.1,
                    "frame_rate": frame_rate,
                }
                filtered = {k: v for k, v in kwargs.items() if k in valid_keys}
                return bt_cls(**filtered)

            log(f"ByteTrack signature keys: {sorted(valid_keys)}")
            return create
        except Exception as exc:
            log(f"ByteTrack unavailable: {exc}")
            return None


# ================== tracking / counting logic ==================
@dataclass
class TrackMemory:
    centers: dict[int, tuple[float, float]] = field(default_factory=dict)
    first_seen: dict[int, datetime] = field(default_factory=dict)
    count_once: set[int] = field(default_factory=set)
    long_stay_emitted: set[int] = field(default_factory=set)


class CounterLogic:
    def __init__(self, direction: str, line_points: list[list[int]]):
        self.direction = direction
        self.line_points = line_points
        self.hist_10min = [0] * 144

    def update_cross(self, track_id: int, prev_center: tuple[float, float], center: tuple[float, float], now: datetime) -> bool:
        lx = (self.line_points[0][0] + self.line_points[1][0]) / 2
        crossed = (prev_center[0] < lx <= center[0]) if self.direction == "LtoR" else (prev_center[0] > lx >= center[0])
        if crossed:
            idx = (now.hour * 60 + now.minute) // 10
            if 0 <= idx < 144:
                self.hist_10min[idx] += 1
        return crossed


# ================== congestion logic ==================
def compute_congestion_score(active_tracks: int, avg_movement: float, long_stay_count: int) -> float:
    movement_penalty = max(0.0, min(1.0, 1.0 - (avg_movement / 35.0)))
    track_term = min(60.0, active_tracks * 8.0)
    movement_term = movement_penalty * 30.0
    stay_term = min(20.0, long_stay_count * 7.5)
    return max(0.0, min(100.0, track_term + movement_term + stay_term))


def decide_global_congestion_level(payload_by_camera: dict[int, dict[str, Any]]) -> int:
    cam1 = payload_by_camera.get(1, {})
    cam2 = payload_by_camera.get(2, {})
    cam3 = payload_by_camera.get(3, {})
    if bool(cam3.get("threshold_over", False)):
        return 3
    if bool(cam1.get("threshold_over", False)):
        return 1
    if int(cam2.get("long_stay_count", 0)) >= 1:
        return 2
    return 0


# ================== report writer ==================
class ReportWriter:
    def __init__(self, base_dir: Path):
        self.base_dir = base_dir
        self.daily_dir = self.base_dir / "reports" / "daily"
        self.monthly_dir = self.base_dir / "reports" / "monthly"
        self.daily_dir.mkdir(parents=True, exist_ok=True)
        self.monthly_dir.mkdir(parents=True, exist_ok=True)

    def _read_csv(self, path: Path) -> list[dict[str, str]]:
        if not path.exists():
            return []
        with path.open("r", encoding="utf-8") as f:
            return list(csv.DictReader(f))

    def write_daily_report(self, target_date: date, metrics_root: Path) -> Path:
        day = target_date.strftime("%Y-%m-%d")
        wb = Workbook()
        ws = wb.active
        ws.title = "summary"
        ws.append(["Date", day])
        ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        ws.append(["Camera", "Pass Count", "Max Congestion", "Long Stay Events"])

        for cam_id in [1, 2, 3]:
            cam_dir = metrics_root / f"cam{cam_id}"
            pass_rows = self._read_csv(cam_dir / f"pass_events_{day}.csv")
            metrics_rows = self._read_csv(cam_dir / f"realtime_metrics_{day}.csv")
            long_rows = self._read_csv(cam_dir / f"long_stay_events_{day}.csv")
            max_cong = max([float(r.get("congestion_score", 0)) for r in metrics_rows], default=0)
            ws.append([f"cam{cam_id}", len(pass_rows), round(max_cong, 2), len(long_rows)])

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 24
        for row in ws.iter_rows(min_row=1, max_col=5):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws["A1"].font = Font(bold=True, size=14)

        out = self.daily_dir / f"daily_report_{day}.xlsx"
        wb.save(out)
        return out

    def write_monthly_report(self, month: str, metrics_root: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "monthly"
        ws.append(["Month", month])
        ws.append(["Date", "Total Pass", "Max Congestion", "Long Stay Events"])

        for day_dir in sorted(metrics_root.glob("cam1/realtime_metrics_*.csv")):
            day = day_dir.stem.split("_")[-1]
            total_pass = 0
            max_cong = 0.0
            total_long = 0
            for cam_id in [1, 2, 3]:
                cam_dir = metrics_root / f"cam{cam_id}"
                total_pass += len(self._read_csv(cam_dir / f"pass_events_{day}.csv"))
                rows = self._read_csv(cam_dir / f"realtime_metrics_{day}.csv")
                max_cong = max(max_cong, max([float(r.get("congestion_score", 0)) for r in rows], default=0))
                total_long += len(self._read_csv(cam_dir / f"long_stay_events_{day}.csv"))
            if day.startswith(month):
                ws.append([day, total_pass, round(max_cong, 2), total_long])

        out = self.monthly_dir / f"monthly_report_{month}.xlsx"
        wb.save(out)
        return out


def save_multi_day_plot(metrics_csv_paths: list[Path], output_png: Path) -> Optional[Path]:
    # 【脇村モデル】多日集計可視化
    try:
        import matplotlib.pyplot as plt
        import pandas as pd
    except Exception as exc:
        log(f"plot skipped: {exc}")
        return None

    frames = []
    for p in metrics_csv_paths:
        if p.exists():
            df = pd.read_csv(p)
            if "timestamp" in df.columns and "congestion_score" in df.columns:
                df["timestamp"] = pd.to_datetime(df["timestamp"])
                df["minute"] = df["timestamp"].dt.hour * 60 + df["timestamp"].dt.minute
                frames.append(df)
    if not frames:
        return None

    all_df = pd.concat(frames, ignore_index=True)
    # 【脇村モデル】平均トレンド
    mean_trend = all_df.groupby("minute")["congestion_score"].mean()
    # 【脇村モデル】中央値トレンド
    median_trend = all_df.groupby("minute")["congestion_score"].median()

    plt.figure(figsize=(10, 5))
    plt.plot(mean_trend.index, mean_trend.values, label="Mean", color="cyan")
    plt.plot(median_trend.index, median_trend.values, label="Median", color="yellow")
    plt.grid(True, alpha=0.3)
    plt.legend()
    output_png.parent.mkdir(parents=True, exist_ok=True)
    plt.savefig(output_png, dpi=160)
    plt.close()
    return output_png


# ================== UI widgets ==================
class CameraPanel(QtWidgets.QFrame):
    def __init__(self, camera_cfg: dict[str, Any]):
        super().__init__()
        self.camera_cfg = camera_cfg
        self.setStyleSheet("QFrame {background:#050b12; border:1px solid #00bcd4; color:#c8f8ff;}")
        root = QtWidgets.QHBoxLayout(self)

        self.video_label = QtWidgets.QLabel("No Signal")
        self.video_label.setMinimumSize(640, 240)
        self.video_label.setStyleSheet("background:#000; color:#88ddee; font-size:20px;")
        self.video_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        root.addWidget(self.video_label, 3)

        right = QtWidgets.QVBoxLayout()
        self.cam_title = QtWidgets.QLabel(camera_cfg.get("camera_name", "camera"))
        self.cam_title.setStyleSheet("color:#00e5ff; font-weight:bold; font-size:16px;")
        right.addWidget(self.cam_title)

        self.gauge = QtWidgets.QProgressBar()
        self.gauge.setRange(0, 100)
        self.gauge.setFormat("Congestion %p")
        right.addWidget(self.gauge)

        self.hist_text = QtWidgets.QLabel("10min pass histogram (prev/today)")
        self.hist_text.setStyleSheet("color:#8cf;")
        right.addWidget(self.hist_text)

        self.long_stay = QtWidgets.QTableWidget(0, 2)
        self.long_stay.setHorizontalHeaderLabels(["track_id", "stay_min"])
        self.long_stay.horizontalHeader().setStretchLastSection(True)
        right.addWidget(self.long_stay, 1)

        self.meta = QtWidgets.QLabel("device: cpu | gpu: CPU | fps: 0")
        self.meta.setStyleSheet("color:#ffe082;")
        right.addWidget(self.meta)
        root.addLayout(right, 2)

    def update_payload(self, payload: dict[str, Any]) -> None:
        score = int(payload.get("congestion_score", 0))
        th = int(payload.get("threshold", 60))
        self.gauge.setValue(score)
        chunk = "#ff1744" if score >= th else "#00e5ff"
        self.gauge.setStyleSheet(f"QProgressBar::chunk{{background:{chunk};}}")
        if payload.get("frame") is not None:
            self.video_label.setPixmap(payload["frame"])

        prev_hist = payload.get("hist_prev", [0] * 144)
        today_hist = payload.get("hist_today", [0] * 144)
        self.hist_text.setText(f"Prev total: {sum(prev_hist)} / Today total: {sum(today_hist)}")

        items = payload.get("long_stays", [])
        self.long_stay.setRowCount(len(items))
        for r, (tid, minutes) in enumerate(items):
            self.long_stay.setItem(r, 0, QtWidgets.QTableWidgetItem(str(tid)))
            self.long_stay.setItem(r, 1, QtWidgets.QTableWidgetItem(f"{minutes:.1f}"))

        self.meta.setText(
            f"{payload.get('camera_name','')} | {datetime.now().strftime('%H:%M:%S')} | "
            f"device: {payload.get('device','cpu')} | gpu: {payload.get('gpu_name','CPU')} | fps: {payload.get('fps',0):.1f}"
        )


class CameraSettingsDialog(QtWidgets.QDialog):
    def __init__(self, camera_cfg: dict[str, Any], parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Camera Settings: {camera_cfg.get('camera_name', '')}")
        self.camera_cfg = dict(camera_cfg)
        layout = QtWidgets.QFormLayout(self)

        self.name_edit = QtWidgets.QLineEdit(self.camera_cfg.get("camera_name", ""))
        self.url_edit = QtWidgets.QLineEdit(self.camera_cfg.get("stream_url", ""))
        self.enabled_check = QtWidgets.QCheckBox()
        self.enabled_check.setChecked(bool(self.camera_cfg.get("enabled", True)))
        self.dir_combo = QtWidgets.QComboBox()
        self.dir_combo.addItems(["LtoR", "RtoL"])
        self.dir_combo.setCurrentText(self.camera_cfg.get("direction", "LtoR"))
        self.line_edit = QtWidgets.QLineEdit(json.dumps(self.camera_cfg.get("line_points", [[0, 0], [100, 0]]), ensure_ascii=False))
        self.exclude_edit = QtWidgets.QLineEdit(json.dumps(self.camera_cfg.get("exclude_polygon", []), ensure_ascii=False))
        self.th_spin = QtWidgets.QSpinBox(); self.th_spin.setRange(0, 100); self.th_spin.setValue(int(self.camera_cfg.get("congestion_threshold", 60)))
        self.stay_spin = QtWidgets.QSpinBox(); self.stay_spin.setRange(1, 120); self.stay_spin.setValue(int(self.camera_cfg.get("long_stay_minutes", 15)))

        layout.addRow("camera_name", self.name_edit)
        layout.addRow("stream_url", self.url_edit)
        layout.addRow("enabled", self.enabled_check)
        layout.addRow("direction", self.dir_combo)
        layout.addRow("line_points(JSON)", self.line_edit)
        layout.addRow("exclude_polygon(JSON)", self.exclude_edit)
        layout.addRow("congestion_threshold", self.th_spin)
        layout.addRow("long_stay_minutes", self.stay_spin)

        btn_capture = QtWidgets.QPushButton("静止画取得")
        btn_capture.clicked.connect(self.capture_image)
        layout.addRow(btn_capture)

        btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Ok | QtWidgets.QDialogButtonBox.StandardButton.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addRow(btn_box)

    def capture_image(self) -> None:
        src = self.url_edit.text().strip() or "0"
        source = int(src) if src.isdigit() else src
        cap = cv2.VideoCapture(source)
        ok, frame = cap.read()
        cap.release()
        if not ok:
            QtWidgets.QMessageBox.warning(self, "Capture", "stream に接続できませんでした")
            return
        out = Path("app/ai_monitor/data/snapshots") / f"snapshot_cam{self.camera_cfg.get('camera_id')}_{int(time.time())}.jpg"
        out.parent.mkdir(parents=True, exist_ok=True)
        cv2.imwrite(str(out), frame)
        QtWidgets.QMessageBox.information(self, "Capture", f"保存: {out}")

    def updated(self) -> dict[str, Any]:
        self.camera_cfg["camera_name"] = self.name_edit.text().strip()
        self.camera_cfg["stream_url"] = self.url_edit.text().strip()
        self.camera_cfg["enabled"] = self.enabled_check.isChecked()
        self.camera_cfg["direction"] = self.dir_combo.currentText()
        self.camera_cfg["line_points"] = json.loads(self.line_edit.text() or "[[0,0],[100,0]]")
        self.camera_cfg["exclude_polygon"] = json.loads(self.exclude_edit.text() or "[]")
        self.camera_cfg["congestion_threshold"] = self.th_spin.value()
        self.camera_cfg["long_stay_minutes"] = self.stay_spin.value()
        return self.camera_cfg


class CameraEngine:
    def __init__(self, camera_cfg: dict[str, Any], system_cfg: dict[str, Any], data_root: Path, no_ai: bool, force_cpu: bool, safe_mode: bool):
        self.camera_cfg = camera_cfg
        self.system_cfg = system_cfg
        self.data_root = data_root
        self.no_ai = no_ai
        self.force_cpu = force_cpu
        self.safe_mode = safe_mode
        self.cap: Optional[cv2.VideoCapture] = None
        self.backend: Optional[AIBackendLoader] = None
        self.counter = CounterLogic(camera_cfg.get("direction", "LtoR"), camera_cfg.get("line_points", [[0, 0], [100, 0]]))
        self.track_memory = TrackMemory()
        self.prev_hist = [0] * 144
        self.last_fps = 0.0
        self.load_prev_hist()
        self.realtime_csv, self.pass_csv, self.long_csv = self.ensure_daily_csvs()

    def ensure_daily_csvs(self) -> tuple[Path, Path, Path]:
        today = date.today().strftime("%Y-%m-%d")
        cam_dir = self.data_root / "metrics" / f"cam{self.camera_cfg['camera_id']}"
        cam_dir.mkdir(parents=True, exist_ok=True)
        realtime = cam_dir / f"realtime_metrics_{today}.csv"
        pass_csv = cam_dir / f"pass_events_{today}.csv"
        long_csv = cam_dir / f"long_stay_events_{today}.csv"
        for path, header in [
            (realtime, ["timestamp", "camera_id", "camera_name", "active_tracks", "congestion_score", "congestion_threshold", "threshold_over", "long_stay_count", "fps", "runtime_device", "gpu_name"]),
            (pass_csv, ["timestamp", "camera_id", "track_id", "class_name", "direction"]),
            (long_csv, ["first_seen", "detected_at", "camera_id", "track_id", "stay_minutes", "class_name"]),
        ]:
            if not path.exists():
                with path.open("w", newline="", encoding="utf-8") as f:
                    csv.writer(f).writerow(header)
        return realtime, pass_csv, long_csv

    def load_prev_hist(self) -> None:
        prev = (date.today().fromordinal(date.today().toordinal() - 1)).strftime("%Y-%m-%d")
        pass_csv = self.data_root / "metrics" / f"cam{self.camera_cfg['camera_id']}" / f"pass_events_{prev}.csv"
        if not pass_csv.exists():
            return
        with pass_csv.open("r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                ts = datetime.strptime(row["timestamp"], "%Y-%m-%d %H:%M:%S")
                self.prev_hist[(ts.hour * 60 + ts.minute) // 10] += 1

    def connect(self) -> None:
        src = str(self.camera_cfg.get("stream_url", "0"))
        source = int(src) if src.isdigit() else src
        self.cap = cv2.VideoCapture(source)

    def ensure_backend(self) -> Optional[str]:
        if self.no_ai:
            return None
        if self.backend is not None:
            return None
        preference = "cpu" if self.force_cpu else self.system_cfg.get("device_preference", "auto")
        self.backend = AIBackendLoader(self.system_cfg.get("model_path", "yolo11n.pt"), preference, self.safe_mode)
        try:
            self.backend.load()
            return None
        except Exception as exc:
            self.backend = None
            return str(exc)

    def process_once(self) -> dict[str, Any]:
        st = time.time()
        if self.cap is None or not self.cap.isOpened():
            self.connect()
        if self.cap is None:
            return self.empty_payload("camera not initialized")
        ok, frame = self.cap.read()
        if not ok:
            time.sleep(0.2)
            self.connect()
            return self.empty_payload("camera reconnecting")

        err = self.ensure_backend()
        if err:
            return self.empty_payload(f"AI disabled: {err}", frame)

        tracks: list[dict[str, Any]] = []
        if self.backend and self.backend.model is not None:
            try:
                result = self.backend.model.track(frame, persist=True, verbose=False, device=self.backend.device, tracker="bytetrack.yaml")[0]
                boxes = getattr(result, "boxes", None)
                if boxes is not None and boxes.id is not None:
                    ids = boxes.id.cpu().numpy()
                    xyxy = boxes.xyxy.cpu().numpy()
                    cls_arr = boxes.cls.cpu().numpy() if boxes.cls is not None else []
                    for i, (tid, box) in enumerate(zip(ids, xyxy)):
                        x1, y1, x2, y2 = [int(v) for v in box.tolist()]
                        center = ((x1 + x2) / 2.0, (y1 + y2) / 2.0)
                        cls_idx = int(cls_arr[i]) if len(cls_arr) > i else -1
                        name_map = getattr(self.backend.model, "names", {})
                        cls_name = name_map.get(cls_idx, str(cls_idx)) if isinstance(name_map, dict) else str(cls_idx)
                        tracks.append({"track_id": int(tid), "center": center, "bbox": (x1, y1, x2, y2), "class_name": cls_name})
            except Exception as exc:
                return self.empty_payload(f"AI process failed: {exc}", frame)

        now = datetime.now()
        pass_events = []
        long_stays = []
        move_values = []
        for tr in tracks:
            tid = tr["track_id"]
            center = tr["center"]
            prev = self.track_memory.centers.get(tid)
            if prev:
                move = float(np.hypot(center[0] - prev[0], center[1] - prev[1]))
                move_values.append(move)
                crossed = self.counter.update_cross(tid, prev, center, now)
                if crossed and tid not in self.track_memory.count_once:
                    self.track_memory.count_once.add(tid)
                    pass_events.append([now.strftime("%Y-%m-%d %H:%M:%S"), self.camera_cfg["camera_id"], tid, tr["class_name"], self.camera_cfg.get("direction", "LtoR")])
            self.track_memory.centers[tid] = center
            self.track_memory.first_seen.setdefault(tid, now)

            stay_min = (now - self.track_memory.first_seen[tid]).total_seconds() / 60.0
            if stay_min >= float(self.camera_cfg.get("long_stay_minutes", 15)):
                long_stays.append((tid, stay_min))

        long_stay_events = []
        for tid, stay_min in long_stays:
            if tid in self.track_memory.long_stay_emitted:
                continue
            self.track_memory.long_stay_emitted.add(tid)
            first_seen = self.track_memory.first_seen[tid].strftime("%Y-%m-%d %H:%M:%S")
            long_stay_events.append([first_seen, now.strftime("%Y-%m-%d %H:%M:%S"), self.camera_cfg["camera_id"], tid, round(stay_min, 1), "vehicle"])

        avg_move = float(np.mean(move_values)) if move_values else 0.0
        score = compute_congestion_score(len(tracks), avg_move, len(long_stays))
        threshold = float(self.camera_cfg.get("congestion_threshold", 60))
        over = score >= threshold
        self.last_fps = 1.0 / max(1e-6, time.time() - st)

        self.append_rows(self.pass_csv, pass_events)
        self.append_rows(self.long_csv, long_stay_events)
        self.append_rows(self.realtime_csv, [[
            now.strftime("%Y-%m-%d %H:%M:%S"), self.camera_cfg["camera_id"], self.camera_cfg.get("camera_name", ""),
            len(tracks), round(score, 2), threshold, int(over), len(long_stays), round(self.last_fps, 2),
            self.backend.device if self.backend else "cpu", self.backend.gpu_name if self.backend else "CPU",
        ]])

        draw = self.draw(frame, tracks)
        pix = self.to_pixmap(draw)
        return {
            "camera_id": self.camera_cfg["camera_id"],
            "camera_name": self.camera_cfg.get("camera_name", ""),
            "frame": pix,
            "congestion_score": score,
            "threshold": threshold,
            "threshold_over": over,
            "hist_today": self.counter.hist_10min,
            "hist_prev": self.prev_hist,
            "long_stays": sorted(long_stays, key=lambda x: x[1], reverse=True)[:10],
            "long_stay_count": len(long_stays),
            "fps": self.last_fps,
            "device": self.backend.device if self.backend else "cpu",
            "gpu_name": self.backend.gpu_name if self.backend else "CPU",
            "error": "",
        }

    def append_rows(self, path: Path, rows: list[list[Any]]) -> None:
        if not rows:
            return
        with path.open("a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerows(rows)

    def draw(self, frame: np.ndarray, tracks: list[dict[str, Any]]) -> np.ndarray:
        line = self.camera_cfg.get("line_points", [[80, 200], [520, 200]])
        cv2.line(frame, tuple(line[0]), tuple(line[1]), (0, 255, 255), 2)
        poly = self.camera_cfg.get("exclude_polygon", [])
        if len(poly) >= 3:
            cv2.polylines(frame, [np.array(poly, np.int32)], True, (255, 0, 255), 2)
        for tr in tracks:
            x1, y1, x2, y2 = tr["bbox"]
            cv2.rectangle(frame, (x1, y1), (x2, y2), (255, 255, 0), 2)
            cv2.putText(frame, f"id={tr['track_id']}", (x1, max(10, y1 - 6)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)
        return frame

    def to_pixmap(self, frame: np.ndarray) -> QtGui.QPixmap:
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb.shape
        qimg = QtGui.QImage(rgb.data, w, h, ch * w, QtGui.QImage.Format.Format_RGB888)
        return QtGui.QPixmap.fromImage(qimg).scaled(640, 240, QtCore.Qt.AspectRatioMode.KeepAspectRatio)

    def empty_payload(self, reason: str, frame: Optional[np.ndarray] = None) -> dict[str, Any]:
        pix = None
        if frame is not None:
            pix = self.to_pixmap(frame)
        return {
            "camera_id": self.camera_cfg["camera_id"],
            "camera_name": self.camera_cfg.get("camera_name", ""),
            "frame": pix,
            "congestion_score": 0,
            "threshold": float(self.camera_cfg.get("congestion_threshold", 60)),
            "threshold_over": False,
            "hist_today": self.counter.hist_10min,
            "hist_prev": self.prev_hist,
            "long_stays": [],
            "long_stay_count": 0,
            "fps": 0.0,
            "device": "cpu",
            "gpu_name": "CPU",
            "error": reason,
        }


# ================== main window ==================
class MonitorMainWindow(QtWidgets.QMainWindow):
    def __init__(self, root_dir: Path, config_path: Path, probe_info: dict[str, str], no_ai: bool, force_cpu: bool, safe_mode: bool):
        super().__init__()
        self.root_dir = root_dir
        self.no_ai = no_ai
        self.force_cpu = force_cpu
        self.safe_mode = safe_mode
        self.setWindowTitle("AI Congestion Monitor Single")
        self.setStyleSheet("QMainWindow{background:#01050b; color:#d0f8ff;}")

        self.config_store = ConfigStore(config_path)
        self.config = self.config_store.load()
        self.system_cfg = self.config.get("system_config", {})
        self.cameras = self.config.get("camera_settings", {}).get("cameras", [])
        self.ai_status_path = Path(self.system_cfg.get("ai_status_json_path", "app/config/ai_status.json"))
        if not self.ai_status_path.is_absolute():
            self.ai_status_path = self.root_dir / self.ai_status_path

        self.data_root = self.root_dir / "app/ai_monitor/data"
        self.data_root.mkdir(parents=True, exist_ok=True)
        self.report_writer = ReportWriter(self.data_root)

        central = QtWidgets.QWidget(); self.setCentralWidget(central)
        layout = QtWidgets.QVBoxLayout(central)

        header = QtWidgets.QHBoxLayout()
        self.runtime_label = QtWidgets.QLabel(self._format_probe(probe_info))
        self.runtime_label.setStyleSheet("color:#b2ebf2;")
        header.addWidget(self.runtime_label, 1)
        for text, cb in [
            ("設定", self.open_settings),
            ("日次Excel", self.export_daily),
            ("月次Excel", self.export_monthly),
            ("多日Plot", self.export_plot),
        ]:
            btn = QtWidgets.QPushButton(text)
            btn.clicked.connect(cb)
            header.addWidget(btn)
        layout.addLayout(header)

        self.engines: dict[int, CameraEngine] = {}
        self.panels: dict[int, CameraPanel] = {}
        for cam in self.cameras:
            panel = CameraPanel(cam)
            layout.addWidget(panel, 1)
            self.panels[int(cam["camera_id"])] = panel
            if cam.get("enabled", True):
                self.engines[int(cam["camera_id"])] = CameraEngine(cam, self.system_cfg, self.data_root, no_ai, force_cpu, safe_mode)

        self.timer = QtCore.QTimer(self)
        self.timer.timeout.connect(self.tick)
        self.timer.start(int(self.system_cfg.get("ui_refresh_interval_ms", 500)))
        self.showMaximized()

    def _format_probe(self, probe: dict[str, str]) -> str:
        return (
            f"Python {probe.get('python_version')} | torch {probe.get('torch_version')} | "
            f"CUDA {probe.get('cuda_available')} | GPU {probe.get('gpu_name')} | "
            f"OpenCV {probe.get('opencv_version')} | ultralytics {probe.get('ultralytics_version')}"
        )

    def tick(self) -> None:
        payloads: dict[int, dict[str, Any]] = {}
        for cid, engine in self.engines.items():
            try:
                payload = engine.process_once()
                payloads[cid] = payload
                self.panels[cid].update_payload(payload)
                if payload.get("error"):
                    self.panels[cid].video_label.setText(payload["error"])
            except Exception as exc:
                self.panels[cid].video_label.setText(f"camera {cid} error: {exc}")
        level = decide_global_congestion_level(payloads)
        try:
            atomic_update_ai_status(self.ai_status_path, level)
        except Exception as exc:
            log(f"ai_status update failed: {exc}")

    def open_settings(self) -> None:
        names = [f"{c['camera_id']}: {c['camera_name']}" for c in self.cameras]
        selected, ok = QtWidgets.QInputDialog.getItem(self, "設定カメラ", "対象", names, 0, False)
        if not ok:
            return
        cid = int(selected.split(":", 1)[0])
        idx = next(i for i, c in enumerate(self.cameras) if int(c["camera_id"]) == cid)
        dlg = CameraSettingsDialog(self.cameras[idx], self)
        if dlg.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            self.cameras[idx] = dlg.updated()
            self.config["camera_settings"]["cameras"] = self.cameras
            self.config_store.save(self.config)
            QtWidgets.QMessageBox.information(self, "保存", "設定保存完了。再起動で完全反映")

    def export_daily(self) -> None:
        out = self.report_writer.write_daily_report(date.today(), self.data_root / "metrics")
        QtWidgets.QMessageBox.information(self, "日次", f"保存: {out}")

    def export_monthly(self) -> None:
        month = datetime.now().strftime("%Y-%m")
        out = self.report_writer.write_monthly_report(month, self.data_root / "metrics")
        QtWidgets.QMessageBox.information(self, "月次", f"保存: {out}")

    def export_plot(self) -> None:
        files = []
        for cam in [1, 2, 3]:
            cam_dir = self.data_root / "metrics" / f"cam{cam}"
            files.extend(sorted(cam_dir.glob("realtime_metrics_*.csv"))[-7:])
        out = self.data_root / "reports" / "daily" / f"multi_day_trend_{date.today().isoformat()}.png"
        r = save_multi_day_plot(files, out)
        if r:
            QtWidgets.QMessageBox.information(self, "plot", f"保存: {r}")
        else:
            QtWidgets.QMessageBox.warning(self, "plot", "plot 出力に必要なデータが不足")


# ================== entry point ==================
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="AI congestion monitor (single file)")
    p.add_argument("--root", default=str(Path(__file__).resolve().parents[2]))
    p.add_argument("--config", default="app/ai_monitor/config_single.json")
    p.add_argument("--probe-device", action="store_true")
    p.add_argument("--no-ai", action="store_true")
    p.add_argument("--cpu", action="store_true")
    p.add_argument("--safe-mode", action="store_true")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    root = Path(args.root)
    config_path = Path(args.config)
    if not config_path.is_absolute():
        config_path = root / config_path

    probe = probe_runtime_environment(skip_torch=bool(args.no_ai))
    log(f"Runtime probe: {json.dumps(probe, ensure_ascii=False)}")

    if args.probe_device:
        print(json.dumps(probe, ensure_ascii=False, indent=2))
        return 0

    app = QtWidgets.QApplication(sys.argv)
    window = MonitorMainWindow(root, config_path, probe, no_ai=args.no_ai, force_cpu=args.cpu, safe_mode=args.safe_mode)
    window.show()

    if probe.get("probe_error"):
        QtWidgets.QMessageBox.warning(window, "Runtime probe warning", probe["probe_error"])

    return app.exec()


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        traceback.print_exc()
        app = QtWidgets.QApplication.instance() or QtWidgets.QApplication(sys.argv)
        QtWidgets.QMessageBox.critical(None, "起動失敗", f"起動に失敗しました:\n{exc}")
        raise
