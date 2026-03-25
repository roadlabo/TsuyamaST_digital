from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


class ReportWriter:
    def __init__(self, output_root: Path):
        self.output_root = output_root

    def write_daily_report(self, target_date: date, cameras: list[dict], metrics_root: Path) -> Path:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "summary"
        self._style_sheet(ws_summary)

        ws_summary["A1"] = f"Daily Report {target_date.isoformat()}"
        ws_summary["A1"].font = Font(bold=True, size=16, color="00D7FF")

        row = 3
        headers = [
            "camera", "total_pass", "max_congestion", "over_threshold_points", "long_stay_events",
        ]
        ws_summary.append(headers)

        total_pass_all = 0
        for cam in cameras:
            cid = cam["camera_id"]
            date_str = target_date.isoformat()
            cam_dir = metrics_root / f"cam{cid}"
            pass_df = self._safe_read(cam_dir / f"pass_events_{date_str}.csv")
            metric_df = self._safe_read(cam_dir / f"realtime_metrics_{date_str}.csv")
            long_df = self._safe_read(cam_dir / f"long_stay_events_{date_str}.csv")

            total_pass = len(pass_df)
            total_pass_all += total_pass
            max_cong = float(metric_df["congestion_score"].max()) if not metric_df.empty else 0.0
            over_count = int(metric_df["threshold_over"].sum()) if "threshold_over" in metric_df.columns else 0
            long_count = len(long_df)
            ws_summary.append([cam["camera_name"], total_pass, round(max_cong, 1), over_count, long_count])

            self._add_camera_sheet(wb, cam, target_date, pass_df, metric_df, long_df)
            row += 1

        ws_summary["A2"] = "date"
        ws_summary["B2"] = target_date.isoformat()
        ws_summary["A3"] = "all_cameras_total_pass"
        ws_summary["B3"] = total_pass_all

        out_path = self.output_root / "reports" / "daily" / f"daily_report_{target_date.isoformat()}.xlsx"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def write_monthly_report(self, target_month: str, metrics_root: Path, cameras: list[dict]) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "monthly"
        self._style_sheet(ws)
        ws.append(["date", "total_pass", "max_congestion", "long_stay_events"])

        daily_map: dict[str, dict[str, float]] = {}
        for cam in cameras:
            cam_dir = metrics_root / f"cam{cam['camera_id']}"
            for file in cam_dir.glob(f"realtime_metrics_{target_month}-*.csv"):
                date_str = file.stem.split("_")[-1]
                d = daily_map.setdefault(date_str, {"pass": 0, "max": 0, "long": 0})
                metric_df = self._safe_read(file)
                pass_df = self._safe_read(cam_dir / f"pass_events_{date_str}.csv")
                long_df = self._safe_read(cam_dir / f"long_stay_events_{date_str}.csv")
                d["pass"] += len(pass_df)
                if not metric_df.empty:
                    d["max"] = max(d["max"], float(metric_df["congestion_score"].max()))
                d["long"] += len(long_df)

        for key in sorted(daily_map.keys()):
            d = daily_map[key]
            ws.append([key, int(d["pass"]), round(d["max"], 1), int(d["long"])])

        out_path = self.output_root / "reports" / "monthly" / f"monthly_report_{target_month}.xlsx"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def _add_camera_sheet(self, wb: Workbook, cam: dict, target_date: date, pass_df: pd.DataFrame, metric_df: pd.DataFrame, long_df: pd.DataFrame) -> None:
        ws = wb.create_sheet(f"cam{cam['camera_id']}")
        self._style_sheet(ws)
        ws["A1"] = f"{cam['camera_name']} Detail"
        ws["A1"].font = Font(bold=True, size=14, color="00D7FF")

        hist = [0] * 144
        if not pass_df.empty:
            pass_df["timestamp"] = pd.to_datetime(pass_df["timestamp"])
            for ts in pass_df["timestamp"]:
                idx = (ts.hour * 60 + ts.minute) // 10
                hist[idx] += 1

        ws.append(["bin", "pass_count"])
        for i, v in enumerate(hist):
            ws.append([i, v])

        bar = BarChart()
        bar.title = "Pass Histogram"
        data = Reference(ws, min_col=2, min_row=3, max_row=146)
        cats = Reference(ws, min_col=1, min_row=3, max_row=146)
        bar.add_data(data, titles_from_data=False)
        bar.set_categories(cats)
        bar.height = 6
        bar.width = 13
        ws.add_chart(bar, "D3")

        start_row = 150
        ws[f"A{start_row}"] = "congestion_time_series"
        ws.append(["timestamp", "congestion_score"])
        for _, r in metric_df[["timestamp", "congestion_score"]].iterrows() if not metric_df.empty else []:
            ws.append([r["timestamp"], float(r["congestion_score"])])

        if len(metric_df) > 2:
            line = LineChart()
            line.title = "Congestion Score"
            dref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + len(metric_df))
            cref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(metric_df))
            line.add_data(dref, titles_from_data=False)
            line.set_categories(cref)
            line.height = 5
            line.width = 13
            ws.add_chart(line, "D18")

        ls_row = 18
        ws[f"A{ls_row}"] = "long_stay_list"
        ws[f"A{ls_row+1}"] = "track_id"
        ws[f"B{ls_row+1}"] = "stay_minutes"
        for i, (_, r) in enumerate(long_df.iterrows() if not long_df.empty else []):
            ws[f"A{ls_row+2+i}"] = int(r["track_id"])
            ws[f"B{ls_row+2+i}"] = float(r["stay_minutes"])

    def _style_sheet(self, ws):
        fill = PatternFill("solid", fgColor="101922")
        thin = Side(style="thin", color="1A9FB6")
        for col in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 22
        for row in range(1, 220):
            for col in range(1, 8):
                c = ws.cell(row=row, column=col)
                c.fill = fill
                c.font = Font(color="FFFFFF", size=10)
                c.alignment = Alignment(vertical="center")
                c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    @staticmethod
    def _safe_read(path: Path) -> pd.DataFrame:
        if not path.exists():
            return pd.DataFrame()
        try:
            return pd.read_csv(path)
        except Exception:
            return pd.DataFrame()
